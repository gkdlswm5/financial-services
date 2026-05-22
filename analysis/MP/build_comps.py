"""
Build MP-Comps-Analysis.xlsx — institutional-grade comparable company analysis
following the `comps-analysis` skill in plugins/vertical-plugins/financial-analysis.

Conventions enforced:
  - Blue font = hardcoded input (every input has a cell comment citing source)
  - Black font = formula
  - Stats block (Max / 75th / Median / 25th / Min) on every comparable column
  - Times New Roman, navy/light-blue/grey palette only
  - REE-specific tab: NdPr capacity and EV per MT/yr NdPr

Run:
    python3 build_comps.py
Output:
    MP-Comps-Analysis.xlsx (sibling of this script)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "MP-Comps-Analysis.xlsx"

NAVY = "17365D"
LIGHTBLUE = "D9E1F2"
LIGHTGREY = "F2F2F2"
WHITE = "FFFFFF"

BLUE_FONT = Font(name="Times New Roman", size=11, color="0000FF")   # hardcoded input
BLACK_FONT = Font(name="Times New Roman", size=11, color="000000")  # formula
BOLD_WHITE = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLD_BLACK = Font(name="Times New Roman", size=12, bold=True)
ITAL_GREY = Font(name="Times New Roman", size=10, italic=True, color="595959")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
LIGHTBLUE_FILL = PatternFill("solid", fgColor=LIGHTBLUE)
LIGHTGREY_FILL = PatternFill("solid", fgColor=LIGHTGREY)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

PEERS = [
    # (name, ticker)
    ("MP Materials",                "NYSE: MP"),
    ("Lynas Rare Earths",           "ASX: LYC"),
    ("Energy Fuels",                "NYSE: UUUU"),
    ("USA Rare Earth",              "Nasdaq: USAR"),
    ("China Northern Rare Earth",   "SHA: 600111"),
    ("Albemarle",                   "NYSE: ALB"),
]

# Raw inputs — all dollar figures in $ millions unless noted.
# Sources cited via cell comments. [E] = estimate.
# Format: each peer maps to a dict of (value, source_or_assumption)
INPUTS = {
    "MP Materials": {
        "mkt_cap":         (9810,   "StockAnalysis.com – MP Statistics, accessed May 2026 ($9.81B)"),
        "net_debt":        (200,    "[E] Net debt ~$200M post-DoD preferred; from Q1'26 10-Q balance sheet (mp-20260331)"),
        "revenue_ttm":     (275,    "MP Q1 2026 10-Q + prior 3 quarters; TTM ~$275M"),
        "rev_growth":      (0.49,   "Q1 2026 YoY revenue growth, per MP Q1 2026 earnings release"),
        "gross_profit":    (83,     "[E] ~30% gross margin × $275M TTM rev (per Q1 release commentary)"),
        "ebitda_ttm":      (-53,    "TTM EBITDA per StockAnalysis.com / company filings (negative; inflecting positive in Q1'26 Adj)"),
        "net_income_ttm":  (-86,    "TTM net loss per StockAnalysis.com (-$85.87M)"),
        "ndpr_capacity":   (6000,   "Near-term NdPr oxide capacity (MT/yr), per Stage II ramp commentary"),
    },
    "Lynas Rare Earths": {
        "mkt_cap":         (10800,  "Yahoo Finance LYC.AX / StockAnalysis, May 2026 (~A$15.1B → US$10.8B)"),
        "net_debt":        (0,      "Approx net cash neutral; per H1 FY26 results (Dec 2025)"),
        "revenue_ttm":     (578,    "LTM revenue US$578M per StockAnalysis.com, Jan 2026"),
        "rev_growth":      (1.15,   "Q3 FY26 revenue +115% YoY per stocksdownunder.com"),
        "gross_profit":    (289,    "[E] ~50% gross margin × $578M; consistent with reported NdPr selling prices"),
        "ebitda_ttm":      (211,    "LTM EBITDA US$211M per StockAnalysis.com"),
        "net_income_ttm":  (80,     "H1 FY26 net income A$80.2M ≈ US$53M (annualized US$106M); using H1 as TTM proxy ≈ $80M"),
        "ndpr_capacity":   (8800,   "FY26 NdPr oxide volume forecast 8,800 MT (+35% YoY) per S&P Global, Dec 2025"),
    },
    "Energy Fuels": {
        "mkt_cap":         (1500,   "[E] Approx mkt cap May 2026 (uranium+REE mid-cap)"),
        "net_debt":        (-50,    "[E] Net cash position; from UUUU Q1'26 press release"),
        "revenue_ttm":     (70,     "[E] TTM revenue ~$70M (primarily uranium; REE pre-commercial)"),
        "rev_growth":      (None,   "Mixed seg.; NM — uranium-driven, REE segment pre-rev"),
        "gross_profit":    (None,   "NM at current scale"),
        "ebitda_ttm":      (-30,    "[E] Negative TTM EBITDA on capex spend"),
        "net_income_ttm":  (-50,    "[E] Net loss TTM"),
        "ndpr_capacity":   (6000,   "Planned NdPr capacity per Phase 2 BFS, Jan 2026 (6,000 tpa NdPr + 240 Dy + 66 Tb)"),
    },
    "USA Rare Earth": {
        "mkt_cap":         (1000,   "[E] At $24.39 stock price May 2026 × shares post-PIPE"),
        "net_debt":        (-1500,  "Net cash; $1.5B PIPE closed (Inflection Point anchor)"),
        "revenue_ttm":     (5,      "[E] Pre-commercial; magnet line commercial Q2'26"),
        "rev_growth":      (None,   "Pre-revenue scale"),
        "gross_profit":    (None,   "NM"),
        "ebitda_ttm":      (-20,    "[E] Operating loss on build phase"),
        "net_income_ttm":  (-30,    "[E] Net loss"),
        "ndpr_capacity":   (1200,   "Magnet line nameplate (sintered NdFeB), Stillwater OK; no oxide capacity"),
    },
    "China Northern Rare Earth": {
        "mkt_cap":         (11500,  "[E] Approx US$11.5B mkt cap May 2026 (SHA: 600111 at recent prices)"),
        "net_debt":        (1000,   "[E] Modest net debt; state-controlled balance sheet"),
        "revenue_ttm":     (5910,   "FY25 revenue ¥42.563B (+29% YoY) per SMM; ÷7.2 = US$5,910M"),
        "rev_growth":      (0.29,   "FY25 revenue growth per SMM"),
        "gross_profit":    (1182,   "[E] ~20% gross margin × $5,910M (per industry norms for integrated REE producer)"),
        "ebitda_ttm":      (700,    "[E] ~12% EBITDA margin × $5,910M; consistent with $313M net income + D&A + tax"),
        "net_income_ttm":  (313,    "FY25 NI ¥2.251B (+124% YoY) per SMM; ÷7.2 = US$313M"),
        "ndpr_capacity":   (60000,  "[E] Industry-leading position; >60% global REE supply per Discovery Alert / Farmonaut"),
    },
    "Albemarle": {
        "mkt_cap":         (22700,  "[E] At $193.58 stock price (Apr 2026 close) × ~117M shares"),
        "net_debt":        (2300,   "[E] Net debt position; from recent 10-Q"),
        "revenue_ttm":     (5850,   "FY26 guidance ~$5.85B per Q1'26 release"),
        "rev_growth":      (0.33,   "Q1'26 revenue +32.67% YoY"),
        "gross_profit":    (1463,   "[E] ~25% gross margin × $5,850M"),
        "ebitda_ttm":      (2500,   "FY26 EBITDA guidance ~$2.5B per Q1'26 release"),
        "net_income_ttm":  (200,    "[E] Recovering lithium prices; modest profitability"),
        "ndpr_capacity":   (0,      "N/A — lithium / bromine producer, not REE"),
    },
}

# =========================================================
# WORKBOOK
# =========================================================
wb = Workbook()
wb.remove(wb.active)

# ---------- Helpers ----------
def section_header(ws, row, start_col, end_col, text):
    ws.cell(row=row, column=start_col, value=text)
    rng = f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
    ws.merge_cells(rng)
    c = ws.cell(row=row, column=start_col)
    c.font = BOLD_WHITE
    c.fill = NAVY_FILL
    c.alignment = CENTER

def column_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD_BLACK
    c.fill = LIGHTBLUE_FILL
    c.alignment = CENTER

def stat_row(ws, row, label_col, label, n_cols, data_first_row, data_last_row, percent_cols=None, mult_cols=None):
    c = ws.cell(row=row, column=label_col, value=label)
    c.font = BOLD_BLACK
    c.fill = LIGHTGREY_FILL
    c.alignment = LEFT
    for col in range(label_col + 1, label_col + n_cols + 1):
        col_letter = get_column_letter(col)
        rng = f"{col_letter}{data_first_row}:{col_letter}{data_last_row}"
        if label == "Maximum":
            formula = f"=MAX({rng})"
        elif label == "75th Percentile":
            formula = f"=QUARTILE({rng},3)"
        elif label == "Median":
            formula = f"=MEDIAN({rng})"
        elif label == "25th Percentile":
            formula = f"=QUARTILE({rng},1)"
        elif label == "Minimum":
            formula = f"=MIN({rng})"
        cell = ws.cell(row=row, column=col, value=formula)
        cell.fill = LIGHTGREY_FILL
        cell.font = BOLD_BLACK
        cell.alignment = CENTER
        if percent_cols and col in percent_cols:
            cell.number_format = "0.0%"
        elif mult_cols and col in mult_cols:
            cell.number_format = "0.0\"x\""
        else:
            cell.number_format = "#,##0"

def write_input(ws, row, col, value, source, percent=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLUE_FONT
    cell.alignment = CENTER
    if value is None:
        cell.value = "N/A"
        cell.font = ITAL_GREY
    else:
        if percent:
            cell.number_format = "0.0%"
        else:
            cell.number_format = "#,##0"
    cell.comment = Comment(source, "MP comp analysis")

def write_formula(ws, row, col, formula, fmt="#,##0"):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = BLACK_FONT
    cell.alignment = CENTER
    cell.number_format = fmt

# =========================================================
# Tab 1: INPUTS
# =========================================================
ws = wb.create_sheet("Inputs")
ws.sheet_view.showGridLines = False

# Title block (rows 1-3 per skill)
ws["A1"] = "RARE EARTHS & CRITICAL MINERALS — COMPARABLE COMPANY ANALYSIS"
ws.merge_cells("A1:K1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER

ws["A2"] = " • ".join(f"{n} ({t})" for n,t in PEERS)
ws.merge_cells("A2:K2"); ws["A2"].font = BOLD_BLACK; ws["A2"].alignment = CENTER

ws["A3"] = "As of May 2026 | All figures in USD Millions except ratios and capacity (MT/yr)"
ws.merge_cells("A3:K3"); ws["A3"].font = ITAL_GREY; ws["A3"].alignment = CENTER

# Section header
section_header(ws, 5, 1, 11, "RAW INPUTS — cell comments cite source / assumption")

# Column headers
headers = ["Company", "Ticker", "Mkt Cap", "Net Debt", "Revenue (TTM)",
           "Rev Growth %", "Gross Profit", "EBITDA (TTM)", "Net Income (TTM)",
           "NdPr Cap (MT/yr)", "Notes"]
for i, h in enumerate(headers, start=1):
    column_header(ws, 6, i, h)

# Data rows
for idx, (name, ticker) in enumerate(PEERS):
    r = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    ws.cell(row=r, column=2, value=ticker).font = BLACK_FONT
    d = INPUTS[name]
    write_input(ws, r, 3, d["mkt_cap"][0],          d["mkt_cap"][1])
    write_input(ws, r, 4, d["net_debt"][0],         d["net_debt"][1])
    write_input(ws, r, 5, d["revenue_ttm"][0],      d["revenue_ttm"][1])
    write_input(ws, r, 6, d["rev_growth"][0],       d["rev_growth"][1], percent=True)
    write_input(ws, r, 7, d["gross_profit"][0],     d["gross_profit"][1])
    write_input(ws, r, 8, d["ebitda_ttm"][0],       d["ebitda_ttm"][1])
    write_input(ws, r, 9, d["net_income_ttm"][0],   d["net_income_ttm"][1])
    write_input(ws, r,10, d["ndpr_capacity"][0],    d["ndpr_capacity"][1])
    notes = ""
    if any("[E]" in str(d[k][1]) for k in d):
        notes = "Contains [E] estimates — see cell comments"
    ws.cell(row=r, column=11, value=notes).font = ITAL_GREY

# Column widths
widths = [26, 14, 12, 12, 14, 13, 14, 14, 16, 16, 36]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 2: OPERATING METRICS
# =========================================================
ws = wb.create_sheet("Operating Metrics")
ws.sheet_view.showGridLines = False
ws["A1"] = "OPERATING STATISTICS & FINANCIAL METRICS"
ws.merge_cells("A1:G1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER
ws["A2"] = "All figures USD Millions except margins/growth. Formulas reference Inputs tab."
ws.merge_cells("A2:G2"); ws["A2"].font = ITAL_GREY; ws["A2"].alignment = CENTER

op_headers = ["Company", "Revenue (TTM)", "Rev Growth (YoY)", "Gross Profit", "Gross Margin", "EBITDA (TTM)", "EBITDA Margin"]
for i, h in enumerate(op_headers, start=1):
    column_header(ws, 4, i, h)

# Data — formulas reference Inputs!
for idx, (name, _) in enumerate(PEERS):
    r = 5 + idx
    src = 7 + idx   # corresponding inputs row
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    write_formula(ws, r, 2, f"=Inputs!E{src}")                          # revenue
    write_formula(ws, r, 3, f"=IFERROR(Inputs!F{src},\"N/A\")", "0.0%") # rev growth
    write_formula(ws, r, 4, f"=IFERROR(Inputs!G{src},\"N/A\")")         # gross profit
    write_formula(ws, r, 5, f"=IFERROR(Inputs!G{src}/Inputs!E{src},\"N/A\")", "0.0%")  # gross margin
    write_formula(ws, r, 6, f"=Inputs!H{src}")                          # ebitda
    write_formula(ws, r, 7, f"=IFERROR(Inputs!H{src}/Inputs!E{src},\"N/A\")", "0.0%")  # ebitda margin

# Blank row, then stats — only on comparable columns (3=growth, 5=GM, 7=EBITDA margin)
DATA_FIRST = 5
DATA_LAST  = 5 + len(PEERS) - 1
stat_first = DATA_LAST + 2
labels = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]
percent_cols = {3, 5, 7}
for i, lab in enumerate(labels):
    stat_row(ws, stat_first + i, 1, lab, 6, DATA_FIRST, DATA_LAST, percent_cols=percent_cols)
    # clear non-comparable size cols (2, 4, 6 — absolute $)
    for col in (2, 4, 6):
        cell = ws.cell(row=stat_first + i, column=col, value="")
        cell.fill = LIGHTGREY_FILL

widths = [26, 14, 16, 14, 14, 14, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 3: VALUATION
# =========================================================
ws = wb.create_sheet("Valuation")
ws.sheet_view.showGridLines = False
ws["A1"] = "VALUATION MULTIPLES"
ws.merge_cells("A1:G1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER
ws["A2"] = "EV = Mkt Cap + Net Debt. NM (\"not meaningful\") shown for negative denominators."
ws.merge_cells("A2:G2"); ws["A2"].font = ITAL_GREY; ws["A2"].alignment = CENTER

val_headers = ["Company", "Mkt Cap", "EV", "EV / Rev", "EV / EBITDA", "P / E (TTM)", "Comment"]
for i, h in enumerate(val_headers, start=1):
    column_header(ws, 4, i, h)

for idx, (name, _) in enumerate(PEERS):
    r = 5 + idx
    src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    write_formula(ws, r, 2, f"=Inputs!C{src}")                                       # mkt cap
    write_formula(ws, r, 3, f"=Inputs!C{src}+Inputs!D{src}")                         # EV
    write_formula(ws, r, 4, f"=IFERROR((Inputs!C{src}+Inputs!D{src})/Inputs!E{src},\"NM\")", "0.0\"x\"")  # EV/Rev
    write_formula(ws, r, 5, f"=IF(Inputs!H{src}>0,(Inputs!C{src}+Inputs!D{src})/Inputs!H{src},\"NM\")", "0.0\"x\"")  # EV/EBITDA
    write_formula(ws, r, 6, f"=IF(Inputs!I{src}>0,Inputs!C{src}/Inputs!I{src},\"NM\")", "0.0\"x\"")        # P/E
    comment = ""
    if name == "MP Materials":
        comment = "Negative TTM EBITDA → NM; richly priced on revenue; Q1'26 Adj EBITDA inflecting +"
    elif name == "Lynas Rare Earths":
        comment = "Most expensive Western on EV/EBITDA; cheaper than MP on EV/Rev"
    elif name == "Energy Fuels":
        comment = "Pre-meaningful-scale; valuation lens is capacity (see REE-Specific tab)"
    elif name == "USA Rare Earth":
        comment = "Pre-rev magnet pure-play; valuation = optionality on US magnet supply"
    elif name == "China Northern Rare Earth":
        comment = "Cheapest on EV/Rev; un-investable for many Western mandates due to geopolitical risk"
    elif name == "Albemarle":
        comment = "Adjacent commodity (lithium, not REE); included as critical-minerals context"
    cc = ws.cell(row=r, column=7, value=comment); cc.font = ITAL_GREY; cc.alignment = LEFT

# stats on comparable cols only (4 EV/Rev, 5 EV/EBITDA, 6 P/E)
for i, lab in enumerate(labels):
    stat_row(ws, stat_first + i, 1, lab, 6, DATA_FIRST, DATA_LAST, mult_cols={4,5,6})
    for col in (2, 3, 7):
        cell = ws.cell(row=stat_first + i, column=col, value="")
        cell.fill = LIGHTGREY_FILL

widths = [26, 12, 12, 12, 14, 12, 50]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 4: REE-SPECIFIC
# =========================================================
ws = wb.create_sheet("REE-Specific")
ws.sheet_view.showGridLines = False
ws["A1"] = "RARE-EARTH-SPECIFIC METRICS — Capacity-Based Valuation"
ws.merge_cells("A1:F1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER
ws["A2"] = "When EBITDA is negative or scaling, capacity multiples are the cleaner comparable."
ws.merge_cells("A2:F2"); ws["A2"].font = ITAL_GREY; ws["A2"].alignment = CENTER

ree_headers = ["Company", "EV ($M)", "NdPr Capacity (MT/yr)", "EV per MT/yr ($M)", "Stage Reached", "Comment"]
for i, h in enumerate(ree_headers, start=1):
    column_header(ws, 4, i, h)

stages = {
    "MP Materials": "Mine → Sep + Stage III magnet under build (2028)",
    "Lynas Rare Earths": "Mine → Sep + Heavy REE (US Seadrift, 2026)",
    "Energy Fuels": "Processing — REE Phase 2 in BFS, FID Q4'26",
    "USA Rare Earth": "Magnet only (Stillwater OK, Q2'26 commercial)",
    "China Northern Rare Earth": "Mine → Magnet (fully integrated, state-owned)",
    "Albemarle": "N/A — lithium / bromine producer",
}

for idx, (name, _) in enumerate(PEERS):
    r = 5 + idx
    src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    write_formula(ws, r, 2, f"=Inputs!C{src}+Inputs!D{src}")              # EV
    write_formula(ws, r, 3, f"=Inputs!J{src}")                            # capacity
    write_formula(ws, r, 4, f"=IFERROR((Inputs!C{src}+Inputs!D{src})/Inputs!J{src},\"N/A\")", "#,##0.00")  # EV/cap
    sc = ws.cell(row=r, column=5, value=stages[name]); sc.font = BLACK_FONT; sc.alignment = LEFT
    cmt = ""
    if name == "MP Materials":
        cmt = "Highest $/capacity in Western set — reflects DoD floor + magnet optionality embedded"
    elif name == "Lynas Rare Earths":
        cmt = "Cheaper on capacity than MP; biggest non-China NdPr volume"
    elif name == "China Northern Rare Earth":
        cmt = "Lowest $/capacity globally — scale advantage of incumbent"
    elif name == "Albemarle":
        cmt = "N/A — different commodity"
    cc = ws.cell(row=r, column=6, value=cmt); cc.font = ITAL_GREY; cc.alignment = LEFT

# stats — only on EV/capacity col (4)
for i, lab in enumerate(labels):
    stat_row(ws, stat_first + i, 1, lab, 5, DATA_FIRST, DATA_LAST)
    for col in (2, 3, 5, 6):
        cell = ws.cell(row=stat_first + i, column=col, value="")
        cell.fill = LIGHTGREY_FILL
    # clear and reformat EV/cap col 4 as decimal
    cell = ws.cell(row=stat_first + i, column=4)
    cell.number_format = "#,##0.00"

widths = [26, 12, 20, 18, 50, 50]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 5: NOTES
# =========================================================
ws = wb.create_sheet("Notes")
ws.sheet_view.showGridLines = False
ws["A1"] = "NOTES, METHODOLOGY & DATA-SOURCE CAVEATS"
ws.merge_cells("A1:E1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER

notes_blocks = [
    ("Data sources", [
        "MCP data sources (S&P Kensho / CapIQ, FactSet, Daloopa) WERE NOT USED — the FSI plugin connectors were not configured in this environment.",
        "Primary public sources: SEC EDGAR (10-Q FY2026 Q1 for MP, USAR), MP investor relations, Lynas reporting centre (H1 FY26, Q3 FY26),",
        "StockAnalysis.com & Yahoo Finance for market data, SMM / Discovery Alert for China Northern Rare Earth FY25 results.",
        "Where required figures were not directly disclosed, an [E] estimate is shown with the assumption documented in the cell comment.",
    ]),
    ("Period definitions", [
        "Revenue TTM: rolling four quarters where available; for MP, sum of Q2-Q4 2025 + Q1 2026.",
        "For Lynas, US$ TTM derived from StockAnalysis (US$578M).",
        "For Chinese peers, FY2025 reported figures used (calendar year) — CNY converted at 7.2 CNY/USD.",
        "USAR is pre-commercial revenue — figures are placeholder.",
    ]),
    ("Definitions", [
        "EV = Market Cap + Net Debt (where Net Debt = Total Debt − Cash). Negative Net Debt = net cash position.",
        "EBITDA: per-company reported, including the negative TTM figure for MP (Adj. EBITDA flipped positive in Q1 2026 +$36.6M).",
        "NdPr Capacity: nameplate or near-term oxide production capacity in MT/yr. For USAR, magnet capacity (no oxide).",
        "[NM] (not meaningful): multiple denominator is negative or zero — multiple is suppressed.",
    ]),
    ("Industry-specific guidance", [
        "Standard EV/EBITDA, EV/Sales, and P/E are limited for this peer set due to negative or pre-scale earnings at multiple comps.",
        "EV per MT/yr NdPr capacity (REE-Specific tab) is the cleaner cross-peer comparable for this industry.",
        "The DoD price-protection agreement creates a structural asymmetry that no other peer enjoys — quantitative",
        "multiples understate MP's downside protection vs. peers exposed to NdPr spot volatility.",
    ]),
    ("Caveats & limitations", [
        "1. Cell comments distinguish sourced figures from [E] estimates — review before using in a binding decision.",
        "2. Chinese major valuation should be triangulated against a CNY-denominated reference; FX assumption is point-in-time.",
        "3. Junior peers (UUUU, USAR) are valuation outliers — capacity-based multiples are more informative than P&L-based.",
        "4. This analysis is research, not investment advice. Cross-check against a primary terminal (Bloomberg, CapIQ, FactSet) before action.",
    ]),
    ("Cross-references", [
        "Companion file: MP-Competitive-Analysis.pptx (slide deck per `competitive-analysis` skill).",
        "Build script: build_comps.py — re-run to refresh after updating INPUTS dict.",
    ]),
]

r = 3
for title, lines in notes_blocks:
    c = ws.cell(row=r, column=1, value=title); c.font = BOLD_BLACK; c.fill = LIGHTBLUE_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    for line in lines:
        cc = ws.cell(row=r, column=1, value=line); cc.font = BLACK_FONT; cc.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Reorder sheets
wb._sheets = [wb["Inputs"], wb["Operating Metrics"], wb["Valuation"], wb["REE-Specific"], wb["Notes"]]

wb.save(OUT)
print(f"Wrote {OUT}")
