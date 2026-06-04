"""
Build RKLB-Comps-Analysis.xlsx — comparable company analysis for Rocket Lab.
Small-cap pre-profit / bifurcated growth structure.
Peer set: ASTS, PL, IRDM, RDW (main); LMT (defense anchor); SpaceX (private memo).
EV/Revenue is the cleaner gauge — most peers have negative or no EBITDA.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "RKLB-Comps-Analysis.xlsx"
NAVY = "17365D"; LB = "D9E1F2"; LG = "F2F2F2"
BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDB = Font(name="Times New Roman", size=12, bold=True)
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")
NAVY_FILL = PatternFill("solid", fgColor=NAVY); LB_FILL = PatternFill("solid", fgColor=LB); LG_FILL = PatternFill("solid", fgColor=LG)
C = Alignment("center", "center"); L = Alignment("left", "center"); WRAP = Alignment("left", "top", wrap_text=True)

PEERS = [
    ("Rocket Lab USA",     "RKLB",  "target"),
    ("AST SpaceMobile",    "ASTS",  "pure"),
    ("Planet Labs",        "PL",    "pure"),
    ("Iridium",            "IRDM",  "pure"),
    ("Redwire",            "RDW",   "pure"),
    ("Lockheed Martin",    "LMT",   "anchor"),
    ("SpaceX (private)",   "—",     "private"),
]

INPUTS = {
    "Rocket Lab USA": {
        "mkt_cap": (74974, "RKLB ~$123.32/sh x ~608M sh = ~$75.0B (June 2, 2026)"),
        "net_debt": (-1442, "Net cash $1.44B: cash+sec $1.48B - debt $38M"),
        "revenue": (679, "[E] TTM revenue ~$679M (FY25 $602M + Q1 26 $200M - Q1 25 $123M)"),
        "growth": (0.635, "Q1 CY26 revenue +63.5% YoY (Rocket Lab IR, May 7, 2026)"),
        "gm": (0.382, "Q1 26 GAAP gross margin 38.2%"),
        "ebitda": (-50, "[E] TTM adj EBITDA ~-$50M (Q1 26 adj EBITDA -$11.8M; narrowing losses)"),
        "ni": (-220, "[E] TTM net loss ~$220M (FY25 NI -$198M + Q1 26 -$45M - Q1 25 ~-$25M)"),
    },
    "AST SpaceMobile": {
        "mkt_cap": (35330, "ASTS ~$110.47/sh x ~320M sh = ~$35.3B (June 3, 2026)"),
        "net_debt": (-500, "Net cash ~$500M: $3.5B liquidity - debt $2.97B (mostly converts)"),
        "revenue": (85, "TTM revenue ~$85M (FY25 $70.9M + Q1 26 $14.7M)"),
        "growth": (20.0, "Q1 26 revenue +2000%+ YoY (off ~$0.7M base; reaffirmed FY26 guide $150-200M)"),
        "gm": (0.69, "[E] Gross margin ~69% on FY25 (milestone-driven)"),
        "ebitda": (-200, "[E] Adj EBITDA loss ~$200M (R&D + satellite buildout)"),
        "ni": (-600, "[E] TTM net loss ~$600M (Q1 26 alone -$191M)"),
    },
    "Planet Labs": {
        "mkt_cap": (15400, "PL ~$43/sh x ~358M sh = ~$15.4B (June 3, 2026)"),
        "net_debt": (-178, "Net cash $178M: cash+sec $640M - debt $462M"),
        "revenue": (308, "TTM revenue $307.7M (+26% YoY; FY26 ended Jan 2026)"),
        "growth": (0.41, "Q4 FY26 revenue +41% YoY (subscription growth + large gov't deals)"),
        "gm": (0.56, "FY26 GAAP gross margin 56% (non-GAAP 59%)"),
        "ebitda": (15.5, "FY26 adj EBITDA $15.5M (first profitable adj EBITDA year)"),
        "ni": (-247, "TTM NI -$247M (includes $161M warrant revaluation loss)"),
    },
    "Iridium": {
        "mkt_cap": (5250, "IRDM ~$45/sh x ~117M sh = ~$5.25B (range $5.0-5.5B)"),
        "net_debt": (1700, "Net debt $1.7B: debt $1.8B - cash $112M"),
        "revenue": (887, "TTM revenue ~$887M (FY25 $871.7M + Q1 26 step-up)"),
        "growth": (0.02, "Q1 CY26 service revenue +2% YoY (mature satcom)"),
        "gm": (0.74, "[E] Service gross margin ~74%"),
        "ebitda": (474, "TTM operational EBITDA ~$474M (~54% margin); FY26 guide $480-490M"),
        "ni": (105, "[E] TTM NI ~$105M"),
    },
    "Redwire": {
        "mkt_cap": (4070, "RDW ~$18.44/sh x ~221M sh = ~$4.07B"),
        "net_debt": (-55, "Net cash $55M: cash $145M - debt ~$90M"),
        "revenue": (432, "[E] TTM revenue ~$432M (post-Edge Autonomy contribution)"),
        "growth": (0.579, "Q1 26 revenue +57.9% YoY"),
        "gm": (0.266, "Q1 26 gross margin 26.6% (up from 14.7% YoY)"),
        "ebitda": (-40, "[E] TTM adj EBITDA loss ~-$40M (FY25 -$50.3M)"),
        "ni": (-227, "[E] TTM net loss ~-$227M (FY25 -$226.6M)"),
    },
    "Lockheed Martin": {
        "mkt_cap": (121000, "LMT ~$510/sh x ~237M sh = ~$121B"),
        "net_debt": (18800, "Net debt $18.8B (debt $20.7B - cash $1.89B)"),
        "revenue": (75100, "TTM revenue ~$75.1B"),
        "growth": (0.0, "Q1 26 revenue ~flat YoY"),
        "gm": (0.11, "[E] Op margin ~11% (consolidated); Space segment ~10%"),
        "ebitda": (10000, "[E] TTM EBITDA ~$10B"),
        "ni": (5500, "[E] TTM NI ~$5.5B"),
    },
    "SpaceX (private)": {
        "mkt_cap": (1750000, "SpaceX IPO target ~$1.75-2T (April 2026 confidential S-1; Bloomberg May)"),
        "net_debt": (0, "[E] SpaceX balance sheet not disclosed; assume net cash neutral"),
        "revenue": (18674, "FY25 revenue $18.674B (per S-1; +43% YoY); 2026E ~$22-24B [E]"),
        "growth": (0.43, "FY25 revenue +43% YoY ($18.7B vs $13.1B); Starlink $11.4B (61% of total)"),
        "gm": (0.40, "[E] Gross margin ~40% (Starlink scaling)"),
        "ebitda": (6584, "FY25 adj EBITDA $6.584B (~35% margin)"),
        "ni": (-2589, "FY25 operating loss -$2.589B (per S-1)"),
    },
}

wb = Workbook(); wb.remove(wb.active)

def section(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text); ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0); cell.font = BOLDW; cell.fill = NAVY_FILL; cell.alignment = C

def colhead(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text); cell.font = BOLDB; cell.fill = LB_FILL; cell.alignment = C

def put_input(ws, r, c, val, src, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=val); cell.font = BLUE; cell.alignment = C; cell.number_format = fmt
    cell.comment = Comment(src, "RKLB comps")

def put_f(ws, r, c, formula, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=formula); cell.font = BLACK; cell.alignment = C; cell.number_format = fmt

# Tab 1 — INPUTS
ws = wb.create_sheet("Inputs"); ws.sheet_view.showGridLines = False
ws["A1"] = "SPACE / SAT-COMMS — COMPARABLE COMPANY ANALYSIS (RKLB target)"
ws.merge_cells("A1:K1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = " • ".join(f"{n} ({t})" for n, t, _ in PEERS)
ws.merge_cells("A2:K2"); ws["A2"].font = BOLDB; ws["A2"].alignment = C
ws["A3"] = "As of June 3, 2026 | $M except ratios | Pre-profit cohort: EV/Revenue is the cleaner gauge | LMT = old-space anchor; SpaceX = private benchmark (memo)"
ws.merge_cells("A3:K3"); ws["A3"].font = ITAL; ws["A3"].alignment = C

section(ws, 5, 1, 11, "RAW INPUTS — cell comments cite source / assumption")
heads = ["Company", "Ticker", "Mkt Cap", "Net Debt", "Revenue (TTM)", "Rev Growth %",
         "Gross Margin %", "EBITDA (TTM)", "Net Income (TTM)", "Role", "Notes"]
for i, h in enumerate(heads, start=1): colhead(ws, 6, i, h)
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; d = INPUTS[name]
    cn = ws.cell(row=r, column=1, value=name); cn.font = BOLDB
    ws.cell(row=r, column=2, value=tic).font = BLACK
    put_input(ws, r, 3, d["mkt_cap"][0], d["mkt_cap"][1])
    put_input(ws, r, 4, d["net_debt"][0], d["net_debt"][1])
    put_input(ws, r, 5, d["revenue"][0], d["revenue"][1])
    put_input(ws, r, 6, d["growth"][0], d["growth"][1], "0.0%")
    put_input(ws, r, 7, d["gm"][0], d["gm"][1], "0.0%")
    put_input(ws, r, 8, d["ebitda"][0], d["ebitda"][1])
    put_input(ws, r, 9, d["ni"][0], d["ni"][1])
    tag = {"target": "TARGET", "pure": "PUBLIC PEER", "anchor": "OLD-SPACE ANCHOR", "private": "PRIVATE (memo)"}[kind]
    ws.cell(row=r, column=10, value=tag).font = ITAL
for i, w in enumerate([22, 9, 13, 11, 14, 13, 14, 14, 16, 18, 22], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Tab 2 — VALUATION (EV/Rev focus)
ws = wb.create_sheet("Valuation"); ws.sheet_view.showGridLines = False
ws["A1"] = "VALUATION — EV/REVENUE THE CLEANER GAUGE FOR THIS COHORT"
ws.merge_cells("A1:H1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "EV = Mkt Cap + Net Debt (negative net debt = net cash). EV/EBITDA shown but mostly NM (negative or NM)."
ws.merge_cells("A2:H2"); ws["A2"].font = ITAL; ws["A2"].alignment = C
vh = ["Company", "Mkt Cap", "EV", "EV / Rev", "EV / EBITDA", "P/E (TTM)", "Comment"]
for i, h in enumerate(vh, start=1): colhead(ws, 6, i, h)
comments = {
    "Rocket Lab USA": "Target. ~110x EV/Rev — top of public cohort ex-ASTS. Bifurcated story; pre-profit but scaling fast.",
    "AST SpaceMobile": "~410x EV/Rev — pure pre-revenue story-stock; not useful as a multiple anchor.",
    "Planet Labs": "~49x EV/Rev — subscription Earth-obs; NDR 116%; first adj EBITDA+ year.",
    "Iridium": "~7-8x EV/Rev — the FCF/maturity-state anchor; profitable; cash returns.",
    "Redwire": "~9x EV/Rev — closest space-infra growth comp; pre-profit; Edge Autonomy reshape.",
    "Lockheed Martin": "~1.85x EV/Rev — old-space anchor; mature defense multiple; Space segment ~$13B/yr at ~10% margin.",
    "SpaceX (private)": "~95x EV/Rev at $1.75T IPO target (~$18.7B FY25 rev). Private mark. Set the upper bound of public valuation.",
}
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLDB
    put_f(ws, r, 2, f"=Inputs!C{src}")
    put_f(ws, r, 3, f"=Inputs!C{src}+Inputs!D{src}")
    put_f(ws, r, 4, f"=(Inputs!C{src}+Inputs!D{src})/Inputs!E{src}", '0.0"x"')
    put_f(ws, r, 5, f"=IF(Inputs!H{src}>0,(Inputs!C{src}+Inputs!D{src})/Inputs!H{src},\"NM\")", '0.0"x"')
    put_f(ws, r, 6, f"=IF(Inputs!I{src}>0,Inputs!C{src}/Inputs!I{src},\"NM\")", '0.0"x"')
    cc = ws.cell(row=r, column=7, value=comments[name]); cc.font = ITAL; cc.alignment = L
for i, w in enumerate([22, 13, 13, 12, 14, 12, 64], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Tab 3 — CAPACITY / OPERATING DETAIL
ws = wb.create_sheet("Capacity"); ws.sheet_view.showGridLines = False
ws["A1"] = "OPERATING CAPACITY — Launches/yr · Backlog · Pure-play status"
ws.merge_cells("A1:F1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A3"] = "Company"; ws["A4"] = "Rocket Lab"; ws["A5"] = "AST SpaceMobile"; ws["A6"] = "Planet Labs"
ws["A7"] = "Iridium"; ws["A8"] = "Redwire"; ws["A9"] = "Lockheed Martin"; ws["A10"] = "SpaceX"
ws["B3"] = "Launches / yr"; ws["B4"] = "21 (FY25)→ Neutron Q4 26"; ws["B5"] = "Sat ops (45 BlueBird YE26)"; ws["B6"] = "Subscription (no launches)"
ws["B7"] = "Sat ops only"; ws["B8"] = "Component supplier"; ws["B9"] = "10+ heavy launches (Atlas/Vulcan-adj)"; ws["B10"] = "~140 Falcon 9 + Starship"
ws["C3"] = "Backlog ($)"; ws["C4"] = "$2.22B (+108% YoY)"; ws["C5"] = "$1.2B+ commercial pipeline"; ws["C6"] = "$900M+ (+79% YoY)"
ws["C7"] = "n/a (subscription)"; ws["C8"] = "$498M (+71% YoY)"; ws["C9"] = "$186.4B company-wide"; ws["C10"] = "n/a (private)"
ws["D3"] = "Pure-play"; ws["D4"] = "Yes — bifurcated"; ws["D5"] = "Yes — direct-to-cell"; ws["D6"] = "Yes — EO subscription"
ws["D7"] = "Yes — satcom"; ws["D8"] = "Yes — space infra"; ws["D9"] = "No — Space is 17% of revenue"; ws["D10"] = "Yes — launch + Starlink"
for c_letter in "ABCD":
    cell = ws[f"{c_letter}3"]
    cell.font = BOLDB; cell.fill = LB_FILL; cell.alignment = L
for r in range(4, 11):
    ws[f"A{r}"].font = BOLDB
for col, w in zip("ABCD", [22, 28, 28, 38]):
    ws.column_dimensions[col].width = w

# Tab 4 — NOTES
ws = wb.create_sheet("Notes"); ws.sheet_view.showGridLines = False
ws["A1"] = "NOTES — small-cap-growth peer framing"
ws.merge_cells("A1:E1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
notes = [
    ("Peer framing", [
        "Pre-profit cohort: EV/Revenue is the cleaner multiple (most peers have negative or NM EBITDA).",
        "Public peers: ASTS (pre-rev option), PL (subscription EO, first adj EBITDA+), IRDM (mature satcom, cash returns), RDW (closest growth comp).",
        "LMT included as 'old-space' anchor — Space segment ~$13B/yr at ~10% op margin; sets the mature-state proxy.",
        "SpaceX as private memo — recently filed S-1 (April 2026, IPO target $1.75-2T per Bloomberg); FY25 revenue $18.7B.",
    ]),
    ("Key observations", [
        "RKLB EV/Rev ~110x vs PL 49x, RDW 9x, IRDM 7.5x, ASTS 410x — RKLB sits in the upper-growth zone.",
        "Analyst consensus PT $103.91 (S&P Global poll of 18 analysts) is ~16% BELOW spot $123.32 — Street view skeptical.",
        "RKLB has rocketed ~5-6x from $20-30 levels in early 2025; current price embeds heroic terminal assumptions.",
        "Companion: RKLB-Model.xlsx (DCF base $9.35 vs $123 market — 'too conservative' or 'right'? You decide).",
    ]),
    ("Data sources & caveats", [
        "RKLB Q1 CY26 10-Q (May 7, 2026); FY25 10-K (Feb 2026); IR press releases.",
        "Peer Q1 26 8-Ks/IR releases. Market data via StockAnalysis/Macrotrends/Public.com (~June 1-3, 2026).",
        "SpaceX figures from April 2026 S-1 disclosure + Bloomberg/Reuters/Satellite Today coverage.",
        "Mynaric acquisition (NOT MDA Space) closed April 14, 2026 — $155.3M cash + 2.28M RKLB shares.",
        "MCP terminals NOT configured. Research only, NOT investment advice.",
    ]),
]
row = 3
for title, lines in notes:
    c = ws.cell(row=row, column=1, value=title); c.font = BOLDB; c.fill = LB_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); row += 1
    for ln in lines:
        cc = ws.cell(row=row, column=1, value=ln); cc.font = BLACK; cc.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 16; row += 1
    row += 1
for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Inputs"], wb["Valuation"], wb["Capacity"], wb["Notes"]]
wb.save(OUT)
print(f"Wrote {OUT}")
