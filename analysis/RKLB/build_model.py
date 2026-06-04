"""
Build RKLB-Model.xlsx — 3-statement model + DCF for Rocket Lab USA.
Adapted for small-cap pre-profit / bifurcated growth structure.

Segments: Launch Services (Electron + Neutron) · Space Systems (incl. Mynaric/optical comms)
Period: FY2024A FY2025A FY2026E | FY2027E-FY2031E
Run: python3 build_model.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "RKLB-Model.xlsx"
NAVY = "17365D"; LB = "D9E1F2"; LG = "F2F2F2"; MB = "BDD7EE"
BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDN = Font(name="Times New Roman", size=11, bold=True, color="17365D")
BOLD = Font(name="Times New Roman", size=11, bold=True)
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")
NAVY_FILL = PatternFill("solid", fgColor=NAVY); LB_FILL = PatternFill("solid", fgColor=LB)
LG_FILL = PatternFill("solid", fgColor=LG); MB_FILL = PatternFill("solid", fgColor=MB)
C = Alignment("center", "center"); L = Alignment("left", "center"); WRAP = Alignment("left", "top", wrap_text=True)
NUM = "#,##0"; PCT = "0.0%"; USD2 = '"$"#,##0.00'

YR_COLS = ["C", "D", "E", "F", "G", "H", "I", "J"]
YR_HEADS = ["FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]

wb = Workbook(); wb.remove(wb.active)

def hdr(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text); ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0); cell.font = BOLDW; cell.fill = NAVY_FILL; cell.alignment = C

def yearrow(ws, row, label="($ in millions)"):
    ws.cell(row=row, column=1, value=label).font = ITAL
    for i, col in enumerate(YR_COLS):
        cell = ws.cell(row=row, column=3 + i, value=YR_HEADS[i])
        cell.font = BOLDN; cell.fill = LB_FILL; cell.alignment = C

def lbl(ws, row, text, *, bold=False, indent=0):
    cell = ws.cell(row=row, column=1, value=("  " * indent) + text)
    cell.font = BOLD if bold else BLACK; cell.alignment = L

def put_input(ws, addr, value, source, fmt=NUM):
    cell = ws[addr]; cell.value = value; cell.font = BLUE; cell.alignment = C; cell.number_format = fmt
    cell.comment = Comment(source, "RKLB model")

def put_f(ws, addr, formula, fmt=NUM, *, bold=False, fill=None):
    cell = ws[addr]; cell.value = formula
    cell.font = BOLDN if bold else BLACK; cell.alignment = C; cell.number_format = fmt
    if fill: cell.fill = fill

# DRIVERS
ws = wb.create_sheet("Drivers"); ws.sheet_view.showGridLines = False
ws["A1"] = "ROCKET LAB USA (RKLB) — INTEGRATED MODEL: DRIVERS & ASSUMPTIONS"
ws.merge_cells("A1:J1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "As of June 3, 2026 | $M unless noted | Blue=input, Black=formula | Pre-profit growth name: EV/Rev + capacity multiples are the cleaner gauge"
ws.merge_cells("A2:J2"); ws["A2"].font = ITAL; ws["A2"].alignment = C

hdr(ws, 4, 1, 10, "SEGMENT REVENUE BUILD ($M)")
yearrow(ws, 5)
r_launch = 6; r_space = 7; r_tot = 8
lbl(ws, r_launch, "Launch Services (Electron + Neutron)", indent=1)
lbl(ws, r_space, "Space Systems (incl. Mynaric)", indent=1)
lbl(ws, r_tot, "Total revenue", bold=True)

launch_h = {"C": (125, "FY24 Launch revenue $125.4M (16 launches)"),
            "D": (199, "FY25 Launch revenue $199.0M (21 launches; +59% YoY)"),
            "E": (260, "[E] FY26 Launch ~$260M: Q1 actual $63.7M + Q2 implied ramp; Neutron Q4 NET")}
space_h = {"C": (311, "FY24 Space Systems $310.8M"),
           "D": (403, "FY25 Space Systems $402.8M (+30%; ~67% of revenue)"),
           "E": (610, "[E] FY26 Space Systems ~$610M: Q1 $136.7M + Mynaric contribution + SDA T2 Beta ramp")}
for col in ("C", "D", "E"):
    put_input(ws, f"{col}{r_launch}", launch_h[col][0], launch_h[col][1])
    put_input(ws, f"{col}{r_space}", space_h[col][0], space_h[col][1])

# Projections — Neutron ramps FY27+
launch_p = {"F": 550, "G": 1100, "H": 1700, "I": 2300, "J": 2900}
space_p = {"F": 825, "G": 1080, "H": 1380, "I": 1700, "J": 2050}
launch_s = "[E] Launch: Neutron Q4 2026 first flight (expendable); FY27 ~5 Neutron @ $50M = $250M + Electron ~25 @ $9-11M = ~$280M; FY30 mature cadence 18-20 Neutron + 30 Electron"
space_s = "[E] Space Systems: SDA T2 Beta delivery FY26-27 ($1.3B contract); Mynaric optical comms scaling; Photon + bus + subsystems +30%/yr base then tapering"
for col in ("F", "G", "H", "I", "J"):
    put_input(ws, f"{col}{r_launch}", launch_p[col], launch_s)
    put_input(ws, f"{col}{r_space}", space_p[col], space_s)
for col in YR_COLS:
    put_f(ws, f"{col}{r_tot}", f"=SUM({col}{r_launch}:{col}{r_space})", bold=True)

r_growth = 9
lbl(ws, r_growth, "Total revenue growth %", indent=1)
for i, col in enumerate(YR_COLS):
    if i == 0:
        ws[f"{col}{r_growth}"] = "n/a"; ws[f"{col}{r_growth}"].font = ITAL; ws[f"{col}{r_growth}"].alignment = C
    else:
        prev = YR_COLS[i - 1]; put_f(ws, f"{col}{r_growth}", f"={col}{r_tot}/{prev}{r_tot}-1", PCT)

# Capacity drivers (small-cap-growth-specific)
hdr(ws, 11, 1, 10, "CAPACITY & UNIT ECONOMICS")
yearrow(ws, 12)
r_elc = 13; r_neut = 14; r_neutprice = 15
lbl(ws, r_elc, "Electron + HASTE launches per year", indent=1)
lbl(ws, r_neut, "Neutron launches per year", indent=1)
lbl(ws, r_neutprice, "Neutron avg price per launch ($M)", indent=1)
elc_v = {"C": 16, "D": 21, "E": 22, "F": 28, "G": 30, "H": 32, "I": 32, "J": 32}
neut_v = {"C": 0, "D": 0, "E": 1, "F": 5, "G": 12, "H": 16, "I": 18, "J": 20}
neutprice_v = {"C": 0, "D": 0, "E": 50, "F": 52, "G": 53, "H": 54, "I": 55, "J": 55}
for col in YR_COLS:
    put_input(ws, f"{col}{r_elc}", elc_v[col], "[E] Electron launches: actual FY24 16, FY25 21; ramp to ~32/yr at mature cadence")
    put_input(ws, f"{col}{r_neut}", neut_v[col], "[E] Neutron: Q4 2026 first flight (expendable); 5 FY27 -> mature 18-20/yr by FY30+ (vs SpaceX Falcon 9 ~140/yr)")
    put_input(ws, f"{col}{r_neutprice}", neutprice_v[col], "[E] Neutron target price $50-55M; first ~3 flights likely loss-leading")

# Margin / FCF drivers
hdr(ws, 17, 1, 10, "MARGIN & FCF DRIVERS")
yearrow(ws, 18)
r_gm = 19; r_opm = 20; r_da = 21; r_capex = 22; r_tax = 23
lbl(ws, r_gm, "Gross margin %", indent=1)
lbl(ws, r_opm, "Operating margin % (EBIT)", indent=1)
lbl(ws, r_da, "D&A ($M)", indent=1)
lbl(ws, r_capex, "Capex ($M)", indent=1)
lbl(ws, r_tax, "Cash tax rate %", indent=1)
gm = {"C": 0.27, "D": 0.30, "E": 0.36, "F": 0.40, "G": 0.42, "H": 0.45, "I": 0.47, "J": 0.48}
opm = {"C": -0.44, "D": -0.38, "E": -0.20, "F": -0.05, "G": 0.05, "H": 0.10, "I": 0.13, "J": 0.15}
da = {"C": 30, "D": 40, "E": 55, "F": 85, "G": 120, "H": 150, "I": 175, "J": 195}
capex = {"C": 80, "D": 120, "E": 180, "F": 220, "G": 200, "H": 170, "I": 150, "J": 140}
tax = {"C": 0, "D": 0, "E": 0, "F": 0, "G": 0.10, "H": 0.15, "I": 0.21, "J": 0.21}
gm_src = {c: "[E] GM expansion: 27%(24) -> 36%(26 Q1 actual 38.2%) -> 48%(31) as Neutron + Mynaric mature" for c in YR_COLS}
opm_src = {
    "C": "FY24 op margin ~-44%", "D": "FY25 op margin ~-38% ($229M op loss / $602M rev)",
    "E": "[E] FY26 -20% (narrowing on revenue scale)", "F": "[E] FY27 ~-5% (near breakeven)",
    "G": "[E] FY28 +5% (first profitable year, Neutron 12 launches)", "H": "[E] FY29 +10%",
    "I": "[E] FY30 +13%", "J": "[E] FY31 +15% (mature margin)",
}
da_src = {c: "[E] D&A grows with Neutron infrastructure depreciation post-FY26" for c in YR_COLS}
capex_src = {c: "[E] FY24 $80M / FY25 $120M / FY26 $180M (peak Neutron build) / taper" for c in YR_COLS}
tax_src = {c: "[E] Cash tax ~0% near-term (NOLs); ramps to ~21% by FY30 once profitable" for c in YR_COLS}
for col in YR_COLS:
    put_input(ws, f"{col}{r_gm}", gm[col], gm_src[col], PCT)
    put_input(ws, f"{col}{r_opm}", opm[col], opm_src[col], PCT)
    put_input(ws, f"{col}{r_da}", da[col], da_src[col])
    put_input(ws, f"{col}{r_capex}", capex[col], capex_src[col])
    put_input(ws, f"{col}{r_tax}", tax[col], tax_src[col], PCT)

# Market data & WACC
hdr(ws, 25, 1, 4, "MARKET DATA & WACC INPUTS")
md = [
    ("Current share price ($)", 123.32, "RKLB ~$123.32 (June 2, 2026 close); 52-wk high $150.23 (May 27)", USD2),
    ("Diluted shares (M)", 608, "~608M shares (Q1 26 w.avg + Mynaric 2.28M)", NUM),
    ("Total debt ($M)", 38, "Convertible 4.25% notes 2029 $37.6M residual (most converted)", NUM),
    ("Cash & marketable securities ($M)", 1480, "Cash + marketable securities $1.48B (Q1 26)", NUM),
    ("Risk-free rate %", 0.043, "[E] 10-yr UST ~4.3%", PCT),
    ("Equity risk premium %", 0.050, "[E] ERP ~5.0%", PCT),
    ("Beta", 1.80, "[E] RKLB beta ~1.8 (small-cap, growth, high-beta space)", "0.00"),
    ("Pre-tax cost of debt %", 0.060, "[E] If RKLB issued debt today ~6% (BB-equiv)", PCT),
    ("Terminal growth rate %", 0.035, "[E] BASE terminal growth 3.5% (space industry secular)", PCT),
]
row = 26
for name, val, src, fmt in md:
    ws.cell(row=row, column=1, value=name).font = BLACK
    put_input(ws, f"C{row}", val, src, fmt); row += 1
R_PRICE, R_SH, R_DEBT, R_CASH, R_RF, R_ERP, R_BETA, R_KD, R_TG = range(26, 35)

hdr(ws, 36, 1, 4, "WACC (computed)")
ws["A37"] = "Cost of equity"; ws["A37"].font = BLACK
put_f(ws, "C37", f"=C{R_RF}+C{R_BETA}*C{R_ERP}", PCT)
ws["A38"] = "After-tax cost of debt"; ws["A38"].font = BLACK
put_f(ws, "C38", f"=C{R_KD}*(1-J{r_tax})", PCT)  # use mature tax rate
ws["A39"] = "Market value of equity ($M)"; ws["A39"].font = BLACK
put_f(ws, "C39", f"=C{R_PRICE}*C{R_SH}", NUM)
ws["A40"] = "Net cash (debt) ($M)"; ws["A40"].font = BLACK
put_f(ws, "C40", f"=C{R_CASH}-C{R_DEBT}", NUM)
ws["A41"] = "Equity weight"; ws["A41"].font = BLACK
put_f(ws, "C41", f"=C39/(C39+C{R_DEBT})", PCT)
ws["A42"] = "Debt weight"; ws["A42"].font = BLACK
put_f(ws, "C42", f"=C{R_DEBT}/(C39+C{R_DEBT})", PCT)
ws["A43"] = "WACC"; ws["A43"].font = BOLDN
put_f(ws, "C43", "=C41*C37+C42*C38", PCT, bold=True, fill=LG_FILL)
R_WACC = 43

ws.column_dimensions["A"].width = 40
for col in YR_COLS:
    ws.column_dimensions[col].width = 11

# INCOME STATEMENT
isw = wb.create_sheet("Income Statement"); isw.sheet_view.showGridLines = False
isw["A1"] = "INCOME STATEMENT ($M)"
isw.merge_cells("A1:J1"); isw["A1"].font = BOLDW; isw["A1"].fill = NAVY_FILL; isw["A1"].alignment = C
yearrow(isw, 4)
lbl(isw, 5, "Total revenue", bold=True)
lbl(isw, 6, "Gross profit")
lbl(isw, 7, "Operating income (EBIT)")
lbl(isw, 8, "(+) D&A")
lbl(isw, 9, "EBITDA", bold=True)
lbl(isw, 10, "Pre-tax income (~ EBIT, modest int.)")
lbl(isw, 11, "(-) Income taxes")
lbl(isw, 12, "Net income (GAAP)", bold=True)
lbl(isw, 13, "Net margin %")
for col in YR_COLS:
    put_f(isw, f"{col}5", f"=Drivers!{col}{r_tot}", bold=True)
    put_f(isw, f"{col}6", f"={col}5*Drivers!{col}{r_gm}")
    put_f(isw, f"{col}7", f"=Drivers!{col}{r_opm}*{col}5")
    put_f(isw, f"{col}8", f"=Drivers!{col}{r_da}")
    put_f(isw, f"{col}9", f"={col}7+{col}8", bold=True)
    put_f(isw, f"{col}10", f"={col}7")  # minimal interest expense (small debt)
    put_f(isw, f"{col}11", f"={col}10*Drivers!{col}{r_tax}")
    put_f(isw, f"{col}12", f"={col}10-{col}11", bold=True)
    put_f(isw, f"{col}13", f"={col}12/{col}5", PCT)
isw.column_dimensions["A"].width = 40
for col in YR_COLS:
    isw.column_dimensions[col].width = 11

# BALANCE SHEET (simplified)
bs = wb.create_sheet("Balance Sheet"); bs.sheet_view.showGridLines = False
bs["A1"] = "BALANCE SHEET (simplified, $M)"
bs.merge_cells("A1:J1"); bs["A1"].font = BOLDW; bs["A1"].fill = NAVY_FILL; bs["A1"].alignment = C
yearrow(bs, 4)
lbl(bs, 5, "Cash + marketable securities")
lbl(bs, 6, "Backlog (memo, $M)")
lbl(bs, 7, "Total debt (incl. converts)")
lbl(bs, 8, "Equity issuance (ATM, $M)")
put_input(bs, "C5", 425, "[E] FY24 cash ~$425M")
put_input(bs, "D5", 1090, "[E] FY25 cash ~$1.09B (post $1.146B ATM)")
put_input(bs, "E5", 1480, "Q1 FY26 cash $1.48B")
put_input(bs, "C6", 1070, "[E] FY24 backlog ~$1.07B")
put_input(bs, "D6", 1850, "FY25 backlog $1.85B (+73% YoY)")
put_input(bs, "E6", 2220, "Q1 FY26 backlog $2.22B (+108% YoY)")
put_input(bs, "C7", 360, "[E] FY24 LT debt incl. converts ~$360M")
put_input(bs, "D7", 100, "[E] FY25 LT debt ~$100M (post-conversions)")
put_input(bs, "E7", 38, "Q1 FY26 LT debt $38M (residual converts)")
# Projected ATM raises (need cash for Neutron)
atm = {"F": (0, "[E] FY26 — no further ATM expected (cash > $1.4B)"),
       "G": (200, "[E] FY27 ATM ~$200M to fund Neutron ramp + working capital"),
       "H": (0, "[E] FY28 — internally funded by then"), "I": (0, ""), "J": (0, "")}
for col in ("F", "G", "H", "I", "J"):
    put_input(bs, f"{col}8", atm[col][0], atm[col][1])
for col in YR_COLS:
    if col in ("F", "G", "H", "I", "J"):
        prev = YR_COLS[YR_COLS.index(col)-1]
        # cash rolls: prior + FCF + ATM
        put_f(bs, f"{col}5", f"={prev}5+'Cash Flow'!{col}11+{col}8")
        put_f(bs, f"{col}7", "=38")  # converts run off
bs.column_dimensions["A"].width = 40
for col in YR_COLS:
    bs.column_dimensions[col].width = 11

# CASH FLOW
cf = wb.create_sheet("Cash Flow"); cf.sheet_view.showGridLines = False
cf["A1"] = "CASH FLOW & FCF ($M)"
cf.merge_cells("A1:J1"); cf["A1"].font = BOLDW; cf["A1"].fill = NAVY_FILL; cf["A1"].alignment = C
yearrow(cf, 4)
lbl(cf, 5, "EBIT")
lbl(cf, 6, "(-) Cash taxes on EBIT")
lbl(cf, 7, "NOPAT")
lbl(cf, 8, "(+) D&A")
lbl(cf, 9, "(-) Capex")
lbl(cf, 10, "(-) Increase in NWC (assumed 0)")
lbl(cf, 11, "Unlevered free cash flow", bold=True)
lbl(cf, 14, "Unlevered FCF margin %")
for col in YR_COLS:
    put_f(cf, f"{col}5", f"='Income Statement'!{col}7")
    put_f(cf, f"{col}6", f"=MAX({col}5,0)*Drivers!{col}{r_tax}")  # no tax shield on losses
    put_f(cf, f"{col}7", f"={col}5-{col}6", bold=True)
    put_f(cf, f"{col}8", f"=Drivers!{col}{r_da}")
    put_f(cf, f"{col}9", f"=Drivers!{col}{r_capex}")
    put_f(cf, f"{col}10", "=0")
    put_f(cf, f"{col}11", f"={col}7+{col}8-{col}9-{col}10", bold=True, fill=LG_FILL)
    put_f(cf, f"{col}14", f"={col}11/Drivers!{col}{r_tot}", PCT)
cf.column_dimensions["A"].width = 40
for col in YR_COLS:
    cf.column_dimensions[col].width = 11

# DCF + SENSITIVITY
dcf = wb.create_sheet("DCF"); dcf.sheet_view.showGridLines = False
dcf["A1"] = "DISCOUNTED CASH FLOW VALUATION (small-cap growth; wide bands)"
dcf.merge_cells("A1:H1"); dcf["A1"].font = BOLDW; dcf["A1"].fill = NAVY_FILL; dcf["A1"].alignment = C
PROJ = ["F", "G", "H", "I", "J"]
dcf["A4"] = "Projection year"; dcf["A4"].font = BOLDN; dcf["A4"].fill = LB_FILL
for i, col in enumerate(PROJ):
    cc = dcf.cell(row=4, column=2 + i, value=YR_HEADS[3 + i]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C
dcf["A5"] = "Period (n)"; dcf["A5"].font = BLACK
for i in range(5):
    put_f(dcf, f"{get_column_letter(2+i)}5", f"={i+1}")
dcf["A6"] = "Unlevered FCF"; dcf["A6"].font = BLACK
for i, col in enumerate(PROJ):
    put_f(dcf, f"{get_column_letter(2+i)}6", f"='Cash Flow'!{col}11")
dcf["A7"] = "Discount factor @ WACC"; dcf["A7"].font = BLACK
for i in range(5):
    cl = get_column_letter(2 + i); put_f(dcf, f"{cl}7", f"=1/(1+Drivers!C{R_WACC})^{cl}5", "0.000")
dcf["A8"] = "PV of unlevered FCF"; dcf["A8"].font = BOLDN
for i in range(5):
    cl = get_column_letter(2 + i); put_f(dcf, f"{cl}8", f"={cl}6*{cl}7")
br = [
    ("Sum of PV of explicit FCF (FY27-31)", "=SUM(B8:F8)", NUM),
    ("Terminal-year FCF (FY2031)", "=F6", NUM),
    ("Terminal value (Gordon)", f"=F6*(1+Drivers!C{R_TG})/(Drivers!C{R_WACC}-Drivers!C{R_TG})", NUM),
    ("PV of terminal value", "=C12*F7", NUM),
    ("Enterprise value", "=C10+C13", NUM),
    ("(+) Net cash", f"=Drivers!C{R_CASH}-Drivers!C{R_DEBT}", NUM),
    ("Equity value", "=C14+C15", NUM),
    ("Diluted shares (M)", f"=Drivers!C{R_SH}", NUM),
    ("Implied value per share ($)", "=C16/C17", USD2),
    ("Current share price ($)", f"=Drivers!C{R_PRICE}", USD2),
    ("Upside / (downside) to current", "=C18/C19-1", PCT),
    ("TV as % of enterprise value", "=C13/C14", PCT),
]
row = 10
for name, formula, fmt in br:
    dcf.cell(row=row, column=1, value=name).font = (BOLDN if "per share" in name or name == "Enterprise value" else BLACK)
    bold = name in ("Implied value per share ($)", "Enterprise value")
    put_f(dcf, f"C{row}", formula, fmt, bold=bold, fill=(MB_FILL if name == "Implied value per share ($)" else None)); row += 1

# Sensitivity — wider WACC band given small-cap risk
dcf["A22"] = "SENSITIVITY — Implied $/share: WACC (down) x Terminal growth (across)"
dcf.merge_cells("A22:H22"); dcf["A22"].font = BOLDW; dcf["A22"].fill = NAVY_FILL; dcf["A22"].alignment = C
g_axis = [0.025, 0.030, 0.035, 0.040, 0.045]
w_axis = [0.110, 0.120, 0.130, 0.140, 0.150]
dcf["B23"] = "Terminal g →"; dcf["B23"].font = ITAL
for j, g in enumerate(g_axis):
    cc = dcf.cell(row=23, column=3 + j, value=g); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT
dcf["A24"] = "WACC ↓"; dcf["A24"].font = ITAL
for r in range(5):
    cc = dcf.cell(row=24 + r, column=2, value=w_axis[r]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT
for r in range(5):
    wref = f"$B${24+r}"
    for cc in range(5):
        gref = f"{get_column_letter(3+cc)}$23"
        pv_expl = "+".join(f"${get_column_letter(2+k)}$6/(1+{wref})^{k+1}" for k in range(5))
        tv = f"$F$6*(1+{gref})/({wref}-{gref})"
        pv_tv = f"({tv})/(1+{wref})^5"
        equity = f"(({pv_expl})+({pv_tv})+Drivers!$C${R_CASH}-Drivers!$C${R_DEBT})"
        formula = f"={equity}/Drivers!$C${R_SH}"
        cell = dcf.cell(row=24 + r, column=3 + cc, value=formula)
        cell.number_format = USD2; cell.alignment = C
        if r == 2 and cc == 2:
            cell.font = Font(name="Times New Roman", size=11, bold=True); cell.fill = MB_FILL
        else:
            cell.font = BLACK
dcf["A30"] = "Center cell (WACC 13.0% / g 3.5%) = base case. Wider WACC band reflects small-cap + pre-profit + execution risk."
dcf.merge_cells("A30:H30"); dcf["A30"].font = ITAL
dcf.column_dimensions["A"].width = 40
for col in "BCDEFGH":
    dcf.column_dimensions[col].width = 12

# NOTES
nt = wb.create_sheet("Notes"); nt.sheet_view.showGridLines = False
nt["A1"] = "RKLB MODEL — METHODOLOGY & CAVEATS"
nt.merge_cells("A1:E1"); nt["A1"].font = BOLDW; nt["A1"].fill = NAVY_FILL; nt["A1"].alignment = C
notes = [
    ("Methodology — small-cap-growth adaptation", [
        "Standard DCF mechanics, but with wider WACC range (11-15%) and a Neutron-conditional FCF profile.",
        "Revenue split: Launch Services (Electron + Neutron) + Space Systems (incl. Mynaric optical comms).",
        "Pre-profit through FY27; first profitable year FY28 in base case (Neutron 12 launches @ ~$53M).",
        "Wide sensitivity bands reflect single-product (Neutron) execution risk + pre-profit base.",
    ]),
    ("Base case stance", [
        "Revenue: $602M (FY25 actual) -> $870M (FY26) -> $1.65B (FY27 first Neutron commercial yr) -> $4.95B (FY31)",
        "Op margin: -38% (FY25) -> -20% (FY26) -> -5% (FY27) -> +5% (FY28) -> +15% (FY31)",
        "Neutron: Q4 2026 first flight (expendable), 5 in FY27, 18-20/yr at mature cadence",
        "WACC 13.0% (beta 1.8, near-100% equity weight); terminal g 3.5% (space secular)",
    ]),
    ("Data sources & caveats", [
        "Rocket Lab Q1 CY26 10-Q (May 2026); FY25 10-K (Feb 2026); FY24 10-K; investor relations press releases.",
        "Neutron timeline per management commentary (Q4 2026 NET after Q4 2025 tank-test failure pushed mid-2026).",
        "Mynaric acquisition closed April 14, 2026 ($155.3M = nominal cash + 2.28M shares); contribution embedded in Space Systems projection.",
        "MCP terminals NOT configured. Forecasts heavily [E]-flagged given execution dependencies.",
        "Sanity check: ~$69B mkt cap on $602M FY25 rev = ~115x P/S; analyst PT $103.91 (S&P consensus, ~16% below spot $123.32).",
        "Research / decision-framing only — NOT investment advice.",
    ]),
]
row = 3
for title, lines in notes:
    c = nt.cell(row=row, column=1, value=title); c.font = BOLDN; c.fill = LB_FILL
    nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); row += 1
    for ln in lines:
        cc = nt.cell(row=row, column=1, value=ln); cc.font = BLACK; cc.alignment = WRAP
        nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        nt.row_dimensions[row].height = 16; row += 1
    row += 1
for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    nt.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Drivers"], wb["Income Statement"], wb["Balance Sheet"], wb["Cash Flow"], wb["DCF"], wb["Notes"]]
wb.save(OUT)
print(f"Wrote {OUT}")
