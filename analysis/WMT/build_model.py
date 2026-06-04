"""
Build WMT-Model.xlsx — integrated 3-statement model + DCF for Walmart (WMT).
Follows the same pattern as analysis/AVGO/build_model.py: Drivers · IS · BS · CF · DCF · Notes.

Segments: Walmart US · Walmart International · Sam's Club
+ explicit ad/marketplace/membership margin-mix lever (the WMT bull thesis)

Periods: FY24A FY25A FY26A | FY27E-FY31E (WMT fiscal year ends ~Jan 31)
Run: python3 build_model.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "WMT-Model.xlsx"
NAVY = "17365D"; LB = "D9E1F2"; LG = "F2F2F2"; MB = "BDD7EE"
BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLD = Font(name="Times New Roman", size=11, bold=True)
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDN = Font(name="Times New Roman", size=11, bold=True, color="17365D")
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")
NAVY_FILL = PatternFill("solid", fgColor=NAVY); LB_FILL = PatternFill("solid", fgColor=LB)
LG_FILL = PatternFill("solid", fgColor=LG); MB_FILL = PatternFill("solid", fgColor=MB)
C = Alignment("center", "center"); L = Alignment("left", "center"); WRAP = Alignment("left", "top", wrap_text=True)
NUM = "#,##0"; PCT = "0.0%"; USD2 = '"$"#,##0.00'

YR_COLS = ["C", "D", "E", "F", "G", "H", "I", "J"]
YR_HEADS = ["FY2024A", "FY2025A", "FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]

wb = Workbook(); wb.remove(wb.active)


def hdr(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text); ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0); cell.font = BOLDW; cell.fill = NAVY_FILL; cell.alignment = C


def yearrow(ws, row, label="($ in millions)"):
    ws.cell(row=row, column=1, value=label).font = ITAL
    for i, col in enumerate(YR_COLS):
        cell = ws.cell(row=row, column=3 + i, value=YR_HEADS[i])
        cell.font = BOLDN; cell.fill = LB_FILL; cell.alignment = C


def lbl(ws, row, text, *, bold=False, indent=0):
    cell = ws.cell(row=row, column=1, value=("  " * indent) + text)
    cell.font = BOLD if bold else BLACK; cell.alignment = L


def put_input(ws, addr, value, source, fmt=NUM):
    cell = ws[addr]; cell.value = value; cell.font = BLUE; cell.alignment = C; cell.number_format = fmt
    cell.comment = Comment(source, "WMT model")


def put_f(ws, addr, formula, fmt=NUM, *, bold=False, fill=None):
    cell = ws[addr]; cell.value = formula
    cell.font = BOLDN if bold else BLACK; cell.alignment = C; cell.number_format = fmt
    if fill: cell.fill = fill


# =====================================================================
# DRIVERS
# =====================================================================
ws = wb.create_sheet("Drivers"); ws.sheet_view.showGridLines = False
ws["A1"] = "WALMART (WMT) — INTEGRATED MODEL: DRIVERS & ASSUMPTIONS"
ws.merge_cells("A1:J1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "As of June 3, 2026 | $ in millions | Blue = input/assumption, Black = formula | WMT FY ends ~Jan 31 (FY26 ended Jan 2026)"
ws.merge_cells("A2:J2"); ws["A2"].font = ITAL; ws["A2"].alignment = C

hdr(ws, 4, 1, 10, "SEGMENT REVENUE BUILD ($M)")
yearrow(ws, 5)
r_us = 6; r_intl = 7; r_sams = 8; r_tot = 9
lbl(ws, r_us, "Walmart US", indent=1)
lbl(ws, r_intl, "Walmart International", indent=1)
lbl(ws, r_sams, "Sam's Club", indent=1)
lbl(ws, r_tot, "Total revenue", bold=True)

# Historicals — WMT reports segments
us_h = {
    "C": (442000, "[E] FY24 Walmart US ~$442B (10-K FY24)"),
    "D": (462000, "[E] FY25 Walmart US ~$462B"),
    "E": (485000, "[E] FY26 Walmart US ~$485B (FY26 closed Jan 2026)"),
}
intl_h = {
    "C": (101000, "[E] FY24 Intl ~$101B"),
    "D": (115000, "[E] FY25 Intl ~$115B (Flipkart, Walmex)"),
    "E": (122000, "[E] FY26 Intl ~$122B"),
}
sams_h = {
    "C": (86000, "[E] FY24 Sam's Club ~$86B"),
    "D": (90000, "[E] FY25 Sam's Club ~$90B"),
    "E": (95000, "[E] FY26 Sam's Club ~$95B"),
}
for col in ("C", "D", "E"):
    put_input(ws, f"{col}{r_us}", us_h[col][0], us_h[col][1])
    put_input(ws, f"{col}{r_intl}", intl_h[col][0], intl_h[col][1])
    put_input(ws, f"{col}{r_sams}", sams_h[col][0], sams_h[col][1])

# Projections — base case: WMT US +5%, Intl +7-8%, Sam's +5-6%
us_p = {"F": 509000, "G": 535000, "H": 562000, "I": 590000, "J": 619000}
intl_p = {"F": 131000, "G": 141000, "H": 152000, "I": 163000, "J": 175000}
sams_p = {"F": 100000, "G": 106000, "H": 112000, "I": 119000, "J": 126000}
src_us = "[E] WMT US +5%/yr base case (sustained comp +3-4% + new units)"
src_intl = "[E] Intl +7-8%/yr (Flipkart, Walmex, Latin America)"
src_sams = "[E] Sam's Club +5-6%/yr (club expansion + member fees)"
for col in ("F", "G", "H", "I", "J"):
    put_input(ws, f"{col}{r_us}", us_p[col], src_us)
    put_input(ws, f"{col}{r_intl}", intl_p[col], src_intl)
    put_input(ws, f"{col}{r_sams}", sams_p[col], src_sams)
for col in YR_COLS:
    put_f(ws, f"{col}{r_tot}", f"=SUM({col}{r_us}:{col}{r_sams})", bold=True)

# Growth row
r_growth = 10
lbl(ws, r_growth, "Total revenue growth %", indent=1)
for i, col in enumerate(YR_COLS):
    if i == 0:
        ws[f"{col}{r_growth}"] = "n/a"; ws[f"{col}{r_growth}"].font = ITAL; ws[f"{col}{r_growth}"].alignment = C
    else:
        prev = YR_COLS[i - 1]; put_f(ws, f"{col}{r_growth}", f"={col}{r_tot}/{prev}{r_tot}-1", PCT)

# Ad / membership margin-mix tier (the WMT bull thesis lever)
hdr(ws, 12, 1, 10, "HIGH-MARGIN MIX TIER (Walmart Connect ads + membership)")
yearrow(ws, 13)
r_ad = 14; r_mbr = 15
lbl(ws, r_ad, "Walmart Connect ad revenue ($M)", indent=1)
lbl(ws, r_mbr, "Membership revenue (Sam's + Walmart+, $M)", indent=1)
ad_v = {"C": 3200, "D": 4200, "E": 5500, "F": 7000, "G": 8800, "H": 10800, "I": 13000, "J": 15500}
mbr_v = {"C": 2600, "D": 2900, "E": 3300, "F": 3700, "G": 4100, "H": 4500, "I": 5000, "J": 5500}
ad_s = "[E] Walmart Connect ad revenue trajectory: ~$3B (FY24) -> ~$15B (FY31); ~40% EBIT margin per mgmt"
mbr_s = "[E] Membership (Sam's $2B+ Walmart+ ~$1B currently) grows mid-teens with member adds + fee increases"
for col in YR_COLS:
    put_input(ws, f"{col}{r_ad}", ad_v[col], ad_s)
    put_input(ws, f"{col}{r_mbr}", mbr_v[col], mbr_s)

# Margin / FCF drivers
hdr(ws, 17, 1, 10, "MARGIN & FREE-CASH-FLOW DRIVERS")
yearrow(ws, 18)
r_opm = 19; r_da = 20; r_capex = 21; r_tax = 22; r_nwc = 23
lbl(ws, r_opm, "Operating margin % (GAAP EBIT)", indent=1)
lbl(ws, r_da, "D&A % of revenue", indent=1)
lbl(ws, r_capex, "Capex % of revenue", indent=1)
lbl(ws, r_tax, "Cash tax rate %", indent=1)
lbl(ws, r_nwc, "Incr. NWC % of revenue growth", indent=1)

opm = {"C": 0.042, "D": 0.044, "E": 0.045, "F": 0.047, "G": 0.049, "H": 0.051, "I": 0.052, "J": 0.053}
da = {"C": 0.018, "D": 0.018, "E": 0.018, "F": 0.018, "G": 0.018, "H": 0.018, "I": 0.018, "J": 0.018}
capex = {"C": 0.030, "D": 0.030, "E": 0.028, "F": 0.027, "G": 0.026, "H": 0.025, "I": 0.024, "J": 0.024}
tax = {"C": 0.245, "D": 0.245, "E": 0.245, "F": 0.245, "G": 0.245, "H": 0.245, "I": 0.245, "J": 0.245}
nwc = {"C": 0.00, "D": 0.00, "E": 0.00, "F": 0.00, "G": 0.00, "H": 0.00, "I": 0.00, "J": 0.00}
opm_src = {
    "C": "FY24 EBIT margin ~4.2%", "D": "FY25 ~4.4%", "E": "FY26 ~4.5%",
    "F": "[E] FY27 4.7% — ad+marketplace+membership mix lift accelerating",
    "G": "[E] 4.9%", "H": "[E] 5.1%", "I": "[E] 5.2%", "J": "[E] 5.3% (mgmt long-term ~5%+)",
}
da_src = {c: "[E] D&A ~1.8% of revenue (WMT historical)" for c in YR_COLS}
capex_src = {
    "C": "[E] FY24 capex ~3.0% of rev (~$20B)", "D": "[E] FY25 ~3.0%",
    "E": "[E] FY26 ~2.8%", "F": "[E] FY27 ~2.7%", "G": "[E] 2.6%",
    "H": "[E] 2.5%", "I": "[E] 2.4%", "J": "[E] 2.4% (DC + automation peak passes)",
}
tax_src = {c: "[E] Cash tax rate ~24.5% (US statutory + state)" for c in YR_COLS}
nwc_src = {c: "[E] WMT runs negative working capital naturally (supplier financing); incremental ~0%" for c in YR_COLS}
for col in YR_COLS:
    put_input(ws, f"{col}{r_opm}", opm[col], opm_src[col], PCT)
    put_input(ws, f"{col}{r_da}", da[col], da_src[col], PCT)
    put_input(ws, f"{col}{r_capex}", capex[col], capex_src[col], PCT)
    put_input(ws, f"{col}{r_tax}", tax[col], tax_src[col], PCT)
    put_input(ws, f"{col}{r_nwc}", nwc[col], nwc_src[col], PCT)

# Market data & WACC
hdr(ws, 25, 1, 4, "MARKET DATA & WACC INPUTS")
md = [
    ("Current share price ($)", 95.50, "WMT ~$95.50/sh (June 3, 2026; ~$760B mkt cap range)", USD2),
    ("Diluted shares (M)", 8050, "~8.05B diluted shares (StockAnalysis)", NUM),
    ("Total debt ($M)", 60000, "[E] Total debt ~$60B (long-term + current portion)", NUM),
    ("Cash & equivalents ($M)", 11000, "[E] Cash ~$11B (latest 10-Q)", NUM),
    ("Risk-free rate %", 0.043, "[E] 10-yr UST ~4.3% (June 2026)", PCT),
    ("Equity risk premium %", 0.050, "[E] US equity risk premium ~5.0%", PCT),
    ("Beta", 0.60, "[E] WMT beta ~0.60 (defensive consumer staple)", "0.00"),
    ("Pre-tax cost of debt %", 0.045, "[E] WMT investment-grade Kd ~4.5%", PCT),
    ("Terminal growth rate %", 0.025, "[E] BASE terminal growth 2.5% (mature US retail)", PCT),
]
row = 26
for name, val, src, fmt in md:
    ws.cell(row=row, column=1, value=name).font = BLACK
    put_input(ws, f"C{row}", val, src, fmt); row += 1
R_PRICE, R_SH, R_DEBT, R_CASH, R_RF, R_ERP, R_BETA, R_KD, R_TG = range(26, 35)

# WACC
hdr(ws, 36, 1, 4, "WACC (computed)")
ws["A37"] = "Cost of equity"; ws["A37"].font = BLACK
put_f(ws, "C37", f"=C{R_RF}+C{R_BETA}*C{R_ERP}", PCT)
ws["A38"] = "After-tax cost of debt"; ws["A38"].font = BLACK
put_f(ws, "C38", f"=C{R_KD}*(1-E{r_tax})", PCT)
ws["A39"] = "Market value of equity ($M)"; ws["A39"].font = BLACK
put_f(ws, "C39", f"=C{R_PRICE}*C{R_SH}", NUM)
ws["A40"] = "Net debt ($M)"; ws["A40"].font = BLACK
put_f(ws, "C40", f"=C{R_DEBT}-C{R_CASH}", NUM)
ws["A41"] = "Equity weight"; ws["A41"].font = BLACK
put_f(ws, "C41", f"=C39/(C39+C{R_DEBT})", PCT)
ws["A42"] = "Debt weight"; ws["A42"].font = BLACK
put_f(ws, "C42", f"=C{R_DEBT}/(C39+C{R_DEBT})", PCT)
ws["A43"] = "WACC"; ws["A43"].font = BOLDN
put_f(ws, "C43", "=C41*C37+C42*C38", PCT, bold=True, fill=LG_FILL)
R_WACC = 43

ws.column_dimensions["A"].width = 36
for col in YR_COLS:
    ws.column_dimensions[col].width = 11

# =====================================================================
# INCOME STATEMENT
# =====================================================================
isw = wb.create_sheet("Income Statement"); isw.sheet_view.showGridLines = False
isw["A1"] = "INCOME STATEMENT ($M)"
isw.merge_cells("A1:J1"); isw["A1"].font = BOLDW; isw["A1"].fill = NAVY_FILL; isw["A1"].alignment = C
yearrow(isw, 4)
lbl(isw, 5, "Total revenue", bold=True)
lbl(isw, 7, "Operating income (EBIT)")
lbl(isw, 8, "(+) D&A")
lbl(isw, 9, "EBITDA", bold=True)
lbl(isw, 10, "(-) Net interest expense")
lbl(isw, 11, "Pre-tax income")
lbl(isw, 12, "(-) Income taxes")
lbl(isw, 13, "Net income (GAAP)", bold=True)
lbl(isw, 14, "Net margin %")
for col in YR_COLS:
    put_f(isw, f"{col}5", f"=Drivers!{col}9", bold=True)
    put_f(isw, f"{col}7", f"=Drivers!{col}19*{col}5")
    put_f(isw, f"{col}8", f"=Drivers!{col}20*{col}5")
    put_f(isw, f"{col}9", f"={col}7+{col}8", bold=True)
int_h = {"C": (2500, "[E] FY24 net interest ~$2.5B"), "D": (2350, "[E] FY25 ~$2.35B"), "E": (2300, "[E] FY26 ~$2.3B")}
for col in ("C", "D", "E"):
    put_input(isw, f"{col}10", int_h[col][0], int_h[col][1])
for col in ("F", "G", "H", "I", "J"):
    put_f(isw, f"{col}10", f"=AVERAGE('Balance Sheet'!{col}7,'Balance Sheet'!{YR_COLS[YR_COLS.index(col)-1]}7)*Drivers!C{R_KD}")
for col in YR_COLS:
    put_f(isw, f"{col}11", f"={col}9-{col}8-{col}10")
    put_f(isw, f"{col}12", f"={col}11*Drivers!{col}22")
    put_f(isw, f"{col}13", f"={col}11-{col}12", bold=True)
    put_f(isw, f"{col}14", f"={col}13/{col}5", PCT)
isw.column_dimensions["A"].width = 36
for col in YR_COLS:
    isw.column_dimensions[col].width = 11

# =====================================================================
# BALANCE SHEET (summary)
# =====================================================================
bs = wb.create_sheet("Balance Sheet"); bs.sheet_view.showGridLines = False
bs["A1"] = "BALANCE SHEET (summary, $M)"
bs.merge_cells("A1:J1"); bs["A1"].font = BOLDW; bs["A1"].fill = NAVY_FILL; bs["A1"].alignment = C
yearrow(bs, 4)
lbl(bs, 5, "Cash & equivalents")
lbl(bs, 6, "Other assets (plug)")
lbl(bs, 7, "Total debt")
lbl(bs, 8, "Shareholders' equity")
lbl(bs, 9, "Net debt", bold=True)
lbl(bs, 10, "Debt repayment (in year)")
lbl(bs, 11, "Dividends + buybacks")
put_input(bs, "C5", 9000, "[E] FY24 cash ~$9B")
put_input(bs, "D5", 9900, "[E] FY25 cash ~$9.9B")
put_input(bs, "E5", 11000, "[E] FY26 cash ~$11B")
put_input(bs, "C7", 60000, "[E] FY24 total debt ~$60B")
put_input(bs, "D7", 61000, "[E] FY25 ~$61B")
put_input(bs, "E7", 60000, "[E] FY26 ~$60B")
put_input(bs, "C8", 83861, "[E] FY24 equity ~$83.9B")
put_input(bs, "D8", 91013, "[E] FY25 equity ~$91B")
put_input(bs, "E8", 95000, "[E] FY26 equity ~$95B")
paydown = {"F": (2000, "[E] $2B paydown FY27"), "G": (2000, "[E] $2B FY28"), "H": (2000, "[E] $2B FY29"),
           "I": (1500, "[E] $1.5B FY30"), "J": (1500, "[E] $1.5B FY31")}
capret = {"F": (15000, "[E] FY27 div + buyback ~$15B (div ~$10B + buyback ~$5B)"),
          "G": (16000, "[E] $16B FY28"), "H": (17500, "[E] $17.5B FY29"),
          "I": (19000, "[E] $19B FY30"), "J": (20500, "[E] $20.5B FY31")}
for col in ("F", "G", "H", "I", "J"):
    put_input(bs, f"{col}10", paydown[col][0], paydown[col][1])
    put_input(bs, f"{col}11", capret[col][0], capret[col][1])
for i, col in enumerate(YR_COLS):
    prev = YR_COLS[i - 1] if i > 0 else None
    if col in ("F", "G", "H", "I", "J"):
        put_f(bs, f"{col}7", f"={prev}7-{col}10")
        put_f(bs, f"{col}8", f"={prev}8+'Income Statement'!{col}13-{col}11")
        put_f(bs, f"{col}5", f"={prev}5+'Cash Flow'!{col}13-{col}10-{col}11")
        put_f(bs, f"{col}6", f"={col}7+{col}8-{col}5")
    else:
        put_f(bs, f"{col}6", f"={col}7+{col}8-{col}5")
    put_f(bs, f"{col}9", f"={col}7-{col}5", bold=True)
bs.column_dimensions["A"].width = 36
for col in YR_COLS:
    bs.column_dimensions[col].width = 11

# =====================================================================
# CASH FLOW
# =====================================================================
cf = wb.create_sheet("Cash Flow"); cf.sheet_view.showGridLines = False
cf["A1"] = "CASH FLOW & FREE CASH FLOW ($M)"
cf.merge_cells("A1:J1"); cf["A1"].font = BOLDW; cf["A1"].fill = NAVY_FILL; cf["A1"].alignment = C
yearrow(cf, 4)
lbl(cf, 5, "EBIT")
lbl(cf, 6, "(-) Cash taxes on EBIT")
lbl(cf, 7, "NOPAT")
lbl(cf, 8, "(+) D&A")
lbl(cf, 9, "(-) Capex")
lbl(cf, 10, "(-) Increase in net working capital")
lbl(cf, 11, "Unlevered free cash flow", bold=True)
lbl(cf, 12, "(-) After-tax net interest")
lbl(cf, 13, "Levered free cash flow", bold=True)
lbl(cf, 14, "Unlevered FCF margin %")
for i, col in enumerate(YR_COLS):
    prev = YR_COLS[i - 1] if i > 0 else None
    put_f(cf, f"{col}5", f"='Income Statement'!{col}7")
    put_f(cf, f"{col}6", f"={col}5*Drivers!{col}22")
    put_f(cf, f"{col}7", f"={col}5-{col}6", bold=True)
    put_f(cf, f"{col}8", f"='Income Statement'!{col}8")
    put_f(cf, f"{col}9", f"=Drivers!{col}21*Drivers!{col}9")
    if prev:
        put_f(cf, f"{col}10", f"=(Drivers!{col}9-Drivers!{prev}9)*Drivers!{col}23")
    else:
        put_f(cf, f"{col}10", "=0")
    put_f(cf, f"{col}11", f"={col}7+{col}8-{col}9-{col}10", bold=True, fill=LG_FILL)
    put_f(cf, f"{col}12", f"='Income Statement'!{col}10*(1-Drivers!{col}22)")
    put_f(cf, f"{col}13", f"={col}11-{col}12", bold=True)
    put_f(cf, f"{col}14", f"={col}11/Drivers!{col}9", PCT)
cf.column_dimensions["A"].width = 36
for col in YR_COLS:
    cf.column_dimensions[col].width = 11

# =====================================================================
# DCF + SENSITIVITY
# =====================================================================
dcf = wb.create_sheet("DCF"); dcf.sheet_view.showGridLines = False
dcf["A1"] = "DISCOUNTED CASH FLOW VALUATION"
dcf.merge_cells("A1:H1"); dcf["A1"].font = BOLDW; dcf["A1"].fill = NAVY_FILL; dcf["A1"].alignment = C
PROJ = ["F", "G", "H", "I", "J"]
dcf["A4"] = "Projection year"; dcf["A4"].font = BOLDN; dcf["A4"].fill = LB_FILL
for i, col in enumerate(PROJ):
    cc = dcf.cell(row=4, column=2 + i, value=YR_HEADS[3 + i]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C
dcf["A5"] = "Period (n)"; dcf["A5"].font = BLACK
for i in range(5):
    put_f(dcf, f"{get_column_letter(2+i)}5", f"={i+1}")
dcf["A6"] = "Unlevered FCF"; dcf["A6"].font = BLACK
for i, col in enumerate(PROJ):
    put_f(dcf, f"{get_column_letter(2+i)}6", f"='Cash Flow'!{col}11")
dcf["A7"] = "Discount factor @ WACC"; dcf["A7"].font = BLACK
for i in range(5):
    cl = get_column_letter(2 + i); put_f(dcf, f"{cl}7", f"=1/(1+Drivers!C{R_WACC})^{cl}5", "0.000")
dcf["A8"] = "PV of unlevered FCF"; dcf["A8"].font = BOLDN
for i in range(5):
    cl = get_column_letter(2 + i); put_f(dcf, f"{cl}8", f"={cl}6*{cl}7")

br = [
    ("Sum of PV of explicit FCF (FY27-31)", "=SUM(B8:F8)", NUM),
    ("Terminal-year FCF (FY2031)", "=F6", NUM),
    ("Terminal value (Gordon)", f"=F6*(1+Drivers!C{R_TG})/(Drivers!C{R_WACC}-Drivers!C{R_TG})", NUM),
    ("PV of terminal value", "=C12*F7", NUM),
    ("Enterprise value", "=C10+C13", NUM),
    ("(-) Net debt", f"=Drivers!C{R_DEBT}-Drivers!C{R_CASH}", NUM),
    ("Equity value", "=C14-C15", NUM),
    ("Diluted shares (M)", f"=Drivers!C{R_SH}", NUM),
    ("Implied value per share ($)", "=C16/C17", USD2),
    ("Current share price ($)", f"=Drivers!C{R_PRICE}", USD2),
    ("Upside / (downside) to current", "=C18/C19-1", PCT),
    ("TV as % of enterprise value", "=C13/C14", PCT),
]
row = 10
for name, formula, fmt in br:
    dcf.cell(row=row, column=1, value=name).font = (BOLDN if "per share" in name or name == "Enterprise value" else BLACK)
    bold = name in ("Implied value per share ($)", "Enterprise value")
    put_f(dcf, f"C{row}", formula, fmt, bold=bold, fill=(MB_FILL if name == "Implied value per share ($)" else None)); row += 1

# 5x5 sensitivity
dcf["A22"] = "SENSITIVITY — Implied $/share: WACC (down) x Terminal growth (across)"
dcf.merge_cells("A22:H22"); dcf["A22"].font = BOLDW; dcf["A22"].fill = NAVY_FILL; dcf["A22"].alignment = C
g_axis = [0.015, 0.020, 0.025, 0.030, 0.035]
w_axis = [0.060, 0.065, 0.070, 0.075, 0.080]
dcf["B23"] = "Terminal g →"; dcf["B23"].font = ITAL
for j, g in enumerate(g_axis):
    cc = dcf.cell(row=23, column=3 + j, value=g); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT
dcf["A24"] = "WACC ↓"; dcf["A24"].font = ITAL
for r in range(5):
    cc = dcf.cell(row=24 + r, column=2, value=w_axis[r]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT
for r in range(5):
    wref = f"$B${24+r}"
    for cc in range(5):
        gref = f"{get_column_letter(3+cc)}$23"
        pv_expl = "+".join(f"${get_column_letter(2+k)}$6/(1+{wref})^{k+1}" for k in range(5))
        tv = f"$F$6*(1+{gref})/({wref}-{gref})"
        pv_tv = f"({tv})/(1+{wref})^5"
        equity = f"(({pv_expl})+({pv_tv})-Drivers!$C${R_DEBT}+Drivers!$C${R_CASH})"
        formula = f"={equity}/Drivers!$C${R_SH}"
        cell = dcf.cell(row=24 + r, column=3 + cc, value=formula)
        cell.number_format = USD2; cell.alignment = C
        if r == 2 and cc == 2:
            cell.font = Font(name="Times New Roman", size=11, bold=True); cell.fill = MB_FILL
        else:
            cell.font = BLACK
dcf["A30"] = "Center cell (WACC 7.0% / g 2.5%) = base case; cross-check vs C18 implied per share."
dcf.merge_cells("A30:H30"); dcf["A30"].font = ITAL
dcf.column_dimensions["A"].width = 38
for col in "BCDEFGH":
    dcf.column_dimensions[col].width = 12

# =====================================================================
# NOTES
# =====================================================================
nt = wb.create_sheet("Notes"); nt.sheet_view.showGridLines = False
nt["A1"] = "WMT MODEL — METHODOLOGY & CAVEATS"
nt.merge_cells("A1:E1"); nt["A1"].font = BOLDW; nt["A1"].fill = NAVY_FILL; nt["A1"].alignment = C
notes = [
    ("Methodology", [
        "Integrated 3-statement model (IS/BS/CF) feeding an unlevered DCF. Drivers tab drives everything.",
        "Revenue built bottom-up: Walmart US + Walmart International + Sam's Club.",
        "Separate visibility on the WMT thesis margin lever: Walmart Connect ad revenue + membership.",
        "DCF: 5-year explicit FCF (FY27-31) + Gordon terminal value; net debt deducted to get equity.",
    ]),
    ("Base case stance", [
        "WMT US +5%/yr (sustained comp +3-4% + new units); Intl +7-8% (Flipkart, Walmex); Sam's +5-6%.",
        "EBIT margin expands 4.5% (FY26) -> 5.3% (FY31) as ad+marketplace+membership mix lifts margin.",
        "Walmart Connect ramps ~$5.5B (FY26) -> ~$15.5B (FY31) at ~40% EBIT margin per mgmt.",
        "Lower WACC (~7%) and terminal growth (2.5%) than AVGO — defensive consumer-staple profile.",
    ]),
    ("Data sources & caveats", [
        "WMT 10-K FY26 (Mar 2026), FY25 / FY24 10-Ks; segment splits per management disclosure (partly [E]).",
        "MCP terminal connectors (CapIQ/FactSet/Daloopa) NOT configured in environment.",
        "Walmart Connect ad-revenue figures are mgmt commentary aggregates — not separately reported line items.",
        "Terminal value typically ~65-70% of EV — sensitive but less so than a high-growth name like AVGO.",
        "Research / decision-framing only — NOT investment advice.",
    ]),
]
row = 3
for title, lines in notes:
    c = nt.cell(row=row, column=1, value=title); c.font = BOLDN; c.fill = LB_FILL
    nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); row += 1
    for ln in lines:
        cc = nt.cell(row=row, column=1, value=ln); cc.font = BLACK; cc.alignment = WRAP
        nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        nt.row_dimensions[row].height = 16; row += 1
    row += 1
for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    nt.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Drivers"], wb["Income Statement"], wb["Balance Sheet"], wb["Cash Flow"], wb["DCF"], wb["Notes"]]
wb.save(OUT)
print(f"Wrote {OUT}")
