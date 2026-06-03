"""
Build WMT-Comps-Analysis.xlsx — institutional-grade comparable company analysis
following the `comps-analysis` skill in plugins/vertical-plugins/financial-analysis.

Conventions enforced:
  - Blue font = hardcoded input (every input has a cell comment citing source)
  - Black font = formula
  - Stats block (Max / 75th / Median / 25th / Min) on every comparable column
  - Times New Roman, navy/light-blue/grey palette only
  - Retail-Specific tab: comp sales growth %, e-com penetration, ad revenue,
    membership economics — the columns that actually differentiate the peer set

Run:
    python3 build_comps.py
Output:
    WMT-Comps-Analysis.xlsx (sibling of this script)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "WMT-Comps-Analysis.xlsx"

NAVY = "17365D"
LIGHTBLUE = "D9E1F2"
LIGHTGREY = "F2F2F2"

BLUE_FONT = Font(name="Times New Roman", size=11, color="0000FF")   # hardcoded input
BLACK_FONT = Font(name="Times New Roman", size=11, color="000000")  # formula
BOLD_WHITE = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLD_BLACK = Font(name="Times New Roman", size=12, bold=True)
ITAL_GREY = Font(name="Times New Roman", size=10, italic=True, color="595959")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
LIGHTBLUE_FILL = PatternFill("solid", fgColor=LIGHTBLUE)
LIGHTGREY_FILL = PatternFill("solid", fgColor=LIGHTGREY)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

PEERS = [
    # (name, ticker)
    ("Walmart",          "NYSE: WMT"),
    ("Costco Wholesale", "Nasdaq: COST"),
    ("Target",           "NYSE: TGT"),
    ("Amazon",           "Nasdaq: AMZN"),
    ("Kroger",           "NYSE: KR"),
    ("BJ's Wholesale",   "NYSE: BJ"),
]

# Raw inputs — all $ figures in $ millions unless noted.
# Period: TTM through FY26 Q1 (calendar Apr 2026) where reported; FY25 final otherwise.
# [E] = estimate / not directly disclosed. Every input has a cell comment.
INPUTS = {
    "Walmart": {
        "mkt_cap":         (765000, "WMT ~$95/share × ~8.05B shares, May 2026 (~$765B)"),
        "net_debt":        (45000,  "[E] Net debt ~$45B excl. operating leases, per FY26 Q1 10-Q balance sheet"),
        "revenue_ttm":     (685000, "TTM revenue ~$685B (FY25 $681B + Q1 FY26 +6% YoY); WMT 10-K FY25 + Q1 FY26 release"),
        "rev_growth":      (0.060,  "FY26 Q1 revenue growth ~+6% YoY (currency-neutral ~+5.4%) per WMT IR May 2026"),
        "gross_profit":    (167000, "[E] Gross margin ~24.4% × $685B TTM (per Q1 FY26 release ~+25 bps YoY)"),
        "ebitda_ttm":      (45000,  "[E] TTM EBITDA ~$45B (operating income $30B + D&A ~$15B), per FY25 10-K + Q1 trend"),
        "net_income_ttm":  (17000,  "[E] TTM net income ~$17B per WMT IR + StockAnalysis"),
        "comp_growth":     (0.045,  "WMT US comparable sales growth +4.5% (ex-fuel) FY26 Q1 per IR"),
        "ecom_mix":        (0.18,   "[E] WMT US e-commerce ~18% of US revenue, ~$110B run-rate per Q1 release"),
        "ad_revenue":      (5000,   "[E] Walmart Connect ad revenue ~$5B run-rate (mgmt commentary; growing ~30% YoY)"),
        "membership_rev":  (3000,   "[E] Walmart+ + Sam's Club membership fees ~$3B (Sam's $2B + Walmart+ ~$1B)"),
    },
    "Costco Wholesale": {
        "mkt_cap":         (420000, "COST ~$945/share × ~444M shares, May 2026 (~$420B)"),
        "net_debt":        (-3000,  "Net cash position ~$3B per Q3 FY26 10-Q (cash $13B vs debt ~$10B)"),
        "revenue_ttm":     (270000, "TTM revenue ~$270B (FY25 $254B + Q1-Q3 FY26 growth) per COST IR"),
        "rev_growth":      (0.080,  "FY26 net sales growth ~+8% YoY per recent quarterly reports"),
        "gross_profit":    (33750,  "[E] Gross margin ~12.5% × $270B (membership accounts for most operating profit)"),
        "ebitda_ttm":      (13000,  "[E] TTM EBITDA ~$13B (operating income ~$10.5B + D&A ~$2.5B)"),
        "net_income_ttm":  (8000,   "[E] TTM net income ~$8B per COST 10-K FY25 + FY26 YTD"),
        "comp_growth":     (0.062,  "Costco US comp sales +6.2% (ex-gas/FX) Q3 FY26"),
        "ecom_mix":        (0.075,  "[E] Costco e-commerce ~7.5% of revenue (vs ~16% for WMT US)"),
        "ad_revenue":      (0,      "Negligible advertising business; Costco does not run a retail-media platform"),
        "membership_rev":  (5000,   "Membership fee revenue ~$5B TTM (mgmt-disclosed; +9% YoY at last report)"),
    },
    "Target": {
        "mkt_cap":         (70000,  "TGT ~$150/share × ~465M shares, May 2026 (~$70B)"),
        "net_debt":        (13000,  "[E] Net debt ~$13B per TGT FY25 10-K balance sheet"),
        "revenue_ttm":     (107000, "TTM revenue ~$107B per TGT IR + FY26 Q1 release"),
        "rev_growth":      (0.015,  "[E] FY26 Q1 revenue growth ~+1.5% YoY, recovering from FY25 decline"),
        "gross_profit":    (29960,  "[E] Gross margin ~28% × $107B (per TGT Q1 release)"),
        "ebitda_ttm":      (10000,  "[E] TTM EBITDA ~$10B (operating income ~$6.5B + D&A ~$3.5B)"),
        "net_income_ttm":  (4200,   "[E] TTM net income ~$4.2B per TGT IR"),
        "comp_growth":     (0.020,  "[E] TGT US comparable sales +2% Q1 FY26 (modest recovery vs FY25 declines)"),
        "ecom_mix":        (0.20,   "[E] TGT digital sales ~20% of revenue per FY25 10-K"),
        "ad_revenue":      (2000,   "[E] Roundel ad business ~$2B annualized revenue (mgmt commentary 2025)"),
        "membership_rev":  (500,    "[E] Target Circle 360 membership rev (small; launched 2024)"),
    },
    "Amazon": {
        "mkt_cap":         (2200000,"AMZN ~$210/share × ~10.6B shares, May 2026 (~$2.2T)"),
        "net_debt":        (30000,  "[E] Net debt ~$30B (excl. operating leases ~$80B+) per Q1 FY26 10-Q"),
        "revenue_ttm":     (650000, "TTM consolidated revenue ~$650B (FY25 $638B + Q1 FY26 ~+10% growth)"),
        "rev_growth":      (0.100,  "[E] Q1 FY26 revenue growth ~+10% YoY (AWS +18%, Retail ~+8%)"),
        "gross_profit":    (305000, "[E] Gross margin ~47% × $650B (consolidated; AWS is high-margin)"),
        "ebitda_ttm":      (160000, "[E] TTM EBITDA ~$160B (operating income ~$80B + D&A ~$80B)"),
        "net_income_ttm":  (65000,  "[E] TTM net income ~$65B per AMZN IR"),
        "comp_growth":     (None,   "N/A — AMZN does not report comparable-store sales; not a retail-store model"),
        "ecom_mix":        (1.00,   "AMZN North America + International segments are ~entirely e-commerce/marketplace"),
        "ad_revenue":      (60000,  "TTM advertising services revenue ~$60B (AMZN Q1 FY26 segment disclosure)"),
        "membership_rev":  (45000,  "[E] Prime membership + subscription services ~$45B TTM (segment disclosure)"),
    },
    "Kroger": {
        "mkt_cap":         (50000,  "KR ~$70/share × ~715M shares, May 2026 (~$50B)"),
        "net_debt":        (13000,  "[E] Net debt ~$13B per KR Q4 FY25 / Q1 FY26 balance sheet"),
        "revenue_ttm":     (148000, "TTM revenue ~$148B per KR IR"),
        "rev_growth":      (0.020,  "[E] FY26 Q1 revenue growth ~+2% YoY (post-Albertsons-merger-block environment)"),
        "gross_profit":    (32560,  "[E] Gross margin ~22% × $148B (per KR FY25 10-K)"),
        "ebitda_ttm":      (7000,   "[E] TTM EBITDA ~$7B (operating income ~$4B + D&A ~$3B)"),
        "net_income_ttm":  (2500,   "[E] TTM net income ~$2.5B per KR IR"),
        "comp_growth":     (0.025,  "[E] KR identical sales (ex-fuel) +2.5% Q1 FY26"),
        "ecom_mix":        (0.10,   "[E] KR digital sales ~10% of revenue; growth ~15% YoY"),
        "ad_revenue":      (1500,   "[E] Kroger Precision Marketing ad revenue ~$1.5B annualized"),
        "membership_rev":  (0,      "No standalone membership program (Boost loyalty is free-tier)"),
    },
    "BJ's Wholesale": {
        "mkt_cap":         (15000,  "BJ ~$110/share × ~133M shares, May 2026 (~$15B)"),
        "net_debt":        (500,    "[E] Modest net debt ~$0.5B per BJ Q4 FY25 / Q1 FY26 balance sheet"),
        "revenue_ttm":     (22000,  "TTM revenue ~$22B per BJ IR"),
        "rev_growth":      (0.050,  "[E] FY26 Q1 revenue growth ~+5% YoY"),
        "gross_profit":    (3960,   "[E] Gross margin ~18% × $22B (per BJ FY25 10-K incl. membership)"),
        "ebitda_ttm":      (1100,   "[E] TTM EBITDA ~$1.1B (operating income ~$0.85B + D&A ~$0.25B)"),
        "net_income_ttm":  (550,    "[E] TTM net income ~$0.55B per BJ IR"),
        "comp_growth":     (0.038,  "[E] BJ comparable club sales (ex-gas) ~+3.8% Q1 FY26"),
        "ecom_mix":        (0.05,   "[E] BJ digital sales ~5% of revenue"),
        "ad_revenue":      (0,      "Negligible ad business at current scale"),
        "membership_rev":  (470,    "[E] BJ membership fee revenue ~$470M TTM"),
    },
}

# =========================================================
# WORKBOOK
# =========================================================
wb = Workbook()
wb.remove(wb.active)

def section_header(ws, row, start_col, end_col, text):
    ws.cell(row=row, column=start_col, value=text)
    rng = f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"
    ws.merge_cells(rng)
    c = ws.cell(row=row, column=start_col)
    c.font = BOLD_WHITE
    c.fill = NAVY_FILL
    c.alignment = CENTER

def column_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD_BLACK
    c.fill = LIGHTBLUE_FILL
    c.alignment = CENTER

def stat_row(ws, row, label_col, label, n_cols, data_first_row, data_last_row,
             percent_cols=None, mult_cols=None, decimal_cols=None):
    c = ws.cell(row=row, column=label_col, value=label)
    c.font = BOLD_BLACK
    c.fill = LIGHTGREY_FILL
    c.alignment = LEFT
    for col in range(label_col + 1, label_col + n_cols + 1):
        col_letter = get_column_letter(col)
        rng = f"{col_letter}{data_first_row}:{col_letter}{data_last_row}"
        if label == "Maximum":
            formula = f"=MAX({rng})"
        elif label == "75th Percentile":
            formula = f"=QUARTILE({rng},3)"
        elif label == "Median":
            formula = f"=MEDIAN({rng})"
        elif label == "25th Percentile":
            formula = f"=QUARTILE({rng},1)"
        elif label == "Minimum":
            formula = f"=MIN({rng})"
        cell = ws.cell(row=row, column=col, value=formula)
        cell.fill = LIGHTGREY_FILL
        cell.font = BOLD_BLACK
        cell.alignment = CENTER
        if percent_cols and col in percent_cols:
            cell.number_format = "0.0%"
        elif mult_cols and col in mult_cols:
            cell.number_format = "0.0\"x\""
        elif decimal_cols and col in decimal_cols:
            cell.number_format = "#,##0.00"
        else:
            cell.number_format = "#,##0"

def write_input(ws, row, col, value, source, percent=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLUE_FONT
    cell.alignment = CENTER
    if value is None:
        cell.value = "N/A"
        cell.font = ITAL_GREY
    else:
        if percent:
            cell.number_format = "0.0%"
        else:
            cell.number_format = "#,##0"
    cell.comment = Comment(source, "WMT comp analysis")

def write_formula(ws, row, col, formula, fmt="#,##0"):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = BLACK_FONT
    cell.alignment = CENTER
    cell.number_format = fmt

# =========================================================
# Tab 1: INPUTS
# =========================================================
ws = wb.create_sheet("Inputs")
ws.sheet_view.showGridLines = False

ws["A1"] = "BIG-BOX & OMNICHANNEL RETAIL — COMPARABLE COMPANY ANALYSIS"
ws.merge_cells("A1:N1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER

ws["A2"] = " • ".join(f"{n} ({t})" for n, t in PEERS)
ws.merge_cells("A2:N2"); ws["A2"].font = BOLD_BLACK; ws["A2"].alignment = CENTER

ws["A3"] = "As of May 2026 | TTM through FY26 Q1 (Apr 2026) where reported | All $ figures in USD millions except ratios and %"
ws.merge_cells("A3:N3"); ws["A3"].font = ITAL_GREY; ws["A3"].alignment = CENTER

section_header(ws, 5, 1, 14, "RAW INPUTS — cell comments cite source / assumption")

headers = ["Company", "Ticker", "Mkt Cap", "Net Debt", "Revenue (TTM)",
           "Rev Growth %", "Gross Profit", "EBITDA (TTM)", "Net Income (TTM)",
           "Comp Sales %", "E-com % of Rev", "Ad Revenue ($M)", "Mbrshp Rev ($M)",
           "Notes"]
for i, h in enumerate(headers, start=1):
    column_header(ws, 6, i, h)

for idx, (name, ticker) in enumerate(PEERS):
    r = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    ws.cell(row=r, column=2, value=ticker).font = BLACK_FONT
    d = INPUTS[name]
    write_input(ws, r, 3, d["mkt_cap"][0],         d["mkt_cap"][1])
    write_input(ws, r, 4, d["net_debt"][0],        d["net_debt"][1])
    write_input(ws, r, 5, d["revenue_ttm"][0],     d["revenue_ttm"][1])
    write_input(ws, r, 6, d["rev_growth"][0],      d["rev_growth"][1], percent=True)
    write_input(ws, r, 7, d["gross_profit"][0],    d["gross_profit"][1])
    write_input(ws, r, 8, d["ebitda_ttm"][0],      d["ebitda_ttm"][1])
    write_input(ws, r, 9, d["net_income_ttm"][0],  d["net_income_ttm"][1])
    write_input(ws, r,10, d["comp_growth"][0],     d["comp_growth"][1], percent=True)
    write_input(ws, r,11, d["ecom_mix"][0],        d["ecom_mix"][1],    percent=True)
    write_input(ws, r,12, d["ad_revenue"][0],      d["ad_revenue"][1])
    write_input(ws, r,13, d["membership_rev"][0],  d["membership_rev"][1])
    notes = ""
    if any("[E]" in str(d[k][1]) for k in d):
        notes = "Contains [E] estimates — see cell comments"
    ws.cell(row=r, column=14, value=notes).font = ITAL_GREY

widths = [22, 14, 13, 12, 14, 13, 13, 14, 16, 13, 14, 14, 14, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 2: OPERATING METRICS
# =========================================================
ws = wb.create_sheet("Operating Metrics")
ws.sheet_view.showGridLines = False
ws["A1"] = "OPERATING STATISTICS & FINANCIAL METRICS"
ws.merge_cells("A1:G1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER
ws["A2"] = "All $ figures USD millions except margins/growth. Formulas reference Inputs tab."
ws.merge_cells("A2:G2"); ws["A2"].font = ITAL_GREY; ws["A2"].alignment = CENTER

op_headers = ["Company", "Revenue (TTM)", "Rev Growth (YoY)", "Gross Profit",
              "Gross Margin", "EBITDA (TTM)", "EBITDA Margin"]
for i, h in enumerate(op_headers, start=1):
    column_header(ws, 4, i, h)

for idx, (name, _) in enumerate(PEERS):
    r = 5 + idx
    src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    write_formula(ws, r, 2, f"=Inputs!E{src}")
    write_formula(ws, r, 3, f"=IFERROR(Inputs!F{src},\"N/A\")", "0.0%")
    write_formula(ws, r, 4, f"=IFERROR(Inputs!G{src},\"N/A\")")
    write_formula(ws, r, 5, f"=IFERROR(Inputs!G{src}/Inputs!E{src},\"N/A\")", "0.0%")
    write_formula(ws, r, 6, f"=Inputs!H{src}")
    write_formula(ws, r, 7, f"=IFERROR(Inputs!H{src}/Inputs!E{src},\"N/A\")", "0.0%")

DATA_FIRST = 5
DATA_LAST = 5 + len(PEERS) - 1
stat_first = DATA_LAST + 2
labels = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]
percent_cols = {3, 5, 7}
for i, lab in enumerate(labels):
    stat_row(ws, stat_first + i, 1, lab, 6, DATA_FIRST, DATA_LAST, percent_cols=percent_cols)
    for col in (2, 4, 6):
        cell = ws.cell(row=stat_first + i, column=col, value="")
        cell.fill = LIGHTGREY_FILL

widths = [22, 16, 16, 16, 14, 16, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 3: VALUATION
# =========================================================
ws = wb.create_sheet("Valuation")
ws.sheet_view.showGridLines = False
ws["A1"] = "VALUATION MULTIPLES"
ws.merge_cells("A1:G1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER
ws["A2"] = "EV = Market Cap + Net Debt (excl. operating leases). NM where denominator <= 0."
ws.merge_cells("A2:G2"); ws["A2"].font = ITAL_GREY; ws["A2"].alignment = CENTER

val_headers = ["Company", "Mkt Cap", "EV", "EV / Rev", "EV / EBITDA", "P / E (TTM)", "Comment"]
for i, h in enumerate(val_headers, start=1):
    column_header(ws, 4, i, h)

for idx, (name, _) in enumerate(PEERS):
    r = 5 + idx
    src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    write_formula(ws, r, 2, f"=Inputs!C{src}")
    write_formula(ws, r, 3, f"=Inputs!C{src}+Inputs!D{src}")
    write_formula(ws, r, 4, f"=IFERROR((Inputs!C{src}+Inputs!D{src})/Inputs!E{src},\"NM\")", "0.0\"x\"")
    write_formula(ws, r, 5, f"=IF(Inputs!H{src}>0,(Inputs!C{src}+Inputs!D{src})/Inputs!H{src},\"NM\")", "0.0\"x\"")
    write_formula(ws, r, 6, f"=IF(Inputs!I{src}>0,Inputs!C{src}/Inputs!I{src},\"NM\")", "0.0\"x\"")
    comment = {
        "Walmart": "Re-rating in progress as ad/marketplace mix expands EBITDA margin",
        "Costco Wholesale": "Premium-priced peer; membership economics support 40%+ P/E premium to WMT",
        "Target": "Cheapest large-cap mass merchant; pricing-in slower comp recovery",
        "Amazon": "Headline multiples reflect AWS economics; retail-only would screen differently",
        "Kroger": "Pure-play grocery; thin EBITDA margin caps multiple expansion",
        "BJ's Wholesale": "Warehouse-club discount to COST; smaller scale, weaker membership renewal data",
    }[name]
    cc = ws.cell(row=r, column=7, value=comment); cc.font = ITAL_GREY; cc.alignment = LEFT

for i, lab in enumerate(labels):
    stat_row(ws, stat_first + i, 1, lab, 6, DATA_FIRST, DATA_LAST, mult_cols={4, 5, 6})
    for col in (2, 3, 7):
        cell = ws.cell(row=stat_first + i, column=col, value="")
        cell.fill = LIGHTGREY_FILL

widths = [22, 12, 12, 12, 14, 12, 52]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 4: RETAIL-SPECIFIC
# =========================================================
ws = wb.create_sheet("Retail-Specific")
ws.sheet_view.showGridLines = False
ws["A1"] = "RETAIL-SPECIFIC METRICS — Comp Sales, Mix, Ads & Membership"
ws.merge_cells("A1:H1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER
ws["A2"] = "Industry-defining metrics: comparable sales, e-com penetration, and the high-margin tiers (ads + membership) that drive margin mix."
ws.merge_cells("A2:H2"); ws["A2"].font = ITAL_GREY; ws["A2"].alignment = CENTER

ret_headers = ["Company", "Comp Sales %", "E-com % of Rev", "Ad Revenue ($M)",
               "Ad % of Rev", "Membership Rev ($M)", "Mbrshp % of Rev", "Posture"]
for i, h in enumerate(ret_headers, start=1):
    column_header(ws, 4, i, h)

postures = {
    "Walmart":          "Scale incumbent — building ad + marketplace + Walmart+ on top of grocery flywheel",
    "Costco Wholesale": "Membership-first warehouse club — premium loyalty, low-margin merchandise",
    "Target":           "Cheap-chic mass — Roundel ad business is the bright spot in a soft topline",
    "Amazon":           "Marketplace-first — ad business is ~10% of revenue and the highest-margin retail tier",
    "Kroger":           "Pure-play grocer — Precision Marketing ad business is the margin lever",
    "BJ's Wholesale":   "Sub-scale COST follower — value-led, limited ad/digital optionality",
}

for idx, (name, _) in enumerate(PEERS):
    r = 5 + idx
    src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLD_BLACK
    write_formula(ws, r, 2, f"=IFERROR(Inputs!J{src},\"N/A\")", "0.0%")
    write_formula(ws, r, 3, f"=IFERROR(Inputs!K{src},\"N/A\")", "0.0%")
    write_formula(ws, r, 4, f"=Inputs!L{src}")
    write_formula(ws, r, 5, f"=IFERROR(Inputs!L{src}/Inputs!E{src},\"N/A\")", "0.0%")
    write_formula(ws, r, 6, f"=Inputs!M{src}")
    write_formula(ws, r, 7, f"=IFERROR(Inputs!M{src}/Inputs!E{src},\"N/A\")", "0.0%")
    pc = ws.cell(row=r, column=8, value=postures[name]); pc.font = ITAL_GREY; pc.alignment = LEFT

# Stats — comparable cols: 2 (comp %), 3 (e-com %), 5 (ad %), 7 (mbrshp %)
for i, lab in enumerate(labels):
    stat_row(ws, stat_first + i, 1, lab, 7, DATA_FIRST, DATA_LAST,
             percent_cols={2, 3, 5, 7})
    for col in (4, 6, 8):
        cell = ws.cell(row=stat_first + i, column=col, value="")
        cell.fill = LIGHTGREY_FILL

widths = [22, 14, 16, 16, 14, 18, 16, 54]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =========================================================
# Tab 5: NOTES
# =========================================================
ws = wb.create_sheet("Notes")
ws.sheet_view.showGridLines = False
ws["A1"] = "NOTES, METHODOLOGY & DATA-SOURCE CAVEATS"
ws.merge_cells("A1:E1"); c = ws["A1"]; c.font = BOLD_WHITE; c.fill = NAVY_FILL; c.alignment = CENTER

notes_blocks = [
    ("Data sources", [
        "MCP data sources (S&P Kensho / CapIQ, FactSet, Daloopa) WERE NOT USED — the FSI plugin connectors were not configured in this environment.",
        "Primary public sources: SEC EDGAR (10-K FY25 and 10-Q FY26 Q1 for each issuer), company IR releases, StockAnalysis.com / Yahoo Finance for market data.",
        "Where required figures were not directly disclosed, an [E] estimate is shown with the underlying assumption documented in the cell comment.",
    ]),
    ("Period definitions", [
        "WMT, TGT, COST use fiscal years offset from calendar — FY26 Q1 maps to calendar Feb-Apr 2026 (WMT) / Mar-May 2026 (TGT) / Feb-May 2026 (COST).",
        "AMZN, KR, BJ use calendar-year reporting; AMZN Q1 FY26 = Jan-Mar 2026.",
        "TTM = trailing four reported quarters as of latest filing; mixed fiscal-year ends are flagged where material.",
        "Comparable / identical sales: per each issuer's published definition; ex-fuel for WMT/KR/BJ; ex-gasoline/FX for COST; ex-gas for KR.",
    ]),
    ("Definitions", [
        "EV = Market Cap + Net Debt (Total Debt - Cash). Operating-lease liabilities excluded — they are material for AMZN especially; flagged in Notes.",
        "Ad revenue: 'retail media' revenue reported as a distinct line where disclosed (AMZN advertising services; mgmt estimates for WMT Connect, Roundel, Kroger Precision Marketing).",
        "Membership revenue: dues / fees only (COST membership, Sam's Club, BJ membership, Walmart+ subscription, AMZN Prime + subscription services).",
        "[NM] = not meaningful (multiple denominator <= 0). [E] = estimate, source of estimate noted in cell comment.",
    ]),
    ("Industry-specific guidance", [
        "Standard EV/EBITDA is reasonably comparable across this set, but headline AMZN multiples are distorted by AWS (~30%+ of consolidated EBITDA at much higher margin than retail).",
        "For Walmart vs. peers, the most differentiated comp lenses are: (a) ad % of revenue (margin mix shift), (b) membership revenue $/customer, (c) comp sales growth.",
        "Costco's premium P/E reflects (i) renewal-rate certainty (~92%+), (ii) negative net debt, (iii) executive-membership pricing power — none of which the others fully replicate.",
        "BJ trades at a structural discount to COST; the gap closing requires demonstrated renewal-rate parity.",
    ]),
    ("Caveats & limitations", [
        "1. Operating-lease liabilities not included in Net Debt — material for AMZN (~$80B+) and TGT (~$3B+). EV understates economic leverage proportionally.",
        "2. AMZN comparability: AWS distorts every consolidated multiple. A clean retail-segment comparable would require disaggregation not performed here.",
        "3. Membership / ad revenue estimates: where mgmt does not separately disclose, figures are [E] based on prepared commentary at recent earnings calls.",
        "4. This analysis is research, not investment advice. Cross-check against a primary terminal (Bloomberg, CapIQ, FactSet) before any binding decision.",
    ]),
    ("Cross-references", [
        "Companion file: WMT-Competitive-Analysis.pptx (slide deck per `competitive-analysis` skill).",
        "Companion HTML view: WMT-Competitive-Analysis.html (browser-renderable summary).",
        "Build script: build_comps.py — re-run to refresh after updating INPUTS dict.",
    ]),
]

r = 3
for title, lines in notes_blocks:
    c = ws.cell(row=r, column=1, value=title); c.font = BOLD_BLACK; c.fill = LIGHTBLUE_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    for line in lines:
        cc = ws.cell(row=r, column=1, value=line); cc.font = BLACK_FONT; cc.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Inputs"], wb["Operating Metrics"], wb["Valuation"], wb["Retail-Specific"], wb["Notes"]]

wb.save(OUT)
print(f"Wrote {OUT}")
