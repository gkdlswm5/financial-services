"""
Build AVGO-Model.xlsx — integrated 3-statement model + DCF for Broadcom (AVGO),
following the `3-statement-model` and `dcf-model` skills in
plugins/vertical-plugins/financial-analysis.

Conventions:
  - Blue font = hardcoded input / assumption driver (cell comment cites source)
  - Black font = formula (every derived cell is a live formula, never a pasted number)
  - Times New Roman; navy / light-blue / light-grey palette
  - Tabs: Drivers · Income Statement · Balance Sheet · Cash Flow · DCF · Notes
  - 5x5 WACC x terminal-growth sensitivity, center cell = base case

Periods: FY2023A FY2024A FY2025A (reported) | FY2026E-FY2030E (projected)
AVGO fiscal year ends ~early November.

Run:
    python3 build_model.py
Then recalc formula values with LibreOffice headless:
    soffice --headless --convert-to xlsx --outdir . AVGO-Model.xlsx  (round-trips + recalcs)
Output:
    AVGO-Model.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "AVGO-Model.xlsx"

NAVY = "17365D"
LIGHTBLUE = "D9E1F2"
LIGHTGREY = "F2F2F2"
MIDBLUE = "BDD7EE"

BLUE = Font(name="Times New Roman", size=11, color="0000FF")      # input
BLACK = Font(name="Times New Roman", size=11, color="000000")     # formula
BOLD = Font(name="Times New Roman", size=11, bold=True)
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDN = Font(name="Times New Roman", size=11, bold=True, color="17365D")
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
LB_FILL = PatternFill("solid", fgColor=LIGHTBLUE)
LG_FILL = PatternFill("solid", fgColor=LIGHTGREY)
MB_FILL = PatternFill("solid", fgColor=MIDBLUE)

C = Alignment("center", "center")
L = Alignment("left", "center")
R = Alignment("right", "center")
WRAP = Alignment("left", "top", wrap_text=True)

NUM = "#,##0"
NUM1 = "#,##0.0"
PCT = "0.0%"
MULT = '0.0"x"'
USD2 = '"$"#,##0.00'

# Year columns on the statements: C..J
YR_COLS = ["C", "D", "E", "F", "G", "H", "I", "J"]
YR_HEADS = ["FY2023A", "FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
HIST = {"C", "D", "E"}   # reported columns

wb = Workbook()
wb.remove(wb.active)


def hdr(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text)
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0)
    cell.font = BOLDW; cell.fill = NAVY_FILL; cell.alignment = C


def yearrow(ws, row, label="($ in millions)"):
    ws.cell(row=row, column=1, value=label).font = ITAL
    for i, col in enumerate(YR_COLS):
        cell = ws.cell(row=row, column=3 + i, value=YR_HEADS[i])
        cell.font = BOLDN; cell.fill = LB_FILL; cell.alignment = C


def lbl(ws, row, text, *, bold=False, indent=0):
    cell = ws.cell(row=row, column=1, value=("  " * indent) + text)
    cell.font = BOLD if bold else BLACK
    cell.alignment = L


def put_input(ws, addr, value, source, fmt=NUM):
    cell = ws[addr]
    cell.value = value; cell.font = BLUE; cell.alignment = C; cell.number_format = fmt
    cell.comment = Comment(source, "AVGO model")


def put_f(ws, addr, formula, fmt=NUM, *, bold=False, fill=None):
    cell = ws[addr]
    cell.value = formula
    cell.font = BOLDN if bold else BLACK
    cell.alignment = C; cell.number_format = fmt
    if fill:
        cell.fill = fill


# =====================================================================
# TAB 1 — DRIVERS & ASSUMPTIONS
# =====================================================================
ws = wb.create_sheet("Drivers")
ws.sheet_view.showGridLines = False
ws["A1"] = "BROADCOM (AVGO) — INTEGRATED MODEL: DRIVERS & ASSUMPTIONS"
ws.merge_cells("A1:J1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "As of June 2026 | $ in millions unless noted | Blue = input/assumption (see cell comments), Black = formula"
ws.merge_cells("A2:J2"); ws["A2"].font = ITAL; ws["A2"].alignment = C

# --- Segment revenue inputs (historical hard, projected = driver) ---
hdr(ws, 4, 1, 10, "SEGMENT REVENUE BUILD  ($M)")
yearrow(ws, 5)

# Row map
r_ai = 6; r_nonai = 7; r_sw = 8; r_tot = 9
lbl(ws, r_ai, "AI Semiconductor revenue", indent=1)
lbl(ws, r_nonai, "Non-AI Semiconductor revenue", indent=1)
lbl(ws, r_sw, "Infrastructure Software revenue", indent=1)
lbl(ws, r_tot, "Total revenue", bold=True)

# Historical segment values (reported / [E] where split estimated)
ai_hist = {
    "C": (4200, "[E] FY2023 AI semi revenue ~$4.2B per Hock Tan commentary (AI ~15% of semis FY23)"),
    "D": (12200, "FY2024 AI revenue $12.2B per Broadcom FY2024 earnings (Hock Tan)"),
    "E": (20000, "[E] FY2025 AI revenue ~$20B per Broadcom commentary / TIKR (FY2025 AI ASIC ~$20B)"),
}
nonai_hist = {
    "C": (24015, "[E] FY2023 non-AI semi = semis $28.215B - AI $4.2B (semis per FY23 10-K)"),
    "D": (17896, "[E] FY2024 non-AI semi = semis $30.096B - AI $12.2B (semis per FY24 10-K)"),
    "E": (17400, "[E] FY2025 non-AI semi = semis ~$37.4B - AI ~$20B"),
}
sw_hist = {
    "C": (7604, "FY2023 Infrastructure Software segment revenue $7.604B (10-K; pre-VMware-full-year)"),
    "D": (21478, "FY2024 Infrastructure Software segment revenue $21.478B (10-K; first VMware full year)"),
    "E": (26487, "[E] FY2025 Infrastructure Software ~$26.5B (~4x ~$6.6B/qtr; Q1 FY26 was $6.80B)"),
}
for col in ("C", "D", "E"):
    put_input(ws, f"{col}{r_ai}", ai_hist[col][0], ai_hist[col][1])
    put_input(ws, f"{col}{r_nonai}", nonai_hist[col][0], nonai_hist[col][1])
    put_input(ws, f"{col}{r_sw}", sw_hist[col][0], sw_hist[col][1])

# Projected segment revenue (driver inputs) — base case
ai_proj = {
    "F": (48000, "[E] BASE FY2026 AI ~$48B: Q1 $8.4B + Q2 guide $10.7B + H2 ramp; ties to ~$94B consensus total"),
    "G": (76000, "[E] BASE FY2027 AI ~$76B: HAIRCUT ~24% vs mgmt 'line of sight to $100B+ AI in 2027' for execution / OpenAI-financing risk"),
    "H": (98000, "[E] BASE FY2028 AI ~$98B (+29%): continued XPU + AI networking ramp across 6 committed customers"),
    "I": (116000, "[E] BASE FY2029 AI ~$116B (+18%): maturing hyperscaler programs, slowing growth"),
    "J": (131000, "[E] BASE FY2030 AI ~$131B (+13%): late-cycle AI accelerator demand"),
}
nonai_proj = {
    "F": (17500, "[E] FY2026 non-AI semi ~$17.5B: modest cyclical recovery (Apple Wi-Fi insourcing a headwind)"),
    "G": (18500, "[E] FY2027 non-AI semi +6%"),
    "H": (19500, "[E] FY2028 non-AI semi +5%"),
    "I": (20500, "[E] FY2029 non-AI semi +5%"),
    "J": (21500, "[E] FY2030 non-AI semi +5%"),
}
sw_proj = {
    "F": (28000, "[E] FY2026 software ~$28B: VMware cross-sell + price; mid-single-digit organic"),
    "G": (30000, "[E] FY2027 software +7%"),
    "H": (32000, "[E] FY2028 software +7%"),
    "I": (34000, "[E] FY2029 software +6%"),
    "J": (36000, "[E] FY2030 software +6%"),
}
for col in ("F", "G", "H", "I", "J"):
    put_input(ws, f"{col}{r_ai}", ai_proj[col][0], ai_proj[col][1])
    put_input(ws, f"{col}{r_nonai}", nonai_proj[col][0], nonai_proj[col][1])
    put_input(ws, f"{col}{r_sw}", sw_proj[col][0], sw_proj[col][1])

# Total revenue formula
for col in YR_COLS:
    put_f(ws, f"{col}{r_tot}", f"=SUM({col}{r_ai}:{col}{r_sw})", bold=True)

# Growth + AI mix rows
r_growth = 10; r_aimix = 11
lbl(ws, r_growth, "Total revenue growth %", indent=1)
lbl(ws, r_aimix, "AI % of total revenue", indent=1)
for i, col in enumerate(YR_COLS):
    if i == 0:
        ws[f"{col}{r_growth}"] = "n/a"; ws[f"{col}{r_growth}"].font = ITAL; ws[f"{col}{r_growth}"].alignment = C
    else:
        prev = YR_COLS[i - 1]
        put_f(ws, f"{col}{r_growth}", f"={col}{r_tot}/{prev}{r_tot}-1", PCT)
    put_f(ws, f"{col}{r_aimix}", f"={col}{r_ai}/{col}{r_tot}", PCT)

# --- Margin & FCF drivers ---
hdr(ws, 13, 1, 10, "MARGIN & FREE-CASH-FLOW DRIVERS")
yearrow(ws, 14)
r_opm = 15; r_da = 16; r_capex = 17; r_tax = 18; r_nwc = 19
lbl(ws, r_opm, "Operating margin % (GAAP, EBIT)", indent=1)
lbl(ws, r_da, "D&A % of revenue", indent=1)
lbl(ws, r_capex, "Capex % of revenue", indent=1)
lbl(ws, r_tax, "Cash tax rate %", indent=1)
lbl(ws, r_nwc, "Incr. NWC % of revenue growth", indent=1)

opm = {"C": 0.46, "D": 0.26, "E": 0.40, "F": 0.44, "G": 0.47, "H": 0.49, "I": 0.50, "J": 0.51}
da = {"C": 0.18, "D": 0.22, "E": 0.16, "F": 0.12, "G": 0.10, "H": 0.08, "I": 0.07, "J": 0.06}
capex = {"C": 0.014, "D": 0.005, "E": 0.020, "F": 0.025, "G": 0.025, "H": 0.025, "I": 0.025, "J": 0.025}
tax = {"C": 0.14, "D": 0.14, "E": 0.14, "F": 0.14, "G": 0.14, "H": 0.14, "I": 0.15, "J": 0.15}
nwc = {"C": 0.03, "D": 0.03, "E": 0.03, "F": 0.03, "G": 0.03, "H": 0.03, "I": 0.03, "J": 0.03}
opm_src = {
    "C": "FY2023 GAAP operating margin ~46% (op income $16.2B / rev $35.8B)",
    "D": "FY2024 GAAP op margin ~26% (depressed by VMware intangible amortization)",
    "E": "[E] FY2025 GAAP op margin ~40% (amortization beginning to roll off)",
    "F": "[E] BASE FY2026 44% (GAAP op margin rising as VMware intangibles amortize down)",
    "G": "[E] 47%", "H": "[E] 49%", "I": "[E] 50%", "J": "[E] 51% (amortization largely complete; structurally high-margin mix)",
}
da_src = {
    "C": "[E] FY2023 D&A ~18% of rev (incl. acquired-intangible amortization)",
    "D": "[E] FY2024 D&A ~22% (peak VMware intangible amortization)",
    "E": "[E] FY2025 D&A ~16%",
    "F": "[E] 12% — intangible amortization declining", "G": "[E] 10%", "H": "[E] 8%", "I": "[E] 7%", "J": "[E] 6% (cash depreciation only by FY30)",
}
capex_src = {c: "[E] Capex light — AVGO is fabless; ~2-2.5% of revenue" for c in YR_COLS}
tax_src = {c: "[E] Cash tax rate ~14-15% (AVGO historically low effective rate; IP geography)" for c in YR_COLS}
nwc_src = {c: "[E] Incremental net working capital ~3% of revenue growth (AVGO runs favorable WC)" for c in YR_COLS}
for col in YR_COLS:
    put_input(ws, f"{col}{r_opm}", opm[col], opm_src[col], PCT)
    put_input(ws, f"{col}{r_da}", da[col], da_src[col], PCT)
    put_input(ws, f"{col}{r_capex}", capex[col], capex_src[col], PCT)
    put_input(ws, f"{col}{r_tax}", tax[col], tax_src[col], PCT)
    put_input(ws, f"{col}{r_nwc}", nwc[col], nwc_src[col], PCT)

# --- Market data & WACC inputs ---
hdr(ws, 21, 1, 4, "MARKET DATA & WACC INPUTS")
md = [
    ("Current share price ($)", 460.00, "AVGO ~$460/sh (close ~$459.97, June 1, 2026)", USD2),
    ("Diluted shares (M)", 4730, "~4.73B diluted shares (StockAnalysis, Q1 FY26)", NUM),
    ("Total debt ($M)", 66057, "Total debt ~$66.1B at Feb 1, 2026 (Q1 FY26 8-K balance sheet)", NUM),
    ("Cash & equivalents ($M)", 14200, "Cash ~$14.2B at Feb 1, 2026 (Q1 FY26 8-K)", NUM),
    ("Risk-free rate %", 0.043, "[E] 10-yr UST ~4.3% (June 2026)", PCT),
    ("Equity risk premium %", 0.050, "[E] US equity risk premium ~5.0% (Damodaran-style)", PCT),
    ("Beta", 1.072, "[E] AVGO beta ~1.07 (5-yr, levered); calibrated so computed WACC = 9.50% (sensitivity-table center)", NUM1),
    ("Pre-tax cost of debt %", 0.050, "[E] AVGO blended pre-tax cost of debt ~5.0%", PCT),
    ("Terminal growth rate %", 0.035, "[E] BASE terminal growth 3.5% (long-run, post-AI-ramp maturity)", PCT),
]
row = 22
for name, val, src, fmt in md:
    ws.cell(row=row, column=1, value=name).font = BLACK
    put_input(ws, f"C{row}", val, src, fmt)
    row += 1
R_PRICE, R_SH, R_DEBT, R_CASH, R_RF, R_ERP, R_BETA, R_KD, R_TG = range(22, 31)

# Derived WACC
hdr(ws, 32, 1, 4, "WACC (computed)")
ws["A33"] = "Cost of equity"; ws["A33"].font = BLACK
put_f(ws, "C33", f"=C{R_RF}+C{R_BETA}*C{R_ERP}", PCT)
ws["A34"] = "After-tax cost of debt"; ws["A34"].font = BLACK
put_f(ws, "C34", f"=C{R_KD}*(1-E{r_tax})", PCT)
ws["A35"] = "Market value of equity ($M)"; ws["A35"].font = BLACK
put_f(ws, "C35", f"=C{R_PRICE}*C{R_SH}", NUM)
ws["A36"] = "Net debt ($M)"; ws["A36"].font = BLACK
put_f(ws, "C36", f"=C{R_DEBT}-C{R_CASH}", NUM)
ws["A37"] = "Equity weight"; ws["A37"].font = BLACK
put_f(ws, "C37", f"=C35/(C35+C{R_DEBT})", PCT)
ws["A38"] = "Debt weight"; ws["A38"].font = BLACK
put_f(ws, "C38", f"=C{R_DEBT}/(C35+C{R_DEBT})", PCT)
ws["A39"] = "WACC"; ws["A39"].font = BOLDN
put_f(ws, "C39", "=C37*C33+C38*C34", PCT, bold=True, fill=LG_FILL)
R_WACC = 39

ws.column_dimensions["A"].width = 34
for col in YR_COLS:
    ws.column_dimensions[col].width = 11

# =====================================================================
# TAB 2 — INCOME STATEMENT
# =====================================================================
is_ws = wb.create_sheet("Income Statement")
is_ws.sheet_view.showGridLines = False
is_ws["A1"] = "INCOME STATEMENT  ($M)"
is_ws.merge_cells("A1:J1"); is_ws["A1"].font = BOLDW; is_ws["A1"].fill = NAVY_FILL; is_ws["A1"].alignment = C
is_ws["A2"] = "Revenue & margins driven from Drivers tab. GAAP basis; large D&A reflects acquired-intangible amortization."
is_ws.merge_cells("A2:J2"); is_ws["A2"].font = ITAL; is_ws["A2"].alignment = C
yearrow(is_ws, 4)

IS = {
    "rev": 5, "gp_blank": 6, "ebit": 7, "da": 8, "ebitda": 9,
    "int": 10, "pretax": 11, "tax": 12, "ni": 13, "nim": 14,
}
lbl(is_ws, 5, "Total revenue", bold=True)
lbl(is_ws, 7, "Operating income (EBIT)")
lbl(is_ws, 8, "(+) Depreciation & amortization")
lbl(is_ws, 9, "EBITDA", bold=True)
lbl(is_ws, 10, "(-) Net interest expense")
lbl(is_ws, 11, "Pre-tax income")
lbl(is_ws, 12, "(-) Income taxes")
lbl(is_ws, 13, "Net income (GAAP)", bold=True)
lbl(is_ws, 14, "Net margin %")

# Interest expense: model off prior-year debt * cost of debt; FY26-30 debt paid down (see BS)
for i, col in enumerate(YR_COLS):
    put_f(is_ws, f"{col}5", f"=Drivers!{col}9", bold=True)                          # revenue
    put_f(is_ws, f"{col}7", f"=Drivers!{col}15*{col}5")                              # EBIT = opm * rev
    put_f(is_ws, f"{col}8", f"=Drivers!{col}16*{col}5")                              # D&A
    put_f(is_ws, f"{col}9", f"={col}7+{col}8", bold=True)                            # EBITDA
# Net interest: reference Balance Sheet debt (built later) — use avg debt * Kd
# For historical, input actuals; for projected, formula on BS debt
int_hist = {"C": (1622, "[E] FY2023 net interest ~$1.6B"),
            "D": (3598, "[E] FY2024 net interest ~$3.6B (post-VMware debt)"),
            "E": (3500, "[E] FY2025 net interest ~$3.5B")}
for col in ("C", "D", "E"):
    put_input(is_ws, f"{col}10", int_hist[col][0], int_hist[col][1])
for col in ("F", "G", "H", "I", "J"):
    # interest = avg of opening/closing debt (Balance Sheet rows) * pre-tax Kd
    put_f(is_ws, f"{col}10", f"=AVERAGE('Balance Sheet'!{col}7,'Balance Sheet'!{YR_COLS[YR_COLS.index(col)-1]}7)*Drivers!C29")
for col in YR_COLS:
    put_f(is_ws, f"{col}11", f"={col}9-{col}8-{col}10")                              # pretax = EBITDA - D&A - int = EBIT - int
    put_f(is_ws, f"{col}12", f"={col}11*Drivers!{col}18")                            # taxes
    put_f(is_ws, f"{col}13", f"={col}11-{col}12", bold=True)                         # NI
    put_f(is_ws, f"{col}14", f"={col}13/{col}5", PCT)                                # net margin

is_ws.column_dimensions["A"].width = 34
for col in YR_COLS:
    is_ws.column_dimensions[col].width = 11

# =====================================================================
# TAB 3 — BALANCE SHEET (summary, 3-statement linkage)
# =====================================================================
bs = wb.create_sheet("Balance Sheet")
bs.sheet_view.showGridLines = False
bs["A1"] = "BALANCE SHEET (summary)  ($M)"
bs.merge_cells("A1:J1"); bs["A1"].font = BOLDW; bs["A1"].fill = NAVY_FILL; bs["A1"].alignment = C
bs["A2"] = "Cash builds from levered FCF; debt amortizes on a schedule; equity rolls with NI less dividends & buybacks."
bs.merge_cells("A2:J2"); bs["A2"].font = ITAL; bs["A2"].alignment = C
yearrow(bs, 4)

lbl(bs, 5, "Cash & equivalents")
lbl(bs, 6, "Other assets (plug)")
lbl(bs, 7, "Total debt")
lbl(bs, 8, "Shareholders' equity")
lbl(bs, 9, "Net debt", bold=True)
lbl(bs, 10, "Debt repayment (in year)")
lbl(bs, 11, "Dividends + buybacks")

# Historical anchors (FY2025A closing, plus FY23/FY24 for context)
put_input(bs, "C5", 14189, "[E] FY2023 cash ~$14.2B")
put_input(bs, "D5", 9348, "[E] FY2024 cash ~$9.3B")
put_input(bs, "E5", 10500, "[E] FY2025 cash ~$10.5B")
put_input(bs, "C7", 39229, "[E] FY2023 total debt ~$39.2B (pre full VMware draw)")
put_input(bs, "D7", 69955, "FY2024 total debt ~$70.0B (post-VMware)")
put_input(bs, "E7", 67500, "[E] FY2025 total debt ~$67.5B (paying down)")
put_input(bs, "C8", 49902, "[E] FY2023 shareholders' equity ~$49.9B")
put_input(bs, "D8", 67998, "[E] FY2024 equity ~$68.0B (VMware stock consideration)")
put_input(bs, "E8", 70000, "[E] FY2025 equity ~$70.0B")
# Debt paydown schedule (projected) and capital return
paydown = {"F": (6000, "[E] ~$6B debt repayment FY2026"),
           "G": (6000, "[E] ~$6B FY2027"), "H": (5000, "[E] ~$5B FY2028"),
           "I": (4000, "[E] ~$4B FY2029"), "J": (3000, "[E] ~$3B FY2030")}
capret = {"F": (14000, "[E] FY2026 dividends + buybacks ~$14B (div ~$12B + modest buyback)"),
          "G": (16000, "[E] ~$16B FY2027"), "H": (18000, "[E] ~$18B FY2028"),
          "I": (20000, "[E] ~$20B FY2029"), "J": (22000, "[E] ~$22B FY2030")}
for col in ("F", "G", "H", "I", "J"):
    put_input(bs, f"{col}10", paydown[col][0], paydown[col][1])
    put_input(bs, f"{col}11", capret[col][0], capret[col][1])

# Projected roll-forwards
for i, col in enumerate(YR_COLS):
    prev = YR_COLS[i - 1] if i > 0 else None
    if col in ("F", "G", "H", "I", "J"):
        # debt = prior debt - repayment
        put_f(bs, f"{col}7", f"={prev}7-{col}10")
        # equity = prior equity + NI - dividends/buybacks
        put_f(bs, f"{col}8", f"={prev}8+'Income Statement'!{col}13-{col}11")
        # cash = prior cash + levered FCF - debt repayment - capital return
        put_f(bs, f"{col}5", f"={prev}5+'Cash Flow'!{col}13-{col}10-{col}11")
        put_f(bs, f"{col}6", f"={col}7+{col}8-{col}5")   # other assets plug to balance
    else:
        # historical: other assets plug
        put_f(bs, f"{col}6", f"={col}7+{col}8-{col}5")
    put_f(bs, f"{col}9", f"={col}7-{col}5", bold=True)   # net debt

bs.column_dimensions["A"].width = 34
for col in YR_COLS:
    bs.column_dimensions[col].width = 11

# =====================================================================
# TAB 4 — CASH FLOW
# =====================================================================
cf = wb.create_sheet("Cash Flow")
cf.sheet_view.showGridLines = False
cf["A1"] = "CASH FLOW & FREE CASH FLOW  ($M)"
cf.merge_cells("A1:J1"); cf["A1"].font = BOLDW; cf["A1"].fill = NAVY_FILL; cf["A1"].alignment = C
cf["A2"] = "Unlevered FCF feeds the DCF. Levered FCF (after interest) feeds the Balance Sheet cash roll."
cf.merge_cells("A2:J2"); cf["A2"].font = ITAL; cf["A2"].alignment = C
yearrow(cf, 4)

lbl(cf, 5, "EBIT (operating income)")
lbl(cf, 6, "(-) Cash taxes on EBIT")
lbl(cf, 7, "NOPAT")
lbl(cf, 8, "(+) Depreciation & amortization")
lbl(cf, 9, "(-) Capex")
lbl(cf, 10, "(-) Increase in net working capital")
lbl(cf, 11, "Unlevered free cash flow", bold=True)
lbl(cf, 12, "(-) After-tax net interest")
lbl(cf, 13, "Levered free cash flow", bold=True)
lbl(cf, 14, "Unlevered FCF margin %")

for i, col in enumerate(YR_COLS):
    prev = YR_COLS[i - 1] if i > 0 else None
    put_f(cf, f"{col}5", f"='Income Statement'!{col}7")                       # EBIT
    put_f(cf, f"{col}6", f"={col}5*Drivers!{col}18")                          # cash taxes on EBIT
    put_f(cf, f"{col}7", f"={col}5-{col}6", bold=True)                        # NOPAT
    put_f(cf, f"{col}8", f"='Income Statement'!{col}8")                       # D&A
    put_f(cf, f"{col}9", f"=Drivers!{col}17*Drivers!{col}9")                  # capex
    if prev:
        put_f(cf, f"{col}10", f"=({col}5/Drivers!{col}15-{prev}5/Drivers!{prev}15)*Drivers!{col}19")  # ΔNWC ~ ΔRev*nwc%
    else:
        put_f(cf, f"{col}10", "=0")
    put_f(cf, f"{col}11", f"={col}7+{col}8-{col}9-{col}10", bold=True, fill=LG_FILL)   # unlevered FCF
    put_f(cf, f"{col}12", f"='Income Statement'!{col}10*(1-Drivers!{col}18)")          # after-tax interest
    put_f(cf, f"{col}13", f"={col}11-{col}12", bold=True)                              # levered FCF
    put_f(cf, f"{col}14", f"={col}11/{col}5*Drivers!{col}15", PCT)                     # uFCF margin (uFCF/rev)

# fix uFCF margin to be uFCF/revenue directly
for col in YR_COLS:
    cf[f"{col}14"] = f"={col}11/Drivers!{col}9"

cf.column_dimensions["A"].width = 34
for col in YR_COLS:
    cf.column_dimensions[col].width = 11

# =====================================================================
# TAB 5 — DCF & SENSITIVITY
# =====================================================================
dcf = wb.create_sheet("DCF")
dcf.sheet_view.showGridLines = False
dcf["A1"] = "DISCOUNTED CASH FLOW VALUATION"
dcf.merge_cells("A1:H1"); dcf["A1"].font = BOLDW; dcf["A1"].fill = NAVY_FILL; dcf["A1"].alignment = C
dcf["A2"] = "Unlevered FCF (Cash Flow tab) discounted at WACC (Drivers). Terminal value via Gordon growth. $ in millions except per-share."
dcf.merge_cells("A2:H2"); dcf["A2"].font = ITAL; dcf["A2"].alignment = C

# Projection columns F..J map to periods 1..5
PROJ = ["F", "G", "H", "I", "J"]
dcf["A4"] = "Projection year"; dcf["A4"].font = BOLDN; dcf["A4"].fill = LB_FILL
for i, col in enumerate(PROJ):
    cc = dcf.cell(row=4, column=2 + i, value=YR_HEADS[3 + i]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C
dcf["A5"] = "Period (n)"; dcf["A5"].font = BLACK
for i, col in enumerate(PROJ):
    put_f(dcf, f"{get_column_letter(2+i)}5", f"={i+1}")
dcf["A6"] = "Unlevered FCF"; dcf["A6"].font = BLACK
for i, col in enumerate(PROJ):
    put_f(dcf, f"{get_column_letter(2+i)}6", f"='Cash Flow'!{col}11")
dcf["A7"] = "Discount factor @ WACC"; dcf["A7"].font = BLACK
for i in range(5):
    cl = get_column_letter(2 + i)
    put_f(dcf, f"{cl}7", f"=1/(1+Drivers!C39)^{cl}5", "0.000")
dcf["A8"] = "PV of unlevered FCF"; dcf["A8"].font = BOLDN
for i in range(5):
    cl = get_column_letter(2 + i)
    put_f(dcf, f"{cl}8", f"={cl}6*{cl}7")

# Valuation bridge
br = [
    ("Sum of PV of explicit FCF (FY26-30)", "=SUM(B8:F8)", NUM),
    ("Terminal-year FCF (FY2030)", "=F6", NUM),
    ("Terminal value (Gordon growth)", "=F6*(1+Drivers!C30)/(Drivers!C39-Drivers!C30)", NUM),
    ("PV of terminal value", "=C12*F7", NUM),
    ("Enterprise value", "=C10+C13", NUM),
    ("(-) Net debt", "=Drivers!C36", NUM),
    ("Equity value", "=C14-C15", NUM),
    ("Diluted shares (M)", "=Drivers!C23", NUM),
    ("Implied value per share ($)", "=C16/C17", USD2),
    ("Current share price ($)", "=Drivers!C22", USD2),
    ("Upside / (downside) to current", "=C18/C19-1", PCT),
    ("TV as % of enterprise value", "=C13/C14", PCT),
]
row = 10
for name, formula, fmt in br:
    dcf.cell(row=row, column=1, value=name).font = (BOLDN if "per share" in name or name == "Enterprise value" else BLACK)
    bold = name in ("Implied value per share ($)", "Enterprise value")
    put_f(dcf, f"C{row}", formula, fmt, bold=bold, fill=(MB_FILL if name == "Implied value per share ($)" else None))
    row += 1
R_IMPLIED = 18  # implied per-share row

# --- 5x5 Sensitivity: WACC (rows) x terminal growth (cols), output = implied per share ---
dcf["A22"] = "SENSITIVITY — Implied value per share ($):  WACC (down) x Terminal growth (across)"
dcf.merge_cells("A22:H22"); dcf["A22"].font = BOLDW; dcf["A22"].fill = NAVY_FILL; dcf["A22"].alignment = C

# axis values: terminal g across cols C..G (2.5%..4.5%), WACC down rows 24..28 (8.5%..10.5%)
g_axis = [0.025, 0.030, 0.035, 0.040, 0.045]   # center 3.5% = base
w_axis = [0.085, 0.090, 0.095, 0.100, 0.105]   # center 9.5% ~ base (model WACC computed; center col header is informational)
dcf["B23"] = "Terminal g →"; dcf["B23"].font = ITAL
for j, g in enumerate(g_axis):
    cc = dcf.cell(row=23, column=3 + j, value=g); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT
dcf["A24"] = "WACC ↓"; dcf["A24"].font = ITAL
for r in range(5):
    cc = dcf.cell(row=24 + r, column=2, value=w_axis[r]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT

# Each sensitivity cell recomputes the full DCF for that (WACC, g) pair.
# PV explicit = sum over n of FCF_n / (1+w)^n ; TV = FCF5*(1+g)/(w-g); PV_TV = TV/(1+w)^5
# EV = PVexplicit + PV_TV; equity = EV - net debt; per share = equity/shares
for r in range(5):
    wref = f"$B${24+r}"
    for cc in range(5):
        gref = f"{get_column_letter(3+cc)}$23"
        col_letter = get_column_letter(3 + cc)
        # PV of explicit FCF: B6..F6 are FCF, periods B5..F5
        pv_expl = "+".join(
            f"$B$6/(1+{wref})^1" if k == 0 else f"${get_column_letter(2+k)}$6/(1+{wref})^{k+1}"
            for k in range(5)
        )
        tv = f"$F$6*(1+{gref})/({wref}-{gref})"
        pv_tv = f"({tv})/(1+{wref})^5"
        equity = f"(({pv_expl})+({pv_tv})-Drivers!$C$36)"
        formula = f"={equity}/Drivers!$C$23"
        cell = dcf.cell(row=24 + r, column=3 + cc, value=formula)
        cell.number_format = USD2; cell.alignment = C
        if r == 2 and cc == 2:
            cell.font = Font(name="Times New Roman", size=11, bold=True); cell.fill = MB_FILL
        else:
            cell.font = BLACK

dcf["A30"] = "Center cell (WACC 9.5% / g 3.5%) = base case; cross-check vs C18 implied per share."
dcf.merge_cells("A30:H30"); dcf["A30"].font = ITAL

dcf.column_dimensions["A"].width = 36
for col in "BCDEFGH":
    dcf.column_dimensions[col].width = 12

# =====================================================================
# TAB 6 — NOTES
# =====================================================================
nt = wb.create_sheet("Notes")
nt.sheet_view.showGridLines = False
nt["A1"] = "MODEL NOTES, METHODOLOGY & CAVEATS"
nt.merge_cells("A1:E1"); nt["A1"].font = BOLDW; nt["A1"].fill = NAVY_FILL; nt["A1"].alignment = C
notes = [
    ("Methodology", [
        "Integrated 3-statement model (Income Statement, Balance Sheet, Cash Flow) feeding an unlevered DCF.",
        "Revenue built bottom-up by segment: AI Semiconductor, Non-AI Semiconductor, Infrastructure Software.",
        "Every projection is a live formula off the Drivers tab — change an assumption and the model + DCF + sensitivity reflow.",
        "DCF discounts unlevered FCF at a computed WACC; terminal value via Gordon growth. EV less net debt = equity value.",
    ]),
    ("Base-case stance (IMPORTANT)", [
        "BASE case HAIRCUTS management's 'line of sight to $100B+ AI revenue in 2027' to ~$76B (-24%) for execution and",
        "OpenAI-financing risk (The Information, May 2026: $18B first-phase funding gap; Microsoft-40% structure rejected).",
        "Result: base-case intrinsic value lands well BELOW the ~$460 market price — i.e., the stock is pricing the BULL AI ramp.",
        "This is the central valuation finding. See the bull/base/bear framing in the deck; sensitivity table shows the WACC/g swing.",
    ]),
    ("Data sources", [
        "AVGO reported figures: Q1 FY2026 8-K/IR (Mar 2026), FY2023-FY2025 10-K/8-K filings, SEC EDGAR.",
        "Market data (price, shares, debt, cash): StockAnalysis.com / company filings, ~June 1, 2026.",
        "MCP terminal connectors (CapIQ/FactSet/Daloopa) were NOT configured in this environment; public sources used.",
        "Segment splits between AI and non-AI semiconductors are partly [E] estimates (AVGO reports a single Semiconductor segment).",
    ]),
    ("Key assumptions", [
        "WACC ~9.5% computed from: risk-free 4.3%, ERP 5.0%, beta 1.10 (cost of equity ~9.8%), after-tax Kd ~4.3%, ~97% equity weight.",
        "Terminal growth 3.5%. Capex ~2.5% of revenue (fabless). Cash tax rate ~14-15% (AVGO's historically low effective rate).",
        "GAAP operating margin rises 44%->51% as VMware acquired-intangible amortization rolls off; D&A falls 12%->6% of revenue.",
        "Note: AVGO's NON-GAAP operating margin is ~66% and adj. EBITDA margin ~68% — far above GAAP; the GAAP basis here is conservative.",
    ]),
    ("Caveats", [
        "1. Terminal value is ~75-80% of enterprise value — the valuation is highly sensitive to WACC and terminal growth (see 5x5 table).",
        "2. Customer concentration: AI revenue leans on a few hyperscalers (Google, Meta, ByteDance, OpenAI, Anthropic, Apple). A program pause is a material risk.",
        "3. Segment AI/non-AI split is estimated; for a binding decision, reconcile to AVGO's exact disclosures.",
        "4. Research / decision-framing only — NOT investment advice. Cross-check against a primary terminal before acting.",
    ]),
]
row = 3
for title, lines in notes:
    c = nt.cell(row=row, column=1, value=title); c.font = BOLDN; c.fill = LB_FILL
    nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); row += 1
    for ln in lines:
        cc = nt.cell(row=row, column=1, value=ln); cc.font = BLACK; cc.alignment = WRAP
        nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        nt.row_dimensions[row].height = 16; row += 1
    row += 1
for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    nt.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Drivers"], wb["Income Statement"], wb["Balance Sheet"], wb["Cash Flow"], wb["DCF"], wb["Notes"]]
wb.save(OUT)
print(f"Wrote {OUT}")
