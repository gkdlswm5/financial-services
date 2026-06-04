"""
Build AVGO-Comps-Analysis.xlsx — comparable company analysis for Broadcom (AVGO),
following the `comps-analysis` skill in plugins/vertical-plugins/financial-analysis.

Peer framing (the user's "second framing"): semiconductor cohort + one software anchor.
  Semis peers (stat set): NVDA, AMD, MRVL, QCOM, TXN
  Software anchor (memo, excluded from semis stats): ORCL  — proxy for the VMware (~40%) segment

Conventions:
  - Blue = hardcoded input (cell comment cites source/assumption); black = formula
  - Stat block (Max/75/Median/25/Min) computed over the 5 SEMIS peers only
  - AVGO shown as target; ORCL shown as a software-anchor memo row
  - Both TTM and forward P/E (forward is the cleaner cross-peer gauge given GAAP distortions)
  - Tabs: Inputs · Operating Metrics · Valuation · Growth & AI Exposure · Notes
Valuation date: ~June 1, 2026.  $ in millions except per-share & ratios.

Run: python3 build_comps.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "AVGO-Comps-Analysis.xlsx"

NAVY = "17365D"; LIGHTBLUE = "D9E1F2"; LIGHTGREY = "F2F2F2"
BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDB = Font(name="Times New Roman", size=12, bold=True)
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
LB_FILL = PatternFill("solid", fgColor=LIGHTBLUE)
LG_FILL = PatternFill("solid", fgColor=LIGHTGREY)
C = Alignment("center", "center"); L = Alignment("left", "center")
WRAP = Alignment("left", "top", wrap_text=True)

# Row order: AVGO (target), 5 semis peers, ORCL (software anchor memo)
PEERS = [
    ("Broadcom",            "AVGO",  "target"),
    ("NVIDIA",              "NVDA",  "semi"),
    ("AMD",                 "AMD",   "semi"),
    ("Marvell Technology",  "MRVL",  "semi"),
    ("Qualcomm",            "QCOM",  "semi"),
    ("Texas Instruments",   "TXN",   "semi"),
    ("Oracle",              "ORCL",  "software"),
]

# All $ in millions. Valuation date ~June 1, 2026.
INPUTS = {
    "Broadcom": {
        "mkt_cap": (2278000, "AVGO ~$481.62/sh post-Q2 FY26 (+4.7% on print) x ~4.73B sh = ~$2.28T (June 3, 2026)"),
        "net_debt": (49000, "[E] Net debt ~$49B at May 3, 2026 (debt ~$65B - cash ~$16B; Q2 FCF $10.3B added)"),
        "revenue": (75470, "TTM revenue $75.47B: Q3-Q4 FY25 ($15.95B+$18.02B) + Q1-Q2 FY26 ($19.3B+$22.2B)"),
        "growth": (0.48, "Q2 FY26 revenue +48% YoY (Broadcom Q2 FY26 IR, June 3 2026); Q3 guide +84%"),
        "gm": (0.68, "[E] GAAP gross margin ~68% (non-GAAP ~77%); Q2 FY26"),
        "ebitda": (50500, "[E] TTM GAAP EBITDA ~$50.5B (adj EBITDA TTM ~$54B; Q2 adj EBITDA $15.2B / 69% margin)"),
        "ni": (28000, "[E] TTM GAAP net income ~$28.0B (Q2 GAAP NI $9.3B; non-GAAP $12.1B)"),
        "fwd_pe": (29.0, "[E] Forward P/E ~29x post-Q2 (consensus FY26-27 EPS upward revision absorbing the +48%)"),
    },
    "NVIDIA": {
        "mkt_cap": (5460000, "NVDA ~$225.6/sh, June 2 2026 (~$5.46T) (companiesmarketcap)"),
        "net_debt": (-4700, "[E] Net cash: debt $8.47B - cash&equiv $13.2B (Q1 FY27 10-Q); much larger incl. marketable securities"),
        "revenue": (253500, "TTM revenue ~$253.5B (GuruFocus; consistent w/ Q1 FY27 $81.6B run-rate)"),
        "growth": (0.85, "Q1 FY27 (Apr 2026) revenue +85% YoY (NVIDIA newsroom)"),
        "gm": (0.749, "GAAP gross margin 74.9% Q1 FY27"),
        "ebitda": (174400, "TTM EBITDA ~$174.4B (valueinvesting.io)"),
        "ni": (159600, "TTM net income ~$159.6B (Macrotrends)"),
        "fwd_pe": (26.0, "[E] Forward P/E ~26x (consensus)"),
    },
    "AMD": {
        "mkt_cap": (839000, "AMD ~$510/sh, June 2 2026 (~$839B) (companiesmarketcap)"),
        "net_debt": (-9100, "Net cash ~$9.1B: cash+ST inv $12.35B - debt $3.22B (Q1 2026 10-Q)"),
        "revenue": (37450, "TTM revenue $37.45B (Macrotrends)"),
        "growth": (0.38, "Q1 2026 revenue +38% YoY (8-K); Data Center $5.8B +57%"),
        "gm": (0.53, "GAAP gross margin ~53% Q1 2026"),
        "ebitda": (8090, "TTM EBITDA ~$8.09B (GuruFocus)"),
        "ni": (5000, "[E] TTM net income ~$5.0B (FY25 $4.3B + ramp; GAAP)"),
        "fwd_pe": (33.0, "[E] Forward P/E ~33x (consensus)"),
    },
    "Marvell Technology": {
        "mkt_cap": (192000, "MRVL ~$291/sh, June 2 2026 (~$192B); +51% on Jensen Huang endorsement"),
        "net_debt": (1430, "Net debt ~$1.43B: cash $3.84B - debt $5.28B (StockAnalysis)"),
        "revenue": (8600, "[E] TTM revenue ~$8.6B (Q1 FY27 $2.418B run-rate; FY27 guide ~$11.5B)"),
        "growth": (0.28, "Q1 FY27 revenue +28% YoY; Data Center 76% of revenue"),
        "gm": (0.521, "GAAP gross margin 52.1% (non-GAAP 58.9%) Q1 FY27"),
        "ebitda": (3000, "[E] TTM EBITDA ~$3.0B (derived from leverage ratios / EV-EBITDA)"),
        "ni": (300, "[E] TTM GAAP net income ~$0.3B (depressed by amortization; non-GAAP MRQ $718M)"),
        "fwd_pe": (48.0, "Forward P/E ~48x (StockAnalysis)"),
    },
    "Qualcomm": {
        "mkt_cap": (215000, "QCOM ~$181/sh mid-May 2026 (~$215B); date-sensitive (range $137-254B)"),
        "net_debt": (5500, "Net debt ~$5.5B: cash+ST sec $9.8B - debt $15.27B (Q2 FY26 10-Q)"),
        "revenue": (44490, "TTM revenue ~$44.49B (Macrotrends)"),
        "growth": (-0.03, "[E] Q2 FY26 revenue ~-3% YoY (handset softness; auto/IoT grew)"),
        "gm": (0.54, "GAAP gross margin ~54% Q2 FY26"),
        "ebitda": (13956, "TTM EBITDA ~$13.96B (GuruFocus)"),
        "ni": (9900, "[E] TTM net income ~$9.9B underlying (excl. one-time $5.7B Q2 tax benefit)"),
        "fwd_pe": (14.0, "[E] Forward P/E ~14x (consensus; cheap on mature-semi de-rating)"),
    },
    "Texas Instruments": {
        "mkt_cap": (287000, "TXN ~$287B mid-May 2026 (Macrotrends); date-sensitive"),
        "net_debt": (8000, "Net debt ~$8B: cash+ST inv $5.1B - debt $13.05B (Q1 2026 10-Q)"),
        "revenue": (18440, "TTM revenue ~$18.44B (StockAnalysis)"),
        "growth": (0.19, "Q1 2026 revenue +19% YoY (industrial + data center led)"),
        "gm": (0.58, "GAAP gross margin ~58% Q1 2026"),
        "ebitda": (8821, "TTM EBITDA ~$8.82B (GuruFocus)"),
        "ni": (5370, "TTM net income ~$5.37B (Macrotrends)"),
        "fwd_pe": (30.0, "[E] Forward P/E ~30x (consensus; earnings cyclically depressed by capex cycle)"),
    },
    "Oracle": {
        "mkt_cap": (640000, "[E] ORCL ~$640B early June 2026 (rose from ~$550B May toward ~$718B on OCI momentum)"),
        "net_debt": (96000, "Net debt ~$96B: debt ~$108B - cash ~$11B (Q3 FY26; +$30B raised Feb 2026 for AI buildout)"),
        "revenue": (64080, "TTM revenue ~$64.08B (Macrotrends)"),
        "growth": (0.22, "Q3 FY26 revenue +22% YoY USD; Cloud +44%, OCI +84%; RPO ~$553B (+325%)"),
        "gm": (0.70, "[E] Blended gross margin ~70% (declining as low-margin OCI scales; OCI GM ~32%)"),
        "ebitda": (30618, "TTM EBITDA ~$30.62B (GuruFocus)"),
        "ni": (16200, "TTM net income ~$16.2B (+33% YoY) (Macrotrends)"),
        "fwd_pe": (28.0, "[E] Forward P/E ~28x (consensus; AI-capex growth premium)"),
    },
}

wb = Workbook(); wb.remove(wb.active)
labels = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]


def section(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text)
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0); cell.font = BOLDW; cell.fill = NAVY_FILL; cell.alignment = C


def colhead(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text); cell.font = BOLDB; cell.fill = LB_FILL; cell.alignment = C


def put_input(ws, r, c, val, src, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=val); cell.font = BLUE; cell.alignment = C; cell.number_format = fmt
    cell.comment = Comment(src, "AVGO comps")


def put_f(ws, r, c, formula, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=formula); cell.font = BLACK; cell.alignment = C; cell.number_format = fmt


# Semis-peer rows for stats: NVDA..TXN are data rows 8..12 (AVGO=7 target, ORCL=13 memo)
SEMI_FIRST, SEMI_LAST = 8, 12


def stat_block(ws, start_row, label_col, cols_fmt):
    """cols_fmt: dict col->fmt for the columns to compute stats on."""
    for i, lab in enumerate(labels):
        r = start_row + i
        c = ws.cell(row=r, column=label_col, value=lab); c.font = BOLDB; c.fill = LG_FILL; c.alignment = L
        # fill the label-col gap cells with grey
        for col in range(label_col + 1, max(cols_fmt) + 1):
            cell = ws.cell(row=r, column=col); cell.fill = LG_FILL
        for col, fmt in cols_fmt.items():
            cl = get_column_letter(col); rng = f"{cl}{SEMI_FIRST}:{cl}{SEMI_LAST}"
            f = {"Maximum": f"=MAX({rng})", "75th Percentile": f"=QUARTILE({rng},3)",
                 "Median": f"=MEDIAN({rng})", "25th Percentile": f"=QUARTILE({rng},1)",
                 "Minimum": f"=MIN({rng})"}[lab]
            cell = ws.cell(row=r, column=col, value=f); cell.font = BOLDB; cell.fill = LG_FILL
            cell.alignment = C; cell.number_format = fmt


# =====================================================================
# TAB 1 — INPUTS
# =====================================================================
ws = wb.create_sheet("Inputs"); ws.sheet_view.showGridLines = False
ws["A1"] = "AI SEMICONDUCTORS & INFRASTRUCTURE SOFTWARE — COMPARABLE COMPANY ANALYSIS"
ws.merge_cells("A1:K1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = " • ".join(f"{n} ({t})" for n, t, _ in PEERS)
ws.merge_cells("A2:K2"); ws["A2"].font = BOLDB; ws["A2"].alignment = C
ws["A3"] = "As of June 3, 2026 (post-AVGO Q2 FY26 print) | $ in millions except ratios | Stats over 5 semis peers (NVDA, AMD, MRVL, QCOM, TXN); AVGO=target, ORCL=software anchor (memo)"
ws.merge_cells("A3:K3"); ws["A3"].font = ITAL; ws["A3"].alignment = C

section(ws, 5, 1, 11, "RAW INPUTS — cell comments cite source / assumption")
heads = ["Company", "Ticker", "Mkt Cap", "Net Debt", "Revenue (TTM)", "Rev Growth %",
         "Gross Margin %", "EBITDA (TTM)", "Net Income (TTM)", "Fwd P/E", "Notes"]
for i, h in enumerate(heads, start=1):
    colhead(ws, 6, i, h)
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; d = INPUTS[name]
    cn = ws.cell(row=r, column=1, value=name); cn.font = BOLDB
    ws.cell(row=r, column=2, value=tic).font = BLACK
    put_input(ws, r, 3, d["mkt_cap"][0], d["mkt_cap"][1])
    put_input(ws, r, 4, d["net_debt"][0], d["net_debt"][1])
    put_input(ws, r, 5, d["revenue"][0], d["revenue"][1])
    put_input(ws, r, 6, d["growth"][0], d["growth"][1], "0.0%")
    put_input(ws, r, 7, d["gm"][0], d["gm"][1], "0.0%")
    put_input(ws, r, 8, d["ebitda"][0], d["ebitda"][1])
    put_input(ws, r, 9, d["ni"][0], d["ni"][1])
    put_input(ws, r, 10, d["fwd_pe"][0], d["fwd_pe"][1], '0.0"x"')
    tag = "TARGET" if kind == "target" else ("SOFTWARE ANCHOR (memo)" if kind == "software" else "")
    ws.cell(row=r, column=11, value=tag).font = ITAL
for i, w in enumerate([20, 9, 13, 11, 14, 13, 14, 14, 16, 10, 22], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# TAB 2 — OPERATING METRICS
# =====================================================================
ws = wb.create_sheet("Operating Metrics"); ws.sheet_view.showGridLines = False
ws["A1"] = "OPERATING STATISTICS & FINANCIAL METRICS"
ws.merge_cells("A1:F1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "Formulas reference Inputs tab. Stats over the 5 semis peers only."
ws.merge_cells("A2:F2"); ws["A2"].font = ITAL; ws["A2"].alignment = C
oh = ["Company", "Revenue (TTM)", "Rev Growth (YoY)", "Gross Margin", "EBITDA (TTM)", "EBITDA Margin"]
for i, h in enumerate(oh, start=1):
    colhead(ws, 6, i, h)
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; src = 7 + idx
    cn = ws.cell(row=r, column=1, value=name); cn.font = BOLDB
    put_f(ws, r, 2, f"=Inputs!E{src}")
    put_f(ws, r, 3, f"=Inputs!F{src}", "0.0%")
    put_f(ws, r, 4, f"=Inputs!G{src}", "0.0%")
    put_f(ws, r, 5, f"=Inputs!H{src}")
    put_f(ws, r, 6, f"=Inputs!H{src}/Inputs!E{src}", "0.0%")
stat_block(ws, 14, 1, {3: "0.0%", 4: "0.0%", 6: "0.0%"})
# clear size cols (2,5) in stat rows already grey via stat_block gap fill? ensure
for i in range(5):
    for col in (2, 5):
        ws.cell(row=14 + i, column=col).fill = LG_FILL
for i, w in enumerate([20, 14, 16, 14, 14, 14], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# TAB 3 — VALUATION
# =====================================================================
ws = wb.create_sheet("Valuation"); ws.sheet_view.showGridLines = False
ws["A1"] = "VALUATION MULTIPLES"
ws.merge_cells("A1:H1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "EV = Mkt Cap + Net Debt. NM where denominator <= 0. Forward P/E is the cleaner cross-peer gauge (GAAP earnings distorted by amortization at AMD/MRVL)."
ws.merge_cells("A2:H2"); ws["A2"].font = ITAL; ws["A2"].alignment = C
vh = ["Company", "Mkt Cap", "EV", "EV / Rev", "EV / EBITDA", "P/E (TTM)", "Fwd P/E", "Comment"]
for i, h in enumerate(vh, start=1):
    colhead(ws, 6, i, h)
comments = {
    "Broadcom": "Target. TTM P/E optically high (~87x) on GAAP NI; fwd ~31x. DCF base ~$246 << $460 mkt (see model).",
    "NVIDIA": "Cheapest large-cap on earnings (P/E ~34) despite $5.5T cap — vast profitability; net cash.",
    "AMD": "Extreme TTM multiples (GAAP ramping); read on EV/Rev (~22x) and fwd P/E.",
    "Marvell Technology": "Purest custom-silicon comp; GAAP P/E NM-high (amortization); fwd ~48x. Re-rated on NVDA endorsement.",
    "Qualcomm": "Mature-semi value name; cheapest fwd P/E (~14x). Limited AI-accelerator exposure.",
    "Texas Instruments": "Analog; high multiples reflect trough earnings in heavy-capex cycle, not just quality.",
    "Oracle": "Software anchor (memo). Now an AI-capex story; ~$96B net debt rising; EV/Rev ~12x embeds high growth.",
}
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; src = 7 + idx
    cn = ws.cell(row=r, column=1, value=name); cn.font = BOLDB
    put_f(ws, r, 2, f"=Inputs!C{src}")
    put_f(ws, r, 3, f"=Inputs!C{src}+Inputs!D{src}")
    put_f(ws, r, 4, f"=(Inputs!C{src}+Inputs!D{src})/Inputs!E{src}", '0.0"x"')
    put_f(ws, r, 5, f"=IF(Inputs!H{src}>0,(Inputs!C{src}+Inputs!D{src})/Inputs!H{src},\"NM\")", '0.0"x"')
    put_f(ws, r, 6, f"=IF(Inputs!I{src}>0,Inputs!C{src}/Inputs!I{src},\"NM\")", '0.0"x"')
    put_f(ws, r, 7, f"=Inputs!J{src}", '0.0"x"')
    cc = ws.cell(row=r, column=8, value=comments[name]); cc.font = ITAL; cc.alignment = L
stat_block(ws, 14, 1, {4: '0.0"x"', 5: '0.0"x"', 6: '0.0"x"', 7: '0.0"x"'})
for i in range(5):
    for col in (2, 3, 8):
        ws.cell(row=14 + i, column=col).fill = LG_FILL
for i, w in enumerate([20, 12, 12, 11, 13, 12, 11, 60], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# TAB 4 — GROWTH & AI EXPOSURE (industry-specific lens)
# =====================================================================
ws = wb.create_sheet("Growth & AI Exposure"); ws.sheet_view.showGridLines = False
ws["A1"] = "GROWTH & AI EXPOSURE — Industry-Specific Lens"
ws.merge_cells("A1:F1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "For this peer set the differentiators are AI revenue exposure, growth, and PEG (growth-adjusted valuation). AI revenue figures are [E] where not separately disclosed."
ws.merge_cells("A2:F2"); ws["A2"].font = ITAL; ws["A2"].alignment = C
gh = ["Company", "Rev Growth (YoY)", "Fwd P/E", "PEG (fwd P/E ÷ growth)", "AI exposure", "Posture"]
for i, h in enumerate(gh, start=1):
    colhead(ws, 6, i, h)
ai_expo = {
    "Broadcom": ("High — direct", "Custom XPU/ASIC + AI networking; 6 committed hyperscaler customers; $73B AI backlog"),
    "NVIDIA": ("Very high — merchant GPU leader", "~70%+ of AI accelerator market; the comp everyone anchors to"),
    "AMD": ("High — #2 merchant GPU", "MI-series data-center GPUs; ~10-15% accelerator share"),
    "Marvell Technology": ("High — custom silicon", "Closest AVGO ASIC comp; Amazon Trainium, Microsoft Maia; ~76% DC revenue"),
    "Qualcomm": ("Low — edge/on-device", "Handset/auto/IoT; limited data-center AI accelerator exposure"),
    "Texas Instruments": ("Low — analog content", "Benefits from AI server analog content, but no accelerator franchise"),
    "Oracle": ("High — AI infra demand", "OCI capacity for AI training (incl. OpenAI); software anchor for VMware"),
}
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; src = 7 + idx
    cn = ws.cell(row=r, column=1, value=name); cn.font = BOLDB
    put_f(ws, r, 2, f"=Inputs!F{src}", "0.0%")
    put_f(ws, r, 3, f"=Inputs!J{src}", '0.0"x"')
    # PEG = fwd P/E / (growth% * 100); guard against <=0 growth
    put_f(ws, r, 4, f"=IF(Inputs!F{src}>0,Inputs!J{src}/(Inputs!F{src}*100),\"NM\")", "0.00")
    ws.cell(row=r, column=5, value=ai_expo[name][0]).font = BLACK
    cc = ws.cell(row=r, column=6, value=ai_expo[name][1]); cc.font = ITAL; cc.alignment = L
stat_block(ws, 14, 1, {2: "0.0%", 3: '0.0"x"', 4: "0.00"})
for i in range(5):
    for col in (5, 6):
        ws.cell(row=14 + i, column=col).fill = LG_FILL
for i, w in enumerate([20, 16, 11, 22, 26, 60], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =====================================================================
# TAB 5 — NOTES
# =====================================================================
ws = wb.create_sheet("Notes"); ws.sheet_view.showGridLines = False
ws["A1"] = "NOTES, METHODOLOGY & DATA-SOURCE CAVEATS"
ws.merge_cells("A1:E1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
notes = [
    ("Peer framing", [
        "Second framing (user-selected): semiconductor cohort + one infrastructure-software anchor.",
        "Stat block (Max/75th/Median/25th/Min) is computed over the 5 SEMIS peers only: NVDA, AMD, MRVL, QCOM, TXN.",
        "AVGO is the target (compare to the semis median). ORCL is a software-anchor MEMO row for the VMware (~40%) segment — excluded from semis stats.",
        "Rationale: you don't median a software multiple into a semis set; ORCL frames the software side qualitatively + on its own multiples.",
    ]),
    ("Data sources", [
        "Reported figures: company IR releases and SEC 10-Q/10-K filings for each issuer (most recent quarter as of May/June 2026).",
        "Market data (caps, EV): StockAnalysis.com / companiesmarketcap / GuruFocus, ~June 1-2, 2026.",
        "MCP terminal connectors (CapIQ/FactSet/Daloopa) were NOT configured; public sources used and flagged [E] where estimated.",
    ]),
    ("Key adjustments & caveats", [
        "GAAP TTM P/E is distorted for AMD and MRVL (earnings depressed by acquired-intangible amortization) — use FORWARD P/E for comparability.",
        "QCOM net income excludes a one-time $5.7B Q2 FY26 tax benefit (~$9.9B underlying TTM used).",
        "Market caps were volatile May-June 2026; ORCL in particular ranged ~$550B-$718B — ~$640B used. Re-pull all caps on one date before finalizing.",
        "NVIDIA TTM revenue $253.5B (run-rate-consistent) used; an alternative aggregator window showed ~$216B.",
        "AVGO EBITDA shown on a GAAP basis (~$40.5B) for peer comparability; adjusted EBITDA is ~$46B (~68% margin).",
        "EV excludes operating-lease liabilities and uses simple net debt (total debt - cash). NVDA/AMD net-cash positions are conservative (exclude marketable securities).",
    ]),
    ("Cross-references", [
        "Intrinsic valuation: AVGO-Model.xlsx (3-statement + DCF). Base-case DCF ~$246/share vs ~$460 market.",
        "Narrative: AVGO-Competitive-Analysis.pptx / .html.",
        "Build script: build_comps.py — re-run after updating INPUTS.",
    ]),
]
row = 3
for title, lines in notes:
    c = ws.cell(row=row, column=1, value=title); c.font = BOLDB; c.fill = LB_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); row += 1
    for ln in lines:
        cc = ws.cell(row=row, column=1, value=ln); cc.font = BLACK; cc.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 16; row += 1
    row += 1
for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Inputs"], wb["Operating Metrics"], wb["Valuation"], wb["Growth & AI Exposure"], wb["Notes"]]
wb.save(OUT)
print(f"Wrote {OUT}")
