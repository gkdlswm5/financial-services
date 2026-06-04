"""
Build GOOG-Comps-Analysis.xlsx — comparable company analysis for Alphabet (GOOG).
Pattern follows analysis/AVGO/build_comps.py.

Peer framing: Mega-cap cohort (META, MSFT, AMZN, NFLX) as main stat set.
Ads-pure memo rows (TTD, RDDT) — high-growth ad-tech reference points.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "GOOG-Comps-Analysis.xlsx"
NAVY = "17365D"; LB = "D9E1F2"; LG = "F2F2F2"
BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDB = Font(name="Times New Roman", size=12, bold=True)
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")
NAVY_FILL = PatternFill("solid", fgColor=NAVY); LB_FILL = PatternFill("solid", fgColor=LB); LG_FILL = PatternFill("solid", fgColor=LG)
C = Alignment("center", "center"); L = Alignment("left", "center"); WRAP = Alignment("left", "top", wrap_text=True)

PEERS = [
    ("Alphabet",          "GOOG",  "target"),
    ("Meta Platforms",    "META",  "mega"),
    ("Microsoft",         "MSFT",  "mega"),
    ("Amazon",            "AMZN",  "mega"),
    ("Netflix",           "NFLX",  "mega"),
    ("The Trade Desk",    "TTD",   "memo"),
    ("Reddit",            "RDDT",  "memo"),
]

INPUTS = {
    "Alphabet": {
        "mkt_cap": (4335000, "GOOGL ~$357.73/sh x ~12.12B sh = ~$4.34T (June 3, 2026)"),
        "net_debt": (-80300, "Net cash ~$80B: cash+sec $126.8B - debt $46.5B (Q1 26)"),
        "revenue": (425100, "[E] TTM revenue ~$425B (FY25 $402.8B + Q1 26 step-up)"),
        "growth": (0.22, "Q1 CY26 revenue +22% YoY (+19% cc)"),
        "gm": (0.58, "[E] GAAP gross margin ~58%"),
        "ebitda": (155000, "[E] TTM EBITDA ~$155B (NI + interest + tax + D&A)"),
        "ni": (137000, "[E] TTM NI ~$137B (includes Anthropic mark-to-market gains)"),
        "fwd_pe": (24.0, "[E] Forward P/E ~24x (consensus FY26-27)"),
    },
    "Meta Platforms": {
        "mkt_cap": (1550000, "META ~$1.55T (~$1.50-1.61T spread across sources, late May/early Jun 2026)"),
        "net_debt": (-40000, "[E] Net cash ~$40B (large cash balance, modest debt)"),
        "revenue": (223000, "TTM revenue ~$223B (multiples.vc)"),
        "growth": (0.33, "Q1 CY26 revenue +33% YoY"),
        "gm": (0.82, "[E] Gross margin ~82%"),
        "ebitda": (110000, "TTM EBITDA ~$110B (~51% margin)"),
        "ni": (71000, "TTM NI ~$71B"),
        "fwd_pe": (18.2, "Forward P/E ~18.2x"),
    },
    "Microsoft": {
        "mkt_cap": (4720000, "MSFT ~$4.72T (StockAnalysis, June 1 2026)"),
        "net_debt": (-30000, "[E] Slight net cash; debt ~$50B vs cash+inv ~$80B"),
        "revenue": (318300, "TTM revenue ~$318.3B"),
        "growth": (0.18, "Fiscal Q3 FY26 (CQ1) revenue +18% YoY; Azure +40%"),
        "gm": (0.68, "[E] Gross margin ~68% (down ~70bps YoY on AI capex depreciation)"),
        "ebitda": (274000, "[E] TTM EBITDA ~$274B (implied from 17.1x EV/EBITDA)"),
        "ni": (125200, "TTM NI ~$125.2B"),
        "fwd_pe": (22.6, "Forward P/E ~22.6x"),
    },
    "Amazon": {
        "mkt_cap": (2700000, "AMZN ~$2.7T mkt cap (June 2-3, 2026)"),
        "net_debt": (66000, "[E] Net debt ~$66B (debt $153B - cash $87B, excl. mktable sec)"),
        "revenue": (742800, "TTM revenue ~$742.8B"),
        "growth": (0.17, "Q1 CY26 revenue +17% YoY; AWS +28% to $37.6B"),
        "gm": (0.50, "TTM gross margin ~50.3%"),
        "ebitda": (145700, "TTM EBITDA ~$145.7B"),
        "ni": (91000, "[E] TTM NI ~$91B (Q1 incl $16.8B Anthropic gain)"),
        "fwd_pe": (30.0, "[E] Forward P/E ~30x (consensus)"),
    },
    "Netflix": {
        "mkt_cap": (400000, "NFLX ~$400B (range ~$343-410B; mlq.ai most recent)"),
        "net_debt": (4450, "Net debt ~$4.45B: debt $14.36B - cash+inv $12.30B"),
        "revenue": (46900, "TTM revenue ~$46.9B"),
        "growth": (0.16, "Q1 CY26 revenue +16.2% YoY"),
        "gm": (0.52, "Gross margin ~52% Q1"),
        "ebitda": (31100, "[E] TTM EBITDA ~$31.1B (~66%)"),
        "ni": (11000, "[E] TTM NI ~$11B"),
        "fwd_pe": (26.0, "[E] Forward P/E ~26x"),
    },
    "The Trade Desk": {
        "mkt_cap": (10900, "TTD ~$10.9B (June 1, 2026)"),
        "net_debt": (-1700, "Net cash ~$1.7B (no debt)"),
        "revenue": (2960, "TTM revenue ~$2.96B"),
        "growth": (0.12, "Q1 CY26 revenue +12% YoY (slowing)"),
        "gm": (0.736, "Gross margin ~73.6% TTM"),
        "ebitda": (900, "[E] TTM EBITDA ~$0.9B"),
        "ni": (435, "[E] TTM NI ~$435M"),
        "fwd_pe": (11.0, "[E] Forward P/E ~11x (compressed on deceleration)"),
    },
    "Reddit": {
        "mkt_cap": (30000, "RDDT ~$30B (range $26.7-33.9B across May-Jun)"),
        "net_debt": (-2000, "[E] Net cash ~$2B (no significant debt)"),
        "revenue": (2200, "TTM revenue ~$2.2B"),
        "growth": (0.69, "Q1 CY26 revenue +69% YoY; ads +74%"),
        "gm": (0.915, "Gross margin ~91.5%"),
        "ebitda": (545, "[E] TTM adj EBITDA ~$545M"),
        "ni": (530, "[E] FY25 NI ~$530M (trending higher Q1: $204M)"),
        "fwd_pe": (33.0, "[E] Forward P/E ~33x (deceleration assumed)"),
    },
}

wb = Workbook(); wb.remove(wb.active)
labels = ["Maximum", "75th Percentile", "Median", "25th Percentile", "Minimum"]
SEMI_FIRST, SEMI_LAST = 8, 11  # NVDA..NFLX (rows 8-11), GOOG=7, memo rows 12-13


def section(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text); ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0); cell.font = BOLDW; cell.fill = NAVY_FILL; cell.alignment = C

def colhead(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text); cell.font = BOLDB; cell.fill = LB_FILL; cell.alignment = C

def put_input(ws, r, c, val, src, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=val); cell.font = BLUE; cell.alignment = C; cell.number_format = fmt
    cell.comment = Comment(src, "GOOG comps")

def put_f(ws, r, c, formula, fmt="#,##0"):
    cell = ws.cell(row=r, column=c, value=formula); cell.font = BLACK; cell.alignment = C; cell.number_format = fmt

def stat_block(ws, start_row, label_col, cols_fmt):
    for i, lab in enumerate(labels):
        r = start_row + i
        c = ws.cell(row=r, column=label_col, value=lab); c.font = BOLDB; c.fill = LG_FILL; c.alignment = L
        for col in range(label_col + 1, max(cols_fmt) + 1):
            ws.cell(row=r, column=col).fill = LG_FILL
        for col, fmt in cols_fmt.items():
            cl = get_column_letter(col); rng = f"{cl}{SEMI_FIRST}:{cl}{SEMI_LAST}"
            f = {"Maximum": f"=MAX({rng})", "75th Percentile": f"=QUARTILE({rng},3)",
                 "Median": f"=MEDIAN({rng})", "25th Percentile": f"=QUARTILE({rng},1)",
                 "Minimum": f"=MIN({rng})"}[lab]
            cell = ws.cell(row=r, column=col, value=f); cell.font = BOLDB; cell.fill = LG_FILL; cell.alignment = C; cell.number_format = fmt

# Tab 1 — INPUTS
ws = wb.create_sheet("Inputs"); ws.sheet_view.showGridLines = False
ws["A1"] = "MEGA-CAP TECH — COMPARABLE COMPANY ANALYSIS (GOOG target)"
ws.merge_cells("A1:K1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = " • ".join(f"{n} ({t})" for n, t, _ in PEERS)
ws.merge_cells("A2:K2"); ws["A2"].font = BOLDB; ws["A2"].alignment = C
ws["A3"] = "As of June 3, 2026 | $M except ratios | Stats over the 4 mega-cap peers (META, MSFT, AMZN, NFLX); GOOG=target; TTD, RDDT = ads-pure memo"
ws.merge_cells("A3:K3"); ws["A3"].font = ITAL; ws["A3"].alignment = C
section(ws, 5, 1, 11, "RAW INPUTS — cell comments cite source / assumption")
heads = ["Company", "Ticker", "Mkt Cap", "Net Debt", "Revenue (TTM)", "Rev Growth %",
         "Gross Margin %", "EBITDA (TTM)", "Net Income (TTM)", "Fwd P/E", "Role"]
for i, h in enumerate(heads, start=1): colhead(ws, 6, i, h)
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; d = INPUTS[name]
    cn = ws.cell(row=r, column=1, value=name); cn.font = BOLDB
    ws.cell(row=r, column=2, value=tic).font = BLACK
    put_input(ws, r, 3, d["mkt_cap"][0], d["mkt_cap"][1])
    put_input(ws, r, 4, d["net_debt"][0], d["net_debt"][1])
    put_input(ws, r, 5, d["revenue"][0], d["revenue"][1])
    put_input(ws, r, 6, d["growth"][0], d["growth"][1], "0.0%")
    put_input(ws, r, 7, d["gm"][0], d["gm"][1], "0.0%")
    put_input(ws, r, 8, d["ebitda"][0], d["ebitda"][1])
    put_input(ws, r, 9, d["ni"][0], d["ni"][1])
    put_input(ws, r, 10, d["fwd_pe"][0], d["fwd_pe"][1], '0.0"x"')
    tag = "TARGET" if kind == "target" else ("MEGA-CAP" if kind == "mega" else "ADS-PURE (memo)")
    ws.cell(row=r, column=11, value=tag).font = ITAL
for i, w in enumerate([20, 9, 13, 11, 14, 13, 14, 14, 16, 10, 20], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Tab 2 — OPERATING METRICS
ws = wb.create_sheet("Operating Metrics"); ws.sheet_view.showGridLines = False
ws["A1"] = "OPERATING METRICS"; ws.merge_cells("A1:F1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
oh = ["Company", "Revenue (TTM)", "Rev Growth (YoY)", "Gross Margin", "EBITDA (TTM)", "EBITDA Margin"]
for i, h in enumerate(oh, start=1): colhead(ws, 6, i, h)
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLDB
    put_f(ws, r, 2, f"=Inputs!E{src}")
    put_f(ws, r, 3, f"=Inputs!F{src}", "0.0%")
    put_f(ws, r, 4, f"=Inputs!G{src}", "0.0%")
    put_f(ws, r, 5, f"=Inputs!H{src}")
    put_f(ws, r, 6, f"=Inputs!H{src}/Inputs!E{src}", "0.0%")
stat_block(ws, 15, 1, {3: "0.0%", 4: "0.0%", 6: "0.0%"})
for i, w in enumerate([20, 14, 16, 14, 14, 14], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Tab 3 — VALUATION
ws = wb.create_sheet("Valuation"); ws.sheet_view.showGridLines = False
ws["A1"] = "VALUATION MULTIPLES"; ws.merge_cells("A1:H1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "EV = Mkt Cap + Net Debt (negative net debt = net cash subtracts from EV). NM if EBITDA/NI <= 0."
ws.merge_cells("A2:H2"); ws["A2"].font = ITAL; ws["A2"].alignment = C
vh = ["Company", "Mkt Cap", "EV", "EV / Rev", "EV / EBITDA", "P/E (TTM)", "Fwd P/E", "Comment"]
for i, h in enumerate(vh, start=1): colhead(ws, 6, i, h)
comments = {
    "Alphabet": "Target. Cheap on fwd P/E (~24x) vs the mega-cap median; DCF base $213 << $358 (FY26 capex spike).",
    "Meta Platforms": "Cheapest fwd P/E (~18x) in the mega-cap set; Reality Labs $4B/qtr drag.",
    "Microsoft": "Premium for Azure +40% + OpenAI exposure; $190B FY26 capex absorbs cash.",
    "Amazon": "Highest fwd P/E (~30x); AWS +28% reaccel + ads $70B+ TTM; retail margin still thin.",
    "Netflix": "Premium streaming P/E (~26x); subs no longer disclosed; ad business 2x in 2026.",
    "The Trade Desk": "Memo. Programmatic ad-tech; multiple compressed on deceleration to +12%.",
    "Reddit": "Memo. Hyper-growth (+69% rev) at ~50x P/E; data-licensing revenue accelerating.",
}
for idx, (name, tic, kind) in enumerate(PEERS):
    r = 7 + idx; src = 7 + idx
    ws.cell(row=r, column=1, value=name).font = BOLDB
    put_f(ws, r, 2, f"=Inputs!C{src}")
    put_f(ws, r, 3, f"=Inputs!C{src}+Inputs!D{src}")
    put_f(ws, r, 4, f"=(Inputs!C{src}+Inputs!D{src})/Inputs!E{src}", '0.0"x"')
    put_f(ws, r, 5, f"=IF(Inputs!H{src}>0,(Inputs!C{src}+Inputs!D{src})/Inputs!H{src},\"NM\")", '0.0"x"')
    put_f(ws, r, 6, f"=IF(Inputs!I{src}>0,Inputs!C{src}/Inputs!I{src},\"NM\")", '0.0"x"')
    put_f(ws, r, 7, f"=Inputs!J{src}", '0.0"x"')
    cc = ws.cell(row=r, column=8, value=comments[name]); cc.font = ITAL; cc.alignment = L
stat_block(ws, 15, 1, {4: '0.0"x"', 5: '0.0"x"', 6: '0.0"x"', 7: '0.0"x"'})
for i, w in enumerate([20, 12, 12, 11, 13, 12, 11, 56], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Tab 4 — NOTES
ws = wb.create_sheet("Notes"); ws.sheet_view.showGridLines = False
ws["A1"] = "NOTES & CAVEATS"; ws.merge_cells("A1:E1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
notes = [
    ("Peer framing", [
        "Mega-cap cohort (META, MSFT, AMZN, NFLX) is the main stat set — all platform-economy AI capex spenders.",
        "TTD and RDDT are ads-pure memo rows — useful frame for the Search/YouTube ad business, but too small for direct multiple read.",
        "GOOG is the target. Compare to mega-cap median on fwd P/E (cleanest cross-peer gauge).",
    ]),
    ("Data sources", [
        "All Q1 CY2026 10-Qs / 8-Ks (Apr-May 2026) and FY25 10-Ks. Market data: StockAnalysis / companiesmarketcap / GuruFocus / Macrotrends.",
        "MCP terminals not configured. ENS aggregator figures vary 10-20% across sources (esp. TTM EBITDA definitions).",
    ]),
    ("Key adjustments / caveats", [
        "MSFT TTM EBITDA implied from 17.1x EV/EBITDA (no direct disclosure).",
        "AMZN TTM NI inflated by Q1 $16.8B Anthropic mark-to-market gain.",
        "GOOG TTM NI similarly elevated by Anthropic-related fair-value changes.",
        "NFLX EBITDA depends on content amortization treatment; range $14.9B-$31.1B across definitions.",
        "Companion: GOOG-Model.xlsx (intrinsic DCF: base $213/share vs $358 market).",
    ]),
]
row = 3
for title, lines in notes:
    c = ws.cell(row=row, column=1, value=title); c.font = BOLDB; c.fill = LB_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); row += 1
    for ln in lines:
        cc = ws.cell(row=row, column=1, value=ln); cc.font = BLACK; cc.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); ws.row_dimensions[row].height = 16; row += 1
    row += 1
for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Inputs"], wb["Operating Metrics"], wb["Valuation"], wb["Notes"]]
wb.save(OUT)
print(f"Wrote {OUT}")
