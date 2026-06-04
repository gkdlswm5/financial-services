"""
Build GOOG-Model.xlsx — 3-statement model + DCF for Alphabet (GOOG/GOOGL).
Pattern follows analysis/AVGO/build_model.py.

Segments: Google Services · Google Cloud · Other Bets (Waymo etc.)
Period: FY2023A FY2024A FY2025A | FY2026E-FY2030E (calendar fiscal year).
Run: python3 build_model.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "GOOG-Model.xlsx"
NAVY = "17365D"; LB = "D9E1F2"; LG = "F2F2F2"; MB = "BDD7EE"
BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLD = Font(name="Times New Roman", size=11, bold=True)
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDN = Font(name="Times New Roman", size=11, bold=True, color="17365D")
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")
NAVY_FILL = PatternFill("solid", fgColor=NAVY); LB_FILL = PatternFill("solid", fgColor=LB)
LG_FILL = PatternFill("solid", fgColor=LG); MB_FILL = PatternFill("solid", fgColor=MB)
C = Alignment("center", "center"); L = Alignment("left", "center"); WRAP = Alignment("left", "top", wrap_text=True)
NUM = "#,##0"; PCT = "0.0%"; USD2 = '"$"#,##0.00'

YR_COLS = ["C", "D", "E", "F", "G", "H", "I", "J"]
YR_HEADS = ["FY2023A", "FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]

wb = Workbook(); wb.remove(wb.active)

def hdr(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text); ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0); cell.font = BOLDW; cell.fill = NAVY_FILL; cell.alignment = C

def yearrow(ws, row, label="($ in millions)"):
    ws.cell(row=row, column=1, value=label).font = ITAL
    for i, col in enumerate(YR_COLS):
        cell = ws.cell(row=row, column=3 + i, value=YR_HEADS[i])
        cell.font = BOLDN; cell.fill = LB_FILL; cell.alignment = C

def lbl(ws, row, text, *, bold=False, indent=0):
    cell = ws.cell(row=row, column=1, value=("  " * indent) + text)
    cell.font = BOLD if bold else BLACK; cell.alignment = L

def put_input(ws, addr, value, source, fmt=NUM):
    cell = ws[addr]; cell.value = value; cell.font = BLUE; cell.alignment = C; cell.number_format = fmt
    cell.comment = Comment(source, "GOOG model")

def put_f(ws, addr, formula, fmt=NUM, *, bold=False, fill=None):
    cell = ws[addr]; cell.value = formula
    cell.font = BOLDN if bold else BLACK; cell.alignment = C; cell.number_format = fmt
    if fill: cell.fill = fill

# DRIVERS
ws = wb.create_sheet("Drivers"); ws.sheet_view.showGridLines = False
ws["A1"] = "ALPHABET (GOOG/GOOGL) — INTEGRATED MODEL: DRIVERS & ASSUMPTIONS"
ws.merge_cells("A1:J1"); ws["A1"].font = BOLDW; ws["A1"].fill = NAVY_FILL; ws["A1"].alignment = C
ws["A2"] = "As of June 3, 2026 | $M unless noted | Blue=input, Black=formula | Calendar fiscal year"
ws.merge_cells("A2:J2"); ws["A2"].font = ITAL; ws["A2"].alignment = C

hdr(ws, 4, 1, 10, "SEGMENT REVENUE BUILD ($M)")
yearrow(ws, 5)
r_serv = 6; r_cloud = 7; r_other = 8; r_tot = 9
lbl(ws, r_serv, "Google Services (Search, YouTube, Other)", indent=1)
lbl(ws, r_cloud, "Google Cloud", indent=1)
lbl(ws, r_other, "Other Bets (Waymo etc.)", indent=1)
lbl(ws, r_tot, "Total revenue", bold=True)

serv_h = {
    "C": (272500, "[E] FY23 Services ~$272.5B (per 10-K)"),
    "D": (304922, "[E] FY24 Services ~$304.9B"),
    "E": (342700, "FY25 Services $342.7B (per FY25 10-K segment disclosure)"),
}
cloud_h = {
    "C": (33088, "FY23 Cloud $33.1B (first profitable year, +26%)"),
    "D": (43200, "[E] FY24 Cloud ~$43.2B (+31%)"),
    "E": (58700, "FY25 Cloud $58.7B (+36%)"),
}
other_h = {
    "C": (1527, "[E] FY23 Other Bets ~$1.5B"),
    "D": (1539, "[E] FY24 Other Bets ~$1.5B"),
    "E": (1537, "FY25 Other Bets $1.537B (per FY25 10-K)"),
}
for col in ("C", "D", "E"):
    put_input(ws, f"{col}{r_serv}", serv_h[col][0], serv_h[col][1])
    put_input(ws, f"{col}{r_cloud}", cloud_h[col][0], cloud_h[col][1])
    put_input(ws, f"{col}{r_other}", other_h[col][0], other_h[col][1])

# Projections — Services moderate growth, Cloud hyper-growth tapering, Other Bets stable
serv_p = {"F": 393100, "G": 432400, "H": 471300, "I": 509000, "J": 545600}
cloud_p = {"F": 92000, "G": 132000, "H": 178000, "I": 222000, "J": 266000}
other_p = {"F": 1800, "G": 2100, "H": 2500, "I": 3000, "J": 3600}
serv_s = "[E] Services +13-15%: Search resilient w/ AI Overviews, YouTube ads, Subscription growth; reflects Q1 26 +16%"
cloud_s = "[E] Cloud trajectory: $58.7B (FY25) → $92B (+57%) FY26 / $266B (FY30); Q1 26 was +63%; $460B backlog supports"
other_s = "[E] Other Bets modest; Waymo not yet material-revenue"
for col in ("F", "G", "H", "I", "J"):
    put_input(ws, f"{col}{r_serv}", serv_p[col], serv_s)
    put_input(ws, f"{col}{r_cloud}", cloud_p[col], cloud_s)
    put_input(ws, f"{col}{r_other}", other_p[col], other_s)
for col in YR_COLS:
    put_f(ws, f"{col}{r_tot}", f"=SUM({col}{r_serv}:{col}{r_other})", bold=True)

r_growth = 10
lbl(ws, r_growth, "Total revenue growth %", indent=1)
for i, col in enumerate(YR_COLS):
    if i == 0:
        ws[f"{col}{r_growth}"] = "n/a"; ws[f"{col}{r_growth}"].font = ITAL; ws[f"{col}{r_growth}"].alignment = C
    else:
        prev = YR_COLS[i - 1]; put_f(ws, f"{col}{r_growth}", f"={col}{r_tot}/{prev}{r_tot}-1", PCT)

# Segment operating margins
hdr(ws, 12, 1, 10, "SEGMENT OPERATING MARGINS")
yearrow(ws, 13)
r_serv_m = 14; r_cloud_m = 15; r_other_m = 16; r_corp = 17
lbl(ws, r_serv_m, "Google Services op margin %", indent=1)
lbl(ws, r_cloud_m, "Google Cloud op margin %", indent=1)
lbl(ws, r_other_m, "Other Bets op margin % (loss)", indent=1)
lbl(ws, r_corp, "Corporate / unallocated ($M)", indent=1)
serv_m = {"C": 0.45, "D": 0.45, "E": 0.45, "F": 0.45, "G": 0.45, "H": 0.45, "I": 0.45, "J": 0.45}
cloud_m = {"C": 0.052, "D": 0.14, "E": 0.237, "F": 0.30, "G": 0.32, "H": 0.34, "I": 0.35, "J": 0.36}
other_m = {"C": -2.93, "D": -3.0, "E": -4.88, "F": -3.5, "G": -2.5, "H": -1.5, "I": -1.0, "J": -0.5}
corp_v = {"C": 7000, "D": 7500, "E": 8000, "F": 9000, "G": 10000, "H": 11000, "I": 12000, "J": 13000}
for col in YR_COLS:
    put_input(ws, f"{col}{r_serv_m}", serv_m[col], "[E] Services op margin ~45% (Search-heavy mix)", PCT)
    put_input(ws, f"{col}{r_cloud_m}", cloud_m[col], "[E] Cloud margin: 5%(23)->14%(24)->24%(25)->30%(26)->36%(30); Q1 26 was 32.9%", PCT)
    put_input(ws, f"{col}{r_other_m}", other_m[col], "[E] Other Bets loss/rev ratio improving as Waymo monetizes", PCT)
    put_input(ws, f"{col}{r_corp}", corp_v[col], "[E] Corporate unallocated costs scale with size", NUM)

# Computed total EBIT
r_ebit = 19
lbl(ws, r_ebit, "Total operating income (EBIT)", bold=True)
for col in YR_COLS:
    put_f(ws, f"{col}{r_ebit}",
          f"={col}{r_serv}*{col}{r_serv_m}+{col}{r_cloud}*{col}{r_cloud_m}+{col}{r_other}*{col}{r_other_m}-{col}{r_corp}", bold=True)

# D&A / Capex / tax / NWC
hdr(ws, 21, 1, 10, "D&A / CAPEX / TAX / NWC")
yearrow(ws, 22)
r_da = 23; r_capex = 24; r_tax = 25; r_nwc = 26
lbl(ws, r_da, "D&A ($M)", indent=1)
lbl(ws, r_capex, "Capex ($M)", indent=1)
lbl(ws, r_tax, "Cash tax rate %", indent=1)
lbl(ws, r_nwc, "Incr. NWC % of rev growth", indent=1)
da_v = {"C": 11946, "D": 15311, "E": 22000, "F": 38000, "G": 58000, "H": 76000, "I": 90000, "J": 100000}
capex_v = {"C": 32251, "D": 52535, "E": 92000, "F": 185000, "G": 200000, "H": 180000, "I": 160000, "J": 150000}
tax_v = {"C": 0.155, "D": 0.16, "E": 0.16, "F": 0.18, "G": 0.20, "H": 0.21, "I": 0.21, "J": 0.21}
nwc_v = {"C": 0.05, "D": 0.05, "E": 0.05, "F": 0.05, "G": 0.05, "H": 0.05, "I": 0.05, "J": 0.05}
da_src = {c: "[E] D&A ramps with AI infra capex with ~2-3yr lag; F-J reflect $92B FY25 + $180-190B FY26 ramp" for c in YR_COLS}
capex_src = {c: "FY23 $32B / FY24 $52B / FY25 ~$92B / FY26 guide $180-190B (back-end loaded); model assumes plateau then taper" for c in YR_COLS}
tax_src = {c: "[E] Cash tax rate ~16-21% (GILTI + state); rising as int'l rules tighten" for c in YR_COLS}
nwc_src = {c: "[E] NWC growth modest (large AR/AP balances)" for c in YR_COLS}
for col in YR_COLS:
    put_input(ws, f"{col}{r_da}", da_v[col], da_src[col])
    put_input(ws, f"{col}{r_capex}", capex_v[col], capex_src[col])
    put_input(ws, f"{col}{r_tax}", tax_v[col], tax_src[col], PCT)
    put_input(ws, f"{col}{r_nwc}", nwc_v[col], nwc_src[col], PCT)

# Market data & WACC
hdr(ws, 28, 1, 4, "MARKET DATA & WACC INPUTS")
md = [
    ("Current share price ($)", 357.73, "GOOGL ~$357.73 (June 3, 2026)", USD2),
    ("Diluted shares (M)", 12120, "~12.12B diluted shares (Class A + Class B + Class C)", NUM),
    ("Total debt ($M)", 46500, "LT debt $46.5B Q1 26 (after $31.1B Q1 senior notes)", NUM),
    ("Cash & marketable securities ($M)", 126800, "Cash $30.7B + marketable sec $96.1B = $126.8B (Q1 26)", NUM),
    ("Risk-free rate %", 0.043, "[E] 10-yr UST ~4.3%", PCT),
    ("Equity risk premium %", 0.050, "[E] ERP ~5.0%", PCT),
    ("Beta", 1.00, "[E] GOOG beta ~1.0 (5-yr levered)", "0.00"),
    ("Pre-tax cost of debt %", 0.045, "[E] GOOG investment-grade Kd ~4.5%", PCT),
    ("Terminal growth rate %", 0.030, "[E] BASE terminal growth 3.0% (large-cap tech, secular AI tailwind)", PCT),
]
row = 29
for name, val, src, fmt in md:
    ws.cell(row=row, column=1, value=name).font = BLACK
    put_input(ws, f"C{row}", val, src, fmt); row += 1
R_PRICE, R_SH, R_DEBT, R_CASH, R_RF, R_ERP, R_BETA, R_KD, R_TG = range(29, 38)

hdr(ws, 39, 1, 4, "WACC (computed)")
ws["A40"] = "Cost of equity"; ws["A40"].font = BLACK
put_f(ws, "C40", f"=C{R_RF}+C{R_BETA}*C{R_ERP}", PCT)
ws["A41"] = "After-tax cost of debt"; ws["A41"].font = BLACK
put_f(ws, "C41", f"=C{R_KD}*(1-E{r_tax})", PCT)
ws["A42"] = "Market value of equity ($M)"; ws["A42"].font = BLACK
put_f(ws, "C42", f"=C{R_PRICE}*C{R_SH}", NUM)
ws["A43"] = "Net debt / (cash) ($M)"; ws["A43"].font = BLACK
put_f(ws, "C43", f"=C{R_DEBT}-C{R_CASH}", NUM)
ws["A44"] = "Equity weight"; ws["A44"].font = BLACK
put_f(ws, "C44", f"=C42/(C42+C{R_DEBT})", PCT)
ws["A45"] = "Debt weight"; ws["A45"].font = BLACK
put_f(ws, "C45", f"=C{R_DEBT}/(C42+C{R_DEBT})", PCT)
ws["A46"] = "WACC"; ws["A46"].font = BOLDN
put_f(ws, "C46", "=C44*C40+C45*C41", PCT, bold=True, fill=LG_FILL)
R_WACC = 46

ws.column_dimensions["A"].width = 38
for col in YR_COLS:
    ws.column_dimensions[col].width = 11

# INCOME STATEMENT (lean)
isw = wb.create_sheet("Income Statement"); isw.sheet_view.showGridLines = False
isw["A1"] = "INCOME STATEMENT ($M)"
isw.merge_cells("A1:J1"); isw["A1"].font = BOLDW; isw["A1"].fill = NAVY_FILL; isw["A1"].alignment = C
yearrow(isw, 4)
lbl(isw, 5, "Total revenue", bold=True)
lbl(isw, 7, "Operating income (EBIT)")
lbl(isw, 8, "(+) D&A")
lbl(isw, 9, "EBITDA", bold=True)
lbl(isw, 10, "(-) Net interest expense (or +income)")
lbl(isw, 11, "Pre-tax income")
lbl(isw, 12, "(-) Income taxes")
lbl(isw, 13, "Net income (GAAP)", bold=True)
lbl(isw, 14, "Net margin %")
for col in YR_COLS:
    put_f(isw, f"{col}5", f"=Drivers!{col}9", bold=True)
    put_f(isw, f"{col}7", f"=Drivers!{col}19")
    put_f(isw, f"{col}8", f"=Drivers!{col}23")
    put_f(isw, f"{col}9", f"={col}7+{col}8", bold=True)
# net interest INCOME for GOOG (large net cash position)
int_h = {"C": (-3050, "[E] FY23 interest income net of expense ~+$3B (income, not expense)"),
         "D": (-4000, "[E] FY24 net interest income ~$4B"),
         "E": (-5000, "[E] FY25 net interest income ~$5B")}
for col in ("C", "D", "E"):
    put_input(isw, f"{col}10", int_h[col][0], int_h[col][1])
for col in ("F", "G", "H", "I", "J"):
    # Forward: interest cost rising w/ new debt issuance offset by interest income on cash
    put_f(isw, f"{col}10", f"=Drivers!{col}24/40")  # placeholder: small drag as capex consumes cash
for col in YR_COLS:
    put_f(isw, f"{col}11", f"={col}9-{col}8-{col}10")
    put_f(isw, f"{col}12", f"={col}11*Drivers!{col}25")
    put_f(isw, f"{col}13", f"={col}11-{col}12", bold=True)
    put_f(isw, f"{col}14", f"={col}13/{col}5", PCT)
isw.column_dimensions["A"].width = 38
for col in YR_COLS:
    isw.column_dimensions[col].width = 11

# BALANCE SHEET (summary)
bs = wb.create_sheet("Balance Sheet"); bs.sheet_view.showGridLines = False
bs["A1"] = "BALANCE SHEET (summary, $M)"
bs.merge_cells("A1:J1"); bs["A1"].font = BOLDW; bs["A1"].fill = NAVY_FILL; bs["A1"].alignment = C
yearrow(bs, 4)
lbl(bs, 5, "Cash + marketable securities")
lbl(bs, 6, "Other assets (plug)")
lbl(bs, 7, "Total debt")
lbl(bs, 8, "Shareholders' equity")
lbl(bs, 9, "Net cash (debt)", bold=True)
lbl(bs, 10, "Equity issuance/(repurchase) in year")
lbl(bs, 11, "Dividends + buybacks (cash out)")
put_input(bs, "C5", 110900, "[E] FY23 cash+sec ~$111B")
put_input(bs, "D5", 95700, "[E] FY24 ~$95.7B")
put_input(bs, "E5", 126800, "[E] FY25 / Q1 26 $126.8B (post $31B Q1 raise)")
put_input(bs, "C7", 13253, "[E] FY23 LT debt ~$13.3B")
put_input(bs, "D7", 10883, "[E] FY24 LT debt ~$10.9B")
put_input(bs, "E7", 46500, "FY25/Q1 26 LT debt $46.5B (after $31B Q1 senior notes)")
put_input(bs, "C8", 283379, "[E] FY23 equity ~$283.4B")
put_input(bs, "D8", 325084, "[E] FY24 equity ~$325B")
put_input(bs, "E8", 365000, "[E] FY25 equity ~$365B")
issue = {"F": (-84750, "FY26 ~$85B June equity raise (Berkshire pp + ATM + common)"),
         "G": (0, "[E] no further raise"), "H": (0, ""), "I": (0, ""), "J": (0, "")}
capret = {"F": (50000, "[E] FY26 buyback ~$45B + div ~$10B = $55B (declining as capex absorbs cash)"),
          "G": (60000, "[E] FY27 normalizing"), "H": (70000, "[E] FY28"), "I": (80000, "[E] FY29"), "J": (90000, "[E] FY30")}
for col in ("F", "G", "H", "I", "J"):
    put_input(bs, f"{col}10", issue[col][0], issue[col][1])
    put_input(bs, f"{col}11", capret[col][0], capret[col][1])
for i, col in enumerate(YR_COLS):
    prev = YR_COLS[i - 1] if i > 0 else None
    if col in ("F", "G", "H", "I", "J"):
        put_f(bs, f"{col}7", f"={prev}7")  # debt held flat (no further raise in base)
        put_f(bs, f"{col}8", f"={prev}8+'Income Statement'!{col}13-{col}11-{col}10")
        put_f(bs, f"{col}5", f"={prev}5+'Cash Flow'!{col}13-{col}10-{col}11")
        put_f(bs, f"{col}6", f"={col}7+{col}8-{col}5")
    else:
        put_f(bs, f"{col}6", f"={col}7+{col}8-{col}5")
    put_f(bs, f"{col}9", f"={col}5-{col}7", bold=True)  # net cash positive
bs.column_dimensions["A"].width = 38
for col in YR_COLS:
    bs.column_dimensions[col].width = 11

# CASH FLOW
cf = wb.create_sheet("Cash Flow"); cf.sheet_view.showGridLines = False
cf["A1"] = "CASH FLOW & FREE CASH FLOW ($M)"
cf.merge_cells("A1:J1"); cf["A1"].font = BOLDW; cf["A1"].fill = NAVY_FILL; cf["A1"].alignment = C
yearrow(cf, 4)
lbl(cf, 5, "EBIT")
lbl(cf, 6, "(-) Cash taxes on EBIT")
lbl(cf, 7, "NOPAT")
lbl(cf, 8, "(+) D&A")
lbl(cf, 9, "(-) Capex")
lbl(cf, 10, "(-) Increase in net working capital")
lbl(cf, 11, "Unlevered free cash flow", bold=True)
lbl(cf, 12, "(-) After-tax net interest")
lbl(cf, 13, "Levered free cash flow", bold=True)
lbl(cf, 14, "Unlevered FCF margin %")
for i, col in enumerate(YR_COLS):
    prev = YR_COLS[i - 1] if i > 0 else None
    put_f(cf, f"{col}5", f"='Income Statement'!{col}7")
    put_f(cf, f"{col}6", f"={col}5*Drivers!{col}25")
    put_f(cf, f"{col}7", f"={col}5-{col}6", bold=True)
    put_f(cf, f"{col}8", f"=Drivers!{col}23")
    put_f(cf, f"{col}9", f"=Drivers!{col}24")
    if prev:
        put_f(cf, f"{col}10", f"=(Drivers!{col}9-Drivers!{prev}9)*Drivers!{col}26")
    else:
        put_f(cf, f"{col}10", "=0")
    put_f(cf, f"{col}11", f"={col}7+{col}8-{col}9-{col}10", bold=True, fill=LG_FILL)
    put_f(cf, f"{col}12", f"='Income Statement'!{col}10*(1-Drivers!{col}25)")
    put_f(cf, f"{col}13", f"={col}11-{col}12", bold=True)
    put_f(cf, f"{col}14", f"={col}11/Drivers!{col}9", PCT)
cf.column_dimensions["A"].width = 38
for col in YR_COLS:
    cf.column_dimensions[col].width = 11

# DCF + SENSITIVITY
dcf = wb.create_sheet("DCF"); dcf.sheet_view.showGridLines = False
dcf["A1"] = "DISCOUNTED CASH FLOW VALUATION"
dcf.merge_cells("A1:H1"); dcf["A1"].font = BOLDW; dcf["A1"].fill = NAVY_FILL; dcf["A1"].alignment = C
PROJ = ["F", "G", "H", "I", "J"]
dcf["A4"] = "Projection year"; dcf["A4"].font = BOLDN; dcf["A4"].fill = LB_FILL
for i, col in enumerate(PROJ):
    cc = dcf.cell(row=4, column=2 + i, value=YR_HEADS[3 + i]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C
dcf["A5"] = "Period (n)"; dcf["A5"].font = BLACK
for i in range(5):
    put_f(dcf, f"{get_column_letter(2+i)}5", f"={i+1}")
dcf["A6"] = "Unlevered FCF"; dcf["A6"].font = BLACK
for i, col in enumerate(PROJ):
    put_f(dcf, f"{get_column_letter(2+i)}6", f"='Cash Flow'!{col}11")
dcf["A7"] = "Discount factor @ WACC"; dcf["A7"].font = BLACK
for i in range(5):
    cl = get_column_letter(2 + i); put_f(dcf, f"{cl}7", f"=1/(1+Drivers!C{R_WACC})^{cl}5", "0.000")
dcf["A8"] = "PV of unlevered FCF"; dcf["A8"].font = BOLDN
for i in range(5):
    cl = get_column_letter(2 + i); put_f(dcf, f"{cl}8", f"={cl}6*{cl}7")
br = [
    ("Sum of PV of explicit FCF (FY26-30)", "=SUM(B8:F8)", NUM),
    ("Terminal-year FCF (FY2030)", "=F6", NUM),
    ("Terminal value (Gordon)", f"=F6*(1+Drivers!C{R_TG})/(Drivers!C{R_WACC}-Drivers!C{R_TG})", NUM),
    ("PV of terminal value", "=C12*F7", NUM),
    ("Enterprise value", "=C10+C13", NUM),
    ("(+) Net cash (subtract net debt)", f"=Drivers!C{R_CASH}-Drivers!C{R_DEBT}", NUM),
    ("Equity value", "=C14+C15", NUM),
    ("Diluted shares (M)", f"=Drivers!C{R_SH}", NUM),
    ("Implied value per share ($)", "=C16/C17", USD2),
    ("Current share price ($)", f"=Drivers!C{R_PRICE}", USD2),
    ("Upside / (downside) to current", "=C18/C19-1", PCT),
    ("TV as % of enterprise value", "=C13/C14", PCT),
]
row = 10
for name, formula, fmt in br:
    dcf.cell(row=row, column=1, value=name).font = (BOLDN if "per share" in name or name == "Enterprise value" else BLACK)
    bold = name in ("Implied value per share ($)", "Enterprise value")
    put_f(dcf, f"C{row}", formula, fmt, bold=bold, fill=(MB_FILL if name == "Implied value per share ($)" else None)); row += 1

# Sensitivity
dcf["A22"] = "SENSITIVITY — Implied $/share: WACC (down) x Terminal growth (across)"
dcf.merge_cells("A22:H22"); dcf["A22"].font = BOLDW; dcf["A22"].fill = NAVY_FILL; dcf["A22"].alignment = C
g_axis = [0.020, 0.025, 0.030, 0.035, 0.040]
w_axis = [0.080, 0.085, 0.090, 0.095, 0.100]
dcf["B23"] = "Terminal g →"; dcf["B23"].font = ITAL
for j, g in enumerate(g_axis):
    cc = dcf.cell(row=23, column=3 + j, value=g); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT
dcf["A24"] = "WACC ↓"; dcf["A24"].font = ITAL
for r in range(5):
    cc = dcf.cell(row=24 + r, column=2, value=w_axis[r]); cc.font = BOLDN; cc.fill = LB_FILL; cc.alignment = C; cc.number_format = PCT
for r in range(5):
    wref = f"$B${24+r}"
    for cc in range(5):
        gref = f"{get_column_letter(3+cc)}$23"
        pv_expl = "+".join(f"${get_column_letter(2+k)}$6/(1+{wref})^{k+1}" for k in range(5))
        tv = f"$F$6*(1+{gref})/({wref}-{gref})"
        pv_tv = f"({tv})/(1+{wref})^5"
        equity = f"(({pv_expl})+({pv_tv})+Drivers!$C${R_CASH}-Drivers!$C${R_DEBT})"
        formula = f"={equity}/Drivers!$C${R_SH}"
        cell = dcf.cell(row=24 + r, column=3 + cc, value=formula)
        cell.number_format = USD2; cell.alignment = C
        if r == 2 and cc == 2:
            cell.font = Font(name="Times New Roman", size=11, bold=True); cell.fill = MB_FILL
        else:
            cell.font = BLACK
dcf["A30"] = "Center cell (WACC 9.0% / g 3.0%) = base case; cross-check vs C18 implied per share."
dcf.merge_cells("A30:H30"); dcf["A30"].font = ITAL
dcf.column_dimensions["A"].width = 38
for col in "BCDEFGH":
    dcf.column_dimensions[col].width = 12

# NOTES
nt = wb.create_sheet("Notes"); nt.sheet_view.showGridLines = False
nt["A1"] = "GOOG MODEL — METHODOLOGY & CAVEATS"
nt.merge_cells("A1:E1"); nt["A1"].font = BOLDW; nt["A1"].fill = NAVY_FILL; nt["A1"].alignment = C
notes = [
    ("Methodology", [
        "Integrated 3-statement model feeding unlevered DCF. Drivers tab drives everything.",
        "Revenue split into Google Services, Google Cloud, Other Bets — with explicit segment margins.",
        "Cloud is the marginal-mix story: 5%(23) → 14%(24) → 24%(25) → 30%(26E) → 36%(30E) op margin.",
        "Capex modeled at the announced $180-190B FY26 ramp (back-end loaded per Q1 $22.4B actual),",
        "tapering to ~$150-200B/yr through FY30 as build-out phase moderates.",
    ]),
    ("Base case stance", [
        "GOOG is the unusual case: AI capex is dilutive near-term, but Cloud margin expansion + scale",
        "more than offsets. Base case: Cloud margin reaches 30% in FY26 (vs Q1 26 actual 32.9%),",
        "tapers up to 36% by FY30. WACC ~9.0%, terminal g 3.0%.",
        "Headwinds priced: ~$85B June 2026 equity raise (Berkshire pp + ATM); DOJ search remedies",
        "(no Chrome divestiture, but default-payment restrictions); ~$20B TAC-to-Apple at risk.",
    ]),
    ("Data sources & caveats", [
        "Alphabet 10-K FY25 (Mar 2026), 10-Q Q1 CY26 (Apr 2026); FY23-24 historicals from filings.",
        "Cloud trajectory cross-checked across Q3 25 (+34%, 23.7%), Q4 25 (+48%), Q1 26 (+63%, 32.9%).",
        "MCP terminal connectors NOT configured; all data from SEC filings + reputable aggregators.",
        "Other Bets revenue largely Waymo; segment loss/revenue ratios are [E] derived.",
        "Net interest treatment: historicals show net interest INCOME (large cash); forwards model",
        "modest drag as cash depletes via capex.",
        "Research / decision-framing only — NOT investment advice.",
    ]),
]
row = 3
for title, lines in notes:
    c = nt.cell(row=row, column=1, value=title); c.font = BOLDN; c.fill = LB_FILL
    nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5); row += 1
    for ln in lines:
        cc = nt.cell(row=row, column=1, value=ln); cc.font = BLACK; cc.alignment = WRAP
        nt.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        nt.row_dimensions[row].height = 16; row += 1
    row += 1
for i, w in enumerate([26, 30, 30, 30, 30], start=1):
    nt.column_dimensions[get_column_letter(i)].width = w

wb._sheets = [wb["Drivers"], wb["Income Statement"], wb["Balance Sheet"], wb["Cash Flow"], wb["DCF"], wb["Notes"]]
wb.save(OUT)
print(f"Wrote {OUT}")
