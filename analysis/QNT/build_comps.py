"""
Build QNT-Comps-Analysis.xlsx — Quantum-computing pure-play comps + Honeywell
attribution memo, following the `comps-analysis` skill.

Comps cohort (June 4, 2026):
  - QNT (target — just IPO'd at $60, trading $68, EV ~$15.2B)
  - IONQ (NYSE: IONQ) — trapped-ion DIRECT comp
  - RGTI (Nasdaq: RGTI) — superconducting
  - QBTS (NYSE: QBTS) — quantum annealing (specialty)
  - HON (memo) — Honeywell parent owns ~48% post-IPO

Tabs: Inputs · Operating Metrics · Valuation · Tech & Growth · Notes
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "QNT-Comps-Analysis.xlsx"

NAVY = "17365D"
LB = "D9E1F2"
LG = "F2F2F2"
MB = "BDD7EE"

BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLD = Font(name="Times New Roman", size=11, bold=True)
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDN = Font(name="Times New Roman", size=11, bold=True, color="17365D")
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
LB_FILL = PatternFill("solid", fgColor=LB)
LG_FILL = PatternFill("solid", fgColor=LG)
MB_FILL = PatternFill("solid", fgColor=MB)
TGT_FILL = PatternFill("solid", fgColor="FFF2CC")

C = Alignment("center", "center")
L = Alignment("left", "center")
R = Alignment("right", "center")
WRAP = Alignment("left", "top", wrap_text=True)

NUM = "#,##0"
NUM1 = "#,##0.0"
PCT = "0.0%"
PCT0 = "0%"
MULT = '0.0"x"'
USD2 = '"$"#,##0.00'

wb = Workbook()
wb.remove(wb.active)


def hdr(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text)
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0)
    cell.font = BOLDW
    cell.fill = NAVY_FILL
    cell.alignment = C


# Tickers: Target (QNT) in B, peers C-E, memo F
TICKERS = ["QNT", "IONQ", "RGTI", "QBTS", "HON (memo)"]
COLS = ["B", "C", "D", "E", "F"]
ROLES = ["Target", "Trapped-ion direct comp", "Superconducting", "Annealing (specialty)", "Parent — 48% post-IPO"]


def ticker_header(ws, row):
    ws.cell(row=row, column=1, value="").font = ITAL
    for i, t in enumerate(TICKERS):
        c = ws.cell(row=row, column=2 + i, value=t)
        c.font = BOLDN
        c.fill = LB_FILL if i != 0 else TGT_FILL
        c.alignment = C
    # Role row
    for i, role in enumerate(ROLES):
        c = ws.cell(row=row + 1, column=2 + i, value=role)
        c.font = ITAL
        c.alignment = C


def setw(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def lab(ws, r, txt, bold=False, indent=0, italic=False, fill=None):
    c = ws.cell(row=r, column=1, value=txt)
    c.alignment = Alignment("left", "center", indent=indent)
    c.font = BOLD if bold else (ITAL if italic else BLACK)
    if fill is not None:
        c.fill = fill


def row_inputs(ws, r, values, fmt=NUM1, src_map=None):
    for i, v in enumerate(values):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = BLUE
        c.number_format = fmt
        c.alignment = R
        if i == 0:
            c.fill = TGT_FILL
        if src_map and i in src_map:
            c.comment = Comment(src_map[i], "Model")


def row_formula(ws, r, formulas, fmt=NUM1, bold=False, fill=None):
    for i, f in enumerate(formulas):
        c = ws.cell(row=r, column=2 + i, value=f)
        c.font = BOLD if bold else BLACK
        c.number_format = fmt
        c.alignment = R
        if fill is not None:
            c.fill = fill if i != 0 else TGT_FILL
        elif i == 0:
            c.fill = TGT_FILL


# ============================================================================
# INPUTS TAB
# ============================================================================
ws = wb.create_sheet("Inputs")
setw(ws, {"A": 40, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
hdr(ws, 1, 1, 6, "Quantum-computing comps — Inputs (as of June 4, 2026)")
ticker_header(ws, 3)

r = 6
lab(ws, r, "Share price ($, Jun 3-4 2026)", bold=True)
row_inputs(ws, r, [68.00, 68.21, 24.10, 27.09, 213.00], fmt=USD2,
           src_map={
               0: "QNT first-trade open Jun 4 2026 (IPO priced $60 prior day)",
               1: "IONQ close Jun 3 2026",
               2: "RGTI close Jun 3 2026",
               3: "QBTS close Jun 3 2026",
               4: "HON close Jun 3 2026 [E]"
           })
px = r

r += 1
lab(ws, r, "Shares outstanding (M, fully diluted)", bold=True)
row_inputs(ws, r, [253.9, 373.2, 325.0, 369.9, 650.0],
           src_map={
               0: "S-1: post-IPO 25.9M Class A + 228.0M Class B",
               1: "IONQ 10-Q Q1 2026",
               2: "RGTI Q1 2026 release",
               3: "QBTS Q1 2026 8-K",
               4: "HON share count [E]"
           })
sh = r

r += 1
lab(ws, r, "Market cap ($M)", bold=True)
for i, col in enumerate(COLS):
    c = ws.cell(row=r, column=2 + i, value=f"={col}{px}*{col}{sh}")
    c.font = BOLD
    c.number_format = NUM
    c.alignment = R
    c.fill = TGT_FILL if i == 0 else LB_FILL
mc = r

r += 1
lab(ws, r, "Cash & investments ($M)", indent=1)
row_inputs(ws, r, [2080, 3100, 569, 588, 11000], fmt=NUM,
           src_map={
               0: "$677M pre-IPO + $1,580M IPO net + ITM warrants $80M = ~$2,080M [E]",
               1: "IONQ Q1 2026 (3/31/26)",
               2: "RGTI Q1 2026",
               3: "QBTS Q1 2026",
               4: "HON cash [E]"
           })
cash = r

r += 1
lab(ws, r, "Debt ($M)", indent=1)
row_inputs(ws, r, [0, 0, 0, 0, 22000], fmt=NUM,
           src_map={4: "HON total debt [E]"})
debt = r

r += 1
lab(ws, r, "Enterprise value ($M)", bold=True)
for i, col in enumerate(COLS):
    c = ws.cell(row=r, column=2 + i, value=f"={col}{mc}-{col}{cash}+{col}{debt}")
    c.font = BOLD
    c.number_format = NUM
    c.alignment = R
    c.fill = TGT_FILL if i == 0 else LB_FILL
ev = r

r += 2
lab(ws, r, "Revenue", bold=True, fill=LG_FILL)
r += 1
lab(ws, r, "  FY 2025 revenue ($M)", indent=1)
row_inputs(ws, r, [30.9, 130.0, 7.1, 24.6, 38500],
           src_map={
               0: "S-1: FY25 revenue $30.9M (+34% YoY)",
               1: "IONQ FY2025 release",
               2: "RGTI FY2025 release",
               3: "QBTS FY2025 release",
               4: "HON FY2025 revenue [E]"
           })
fy25 = r

r += 1
lab(ws, r, "  Q1 2026 revenue ($M)", indent=1)
row_inputs(ws, r, [5.2, 64.7, 4.4, 2.9, 9600],
           src_map={
               0: "S-1 Q1 26: $5.2M (-13% vs Q1'25 driven by RIKEN runoff)",
               1: "IONQ Q1 2026 +755% YoY",
               2: "RGTI Q1 2026 +199% YoY",
               3: "QBTS Q1 2026 -81% YoY (system sale comp)",
               4: "HON Q1 2026 [E]"
           })
q1 = r

r += 1
lab(ws, r, "  TTM revenue ($M, est.)", bold=True)
row_inputs(ws, r, [32.6, 187.1, 10.0, 12.5, 38500], fmt=NUM1,
           src_map={
               0: "QNT TTM = FY25 + (Q1'26 - Q1'25): 30.9 + 5.2 - 3.5 = ~$32.6M [E]",
               1: "Peer research agent — IONQ TTM ~$187M",
               2: "RGTI TTM ~$10M",
               3: "QBTS TTM ~$12.5M",
               4: "HON TTM proxy [E]"
           })
ttm = r

r += 2
lab(ws, r, "Most recent quarter YoY growth", bold=True)
row_inputs(ws, r, [-0.13, 7.55, 1.99, -0.81, 0.04], fmt=PCT,
           src_map={
               0: "QNT Q1'26 $5.2M vs Q1'25 [E] ~$6M (RIKEN runoff)",
               1: "IONQ +755% from Q1'25 ~$7.6M to Q1'26 $64.7M",
               2: "RGTI +199% (~$1.5M → $4.4M)",
               3: "QBTS -81% (Q1'25 had $15M system sale)",
               4: "HON ~4% [E]"
           })
yoy = r

# ============================================================================
# OPERATING METRICS TAB
# ============================================================================
ws = wb.create_sheet("Operating Metrics")
setw(ws, {"A": 40, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
hdr(ws, 1, 1, 6, "Operating metrics & profitability")
ticker_header(ws, 3)

r = 6
lab(ws, r, "Latest-Q non-GAAP / Adj. EBITDA ($M)", bold=True)
row_inputs(ws, r, [-130, -96.8, -14.7, -18.4, 9200], fmt=NUM1,
           src_map={
               0: "QNT Q1'26 net loss $136.6M; SBC est ~$7M strip = ~$130M cash loss [E]",
               1: "IONQ Q1'26 Adj. EBITDA loss ($96.8M) (ex-SkyWater $85M)",
               2: "RGTI Q1'26 non-GAAP net loss ($14.7M)",
               3: "QBTS Q1'26 net loss ($18.4M)",
               4: "HON Q1'26 segment profit proxy [E]"
           })
adjeb = r

r += 1
lab(ws, r, "R&D ($M, latest year)", indent=1)
row_inputs(ws, r, [165.4, 254, 35.7, 88, 1500], fmt=NUM,
           src_map={
               0: "S-1: QNT FY25 R&D $165.4M",
               1: "IONQ FY25 R&D [E]",
               2: "RGTI FY25 R&D",
               3: "QBTS FY25 R&D [E]",
               4: "HON FY25 R&D [E]"
           })
rd = r

r += 1
lab(ws, r, "  R&D / revenue", italic=True)
for i, col in enumerate(COLS):
    c = ws.cell(row=r, column=2 + i, value=f"={col}{rd}/{col}{fy25}")
    c.font = BLACK
    c.number_format = PCT
    c.alignment = R
    if i == 0:
        c.fill = TGT_FILL
rd_pct = r

r += 1
lab(ws, r, "Cash burn (annualized FCF, $M)", indent=1)
row_inputs(ws, r, [-440, -350, -55, -80, 6500], fmt=NUM,
           src_map={
               0: "QNT FY25 FCF est -$220M extrapolated 2x = ~-$440M run-rate [E]",
               1: "IONQ extrapolated burn [E]",
               2: "RGTI extrapolated burn [E]",
               3: "QBTS extrapolated burn [E]",
               4: "HON +$6.5B FCF (mature)"
           })
burn = r

r += 1
lab(ws, r, "Years of cash runway (cash / burn)", bold=True, fill=LG_FILL)
for i, col in enumerate(COLS):
    formula = f"=IF({col}{burn}>=0,99,{col}Inputs.{cash}/(-{col}{burn}))"
    # use cross-sheet ref
    formula = f"=IF({col}{burn}>=0,99,Inputs!{col}{cash}/(-{col}{burn}))"
    c = ws.cell(row=r, column=2 + i, value=formula)
    c.font = BOLD
    c.number_format = NUM1
    c.alignment = R
    c.fill = TGT_FILL if i == 0 else LG_FILL

r += 2
lab(ws, r, "Customer concentration (top customer % rev)", bold=True)
row_inputs(ws, r, [0.60, 0.20, 0.30, 0.15, 0.10], fmt=PCT,
           src_map={
               0: "QNT: RIKEN 60% FY25 (7% Q1'26 — cliff)",
               1: "IONQ: govt/AWS large but spread [E]",
               2: "RGTI: govt + Novera concentrated [E]",
               3: "QBTS: 135+ commercial customers [E]",
               4: "HON: very diversified [E]"
           })

# ============================================================================
# VALUATION TAB
# ============================================================================
ws = wb.create_sheet("Valuation")
setw(ws, {"A": 40, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
hdr(ws, 1, 1, 6, "Valuation multiples")
ticker_header(ws, 3)

r = 6
lab(ws, r, "Enterprise value ($M)", bold=True)
for i, col in enumerate(COLS):
    c = ws.cell(row=r, column=2 + i, value=f"=Inputs!{col}{ev}")
    c.font = BOLD
    c.number_format = NUM
    c.alignment = R
    c.fill = TGT_FILL if i == 0 else LB_FILL

r += 1
lab(ws, r, "EV / FY2025 revenue", bold=True, fill=LG_FILL)
for i, col in enumerate(COLS):
    c = ws.cell(row=r, column=2 + i, value=f"=Inputs!{col}{ev}/Inputs!{col}{fy25}")
    c.font = BOLD
    c.number_format = MULT
    c.alignment = R
    c.fill = TGT_FILL if i == 0 else LG_FILL
ev_fy25 = r

r += 1
lab(ws, r, "EV / TTM revenue", bold=True, fill=LG_FILL)
for i, col in enumerate(COLS):
    c = ws.cell(row=r, column=2 + i, value=f"=Inputs!{col}{ev}/Inputs!{col}{ttm}")
    c.font = BOLD
    c.number_format = MULT
    c.alignment = R
    c.fill = TGT_FILL if i == 0 else LG_FILL
ev_ttm = r

r += 2
lab(ws, r, "Peer-set stats (excl. target QNT and memo HON)", bold=True, fill=LG_FILL)
r += 1
lab(ws, r, "  Median EV / FY25 revenue", italic=True)
c = ws.cell(row=r, column=2, value=f"=MEDIAN(C{ev_fy25}:E{ev_fy25})")
c.font = BOLD
c.number_format = MULT
c.alignment = R
c.fill = MB_FILL
med_fy25 = r
r += 1
lab(ws, r, "  Median EV / TTM revenue", italic=True)
c = ws.cell(row=r, column=2, value=f"=MEDIAN(C{ev_ttm}:E{ev_ttm})")
c.font = BOLD
c.number_format = MULT
c.alignment = R
c.fill = MB_FILL
med_ttm = r

r += 2
lab(ws, r, "QNT premium / (discount) to peer median", bold=True, fill=LG_FILL)
r += 1
lab(ws, r, "  on FY2025 revenue", italic=True)
c = ws.cell(row=r, column=2, value=f"=B{ev_fy25}/B{med_fy25}-1")
c.font = BOLD
c.number_format = PCT
c.alignment = R
c.fill = TGT_FILL
r += 1
lab(ws, r, "  on TTM revenue", italic=True)
c = ws.cell(row=r, column=2, value=f"=B{ev_ttm}/B{med_ttm}-1")
c.font = BOLD
c.number_format = PCT
c.alignment = R
c.fill = TGT_FILL

r += 3
lab(ws, r, "Football field — implied price per QNT share", bold=True, fill=LG_FILL)
r += 1
lab(ws, r, "  Low (RGTI multiple ~726x; we hold rev constant)", italic=True)
# (peer_mult * QNT_rev + cash) / QNT_shares
c = ws.cell(row=r, column=2,
            value=f"=(D{ev_fy25}*Inputs!B{fy25}+Inputs!B{cash})/Inputs!B{sh}")
c.font = BLACK
c.number_format = USD2
c.alignment = R
c.fill = TGT_FILL
r += 1
lab(ws, r, "  Mid (median peer multiple)", italic=True)
c = ws.cell(row=r, column=2,
            value=f"=(B{med_fy25}*Inputs!B{fy25}+Inputs!B{cash})/Inputs!B{sh}")
c.font = BOLD
c.number_format = USD2
c.alignment = R
c.fill = MB_FILL
r += 1
lab(ws, r, "  High (apply QBTS multiple ~754x)", italic=True)
c = ws.cell(row=r, column=2,
            value=f"=(E{ev_fy25}*Inputs!B{fy25}+Inputs!B{cash})/Inputs!B{sh}")
c.font = BLACK
c.number_format = USD2
c.alignment = R
c.fill = TGT_FILL
r += 1
lab(ws, r, "  IONQ-anchored (~120x — lowest peer)", italic=True)
c = ws.cell(row=r, column=2,
            value=f"=(C{ev_fy25}*Inputs!B{fy25}+Inputs!B{cash})/Inputs!B{sh}")
c.font = BLACK
c.number_format = USD2
c.alignment = R
c.fill = TGT_FILL
r += 1
lab(ws, r, "  Current trading price", bold=True)
c = ws.cell(row=r, column=2, value=f"=Inputs!B{px}")
c.font = BOLD
c.number_format = USD2
c.alignment = R
c.fill = LB_FILL

# ============================================================================
# TECH & GROWTH TAB
# ============================================================================
ws = wb.create_sheet("Tech & Growth")
setw(ws, {"A": 40, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24})
hdr(ws, 1, 1, 6, "Technology, customers & growth")
ticker_header(ws, 3)

r = 6
specs = [
    ("Modality", ["Trapped-ion (QCCD)", "Trapped-ion", "Superconducting", "Quantum annealing", "Industrial conglomerate"]),
    ("Latest commercial system", ["Helios (Nov 2025)", "Forte / Tempo (2026)", "Cepheus-108Q (Apr 2026)", "Advantage2 (May 2025)", "n/a"]),
    ("Qubit count (latest)", ["98 (Helios)", "64+ AQ; Tempo 5th gen", "108", "4,400+", "n/a"]),
    ("2-qubit gate fidelity", ["99.921% (Helios)", "99.99% (lab)", "99.1% on 108Q; 99.9% prototype", "n/a (annealer)", "n/a"]),
    ("Quantum Volume record", ["2^25 (33M, Sept'25)", "n/d", "n/d", "n/a (annealer)", "n/a"]),
    ("Next milestone", ["Sol 2027 (192q)", "Tempo / 200k qubits", "99.5% fidelity 2H'26", "Quantum Circuits gate-model", "QNT IPO proceeds"]),
    ("FTQC roadmap year", ["2029 (Apollo)", "2028 functional test", "~3 yrs to QC advantage", "QCI gate-model acquisition", "n/a"]),
    ("Marquee customers", ["RIKEN, JPMC, BMW,\nAmgen, MSFT, SoftBank", "AWS, AstraZeneca,\nNVIDIA, AFRL, ORNL", "Novera ($5.7M),\nDOE, AFRL, Japan QPU", "Mastercard, Lockheed,\n135+ commercial", "Aerospace, defense,\nautomation, energy"]),
    ("Govt revenue exposure", ["DARPA QBI Stage B,\n$100M CHIPS LOI", "DARPA HARQ + others", "DOE, AFRL", "Various govt", "Major defense"]),
    ("Modality maturity assessment", ["High fidelity; scaling\nthe key risk", "Production-scale\nrevenue", "Roadmap uncertain", "Annealer NOT gate-model;\nadding via M&A", "Diversified"]),
]
for label, vals in specs:
    lab(ws, r, label, bold=True)
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = BLACK
        c.alignment = Alignment("left", "top", wrap_text=True)
        if i == 0:
            c.fill = TGT_FILL
        ws.row_dimensions[r].height = 30 if "\n" in v else 18
    r += 1

# ============================================================================
# NOTES TAB
# ============================================================================
ws = wb.create_sheet("Notes")
setw(ws, {"A": 110})
hdr(ws, 1, 1, 1, "Comps notes")

notes = [
    "",
    "PEER SET RATIONALE",
    "  Comps statistics computed over IONQ, RGTI, QBTS (the 3 public quantum pure-plays).",
    "  HON is a MEMO column — Honeywell still owns ~48% post-IPO, so worth attributing",
    "  QNT's marginal value to HON's market cap; not a valuation comp.",
    "",
    "VALUATION INSIGHT",
    "  • At $68, QNT trades at ~470-490x EV/FY25 revenue.",
    "  • Peer median EV/FY25 (IONQ, RGTI, QBTS): ~530x — QNT is roughly IN LINE.",
    "  • Peer median EV/TTM: ~720x — QNT trades at a DISCOUNT here.",
    "  • The 'rich vs peers' narrative is wrong. The 'rich vs intrinsic DCF' narrative",
    "    is correct (see QNT-Model.xlsx; +175% premium to traditional DCF).",
    "",
    "FOOTBALL FIELD",
    "  • Range $54 (IONQ-anchored) to ~$150 (QBTS-anchored).",
    "  • Current price ($68) is in the low-mid range.",
    "  • Caveat: TTM-based valuation distorted by Q1 26 hyper-growth at IONQ (+755%).",
    "",
    "CRITICAL HONEYWELL ATTRIBUTION",
    "  • HON owns ~48% post-IPO of QNT equity worth ~$7.5B at $68/share.",
    "  • HON market cap ~$138B → QNT stake = ~5.4% of HON value.",
    "  • Not transformative; modest +ve catalyst for HON shareholders.",
    "  • HON also retains a 'Transaction Committee' veto on QNT M&A >$10M.",
    "",
    "DATA GAPS / [E]",
    "  • IONQ Q1'25 revenue back-solved from +755% YoY statement.",
    "  • Cash burn run-rates extrapolated from latest-Q net loss (excl. SBC where known).",
    "  • HON figures rounded estimates; not core to QNT valuation.",
    "",
    "REFRESH",
    "  Once full Q2'26 prints (Aug/Sep 2026) are out for IONQ/RGTI/QBTS, refresh TTM",
    "  revenue and growth rows. QNT's own Q2 print is the first major catalyst.",
]
for i, txt in enumerate(notes):
    cell = ws.cell(row=2 + i, column=1, value=txt)
    cell.font = BOLD if (txt and not txt.startswith(" ")) else BLACK
    cell.alignment = WRAP

# Tab order: Inputs · Operating Metrics · Valuation · Tech & Growth · Notes
for name in ["Inputs", "Operating Metrics", "Valuation", "Tech & Growth", "Notes"]:
    wb.move_sheet(name, offset=-99)

wb.save(OUT)
print(f"Wrote {OUT}")
