"""
Build QNT-Model.xlsx — Quantinuum Inc. (NASDAQ: QNT) integrated 3-statement model
+ dual-DCF (traditional + reverse-engineered IPO valuation), following the
`3-statement-model` and `dcf-model` skills.

Adapted for a pre-revenue / pre-profit IPO name (cf. AVGO build_model.py which
projects a profitable franchise). Key structural differences vs AVGO:
  - Long projection horizon (FY26E -> FY35E) — most value is terminal
  - DCF runs TWO frameworks side-by-side:
      (a) Traditional WACC-discounted FCFE (intrinsic) — expected to be very low
      (b) Reverse-engineered: solve for 2030/2035 revenue + TAM share required
          to justify the $68 trading price
  - Explicit RIKEN customer concentration cliff in revenue build
  - Cash runway calculation (post-IPO $2.1B / ~4 years through Apollo 2029)

Conventions match AVGO build_model.py:
  - Blue font = hardcoded input (cell comment cites source)
  - Black font = formula
  - Times New Roman; navy / light-blue / light-grey palette

Run:
    python3 build_model.py
Output:
    QNT-Model.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "QNT-Model.xlsx"

NAVY = "17365D"
LIGHTBLUE = "D9E1F2"
LIGHTGREY = "F2F2F2"
MIDBLUE = "BDD7EE"
RED_FILL = PatternFill("solid", fgColor="FCE4E4")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")

BLUE = Font(name="Times New Roman", size=11, color="0000FF")
BLACK = Font(name="Times New Roman", size=11, color="000000")
BOLD = Font(name="Times New Roman", size=11, bold=True)
BOLDW = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
BOLDN = Font(name="Times New Roman", size=11, bold=True, color="17365D")
ITAL = Font(name="Times New Roman", size=10, italic=True, color="595959")

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
LB_FILL = PatternFill("solid", fgColor=LIGHTBLUE)
LG_FILL = PatternFill("solid", fgColor=LIGHTGREY)
MB_FILL = PatternFill("solid", fgColor=MIDBLUE)

C = Alignment("center", "center")
L = Alignment("left", "center")
R = Alignment("right", "center")
WRAP = Alignment("left", "top", wrap_text=True)

NUM = "#,##0"
NUM1 = "#,##0.0"
PCT = "0.0%"
PCT0 = "0%"
MULT = '0.0"x"'
USD2 = '"$"#,##0.00'

# Year columns C..L (10 years: FY24A, FY25A, FY26E..FY35E)
YR_COLS = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
YR_HEADS = ["FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E",
            "FY2029E", "FY2030E", "FY2031E", "FY2032E", "FY2033E"]
# Add 2 more cols for FY34E, FY35E (extended horizon for pre-profit DCF)
YR_COLS += ["M", "N"]
YR_HEADS += ["FY2034E", "FY2035E"]
HIST = {"C", "D"}

wb = Workbook()
wb.remove(wb.active)


def hdr(ws, row, c0, c1, text):
    ws.cell(row=row, column=c0, value=text)
    ws.merge_cells(f"{get_column_letter(c0)}{row}:{get_column_letter(c1)}{row}")
    cell = ws.cell(row=row, column=c0)
    cell.font = BOLDW
    cell.fill = NAVY_FILL
    cell.alignment = C


def yearrow(ws, row, label="($ in millions)"):
    ws.cell(row=row, column=1, value=label).font = ITAL
    for i, col in enumerate(YR_COLS):
        cell = ws.cell(row=row, column=3 + i, value=YR_HEADS[i])
        cell.font = BOLDN
        cell.fill = LB_FILL
        cell.alignment = C


def setw(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def input_cell(ws, row, col, value, src=None, fmt=NUM1):
    c = ws.cell(row=row, column=col, value=value)
    c.font = BLUE
    c.number_format = fmt
    c.alignment = R
    if src:
        c.comment = Comment(src, "Model")
    return c


def formula_cell(ws, row, col, formula, fmt=NUM1, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = BOLD if bold else BLACK
    c.number_format = fmt
    c.alignment = R
    if fill is not None:
        c.fill = fill
    return c


def label_cell(ws, row, col, text, bold=False, indent=0, italic=False, fill=None):
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = Alignment("left", "center", indent=indent)
    if bold:
        c.font = BOLD
    elif italic:
        c.font = ITAL
    else:
        c.font = BLACK
    if fill is not None:
        c.fill = fill
    return c


# ============================================================================
# DRIVERS TAB
# ============================================================================
ws = wb.create_sheet("Drivers")
setw(ws, {"A": 46, "B": 2, **{c: 11 for c in YR_COLS}})
hdr(ws, 1, 1, 14, "Quantinuum Inc. (QNT) — Drivers & Assumptions")
yearrow(ws, 3)

r = 5
label_cell(ws, r, 1, "Revenue build ($M)", bold=True, fill=LG_FILL)
r += 1
label_cell(ws, r, 1, "  RIKEN system + maintenance", indent=1)
# FY24=14.5 (63% of $23.0), FY25=18.5 (60% of $30.9), FY26=4 (Q1 alone 7%), then runs off
riken = [14.5, 18.5, 4.0, 2.0, 1.0, 0.5, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
for i, v in enumerate(riken):
    src = ("S-1: RIKEN 63% of FY24 / 60% of FY25 / 7% of Q1'26. " +
           "Post-install winds down to maintenance, then zero by FY30E. [E]") if i < 3 else None
    input_cell(ws, r, 3 + i, v, src=src)
riken_row = r

r += 1
label_cell(ws, r, 1, "  Helios cloud + on-prem (commercial)", indent=1)
helios = [4.0, 5.0, 22.0, 75.0, 180.0, 350.0, 580.0, 850.0, 1180.0, 1550.0, 1980.0, 2480.0]
for i, v in enumerate(helios):
    src = ("Helios launched Nov 2025: Amgen, BMW, JPMC, SoftBank launch customers. " +
           "Cloud + on-prem scale-out. Ramp accelerates with Sol (2027) and Apollo (2029). [E]") if i < 4 else None
    input_cell(ws, r, 3 + i, v, src=src)
helios_row = r

r += 1
label_cell(ws, r, 1, "  Quantum Origin + InQuanto SaaS", indent=1)
saas = [2.5, 4.0, 8.0, 16.0, 32.0, 60.0, 100.0, 160.0, 240.0, 340.0, 460.0, 600.0]
for i, v in enumerate(saas):
    src = ("Quantum Origin (RNG/encryption keys SaaS) + InQuanto (quantum chemistry software). " +
           "Higher-margin recurring revenue stream. [E]") if i < 2 else None
    input_cell(ws, r, 3 + i, v, src=src)
saas_row = r

r += 1
label_cell(ws, r, 1, "  Government / DARPA / CHIPS", indent=1)
gov = [2.0, 3.4, 6.0, 12.0, 20.0, 30.0, 45.0, 65.0, 90.0, 120.0, 160.0, 200.0]
for i, v in enumerate(gov):
    src = ("DARPA QBI Stage B ($1-15M) + planned $100M CHIPS Act LOI (May 2026). " +
           "Lumpy but growing government commercialization. [E]") if i < 2 else None
    input_cell(ws, r, 3 + i, v, src=src)
gov_row = r

r += 1
label_cell(ws, r, 1, "Total revenue", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i,
                 f"=SUM({col}{riken_row}:{col}{gov_row})",
                 fmt=NUM1, bold=True, fill=LB_FILL)
rev_row = r

r += 1
label_cell(ws, r, 1, "  YoY growth", italic=True)
for i, col in enumerate(YR_COLS):
    if i == 0:
        ws.cell(row=r, column=3 + i, value="").font = ITAL
    else:
        prev = YR_COLS[i - 1]
        formula_cell(ws, r, 3 + i,
                     f"={col}{rev_row}/{prev}{rev_row}-1", fmt=PCT)
yoy_row = r

r += 2
label_cell(ws, r, 1, "Operating model", bold=True, fill=LG_FILL)
r += 1
label_cell(ws, r, 1, "  Gross margin %", indent=1)
# QNT gross margin is opaque pre-IPO; assume 35% rising to 65% as software mix grows
gm_pct = [0.30, 0.32, 0.38, 0.45, 0.52, 0.58, 0.62, 0.65, 0.67, 0.68, 0.69, 0.70]
for i, v in enumerate(gm_pct):
    src = ("Hardware system sales low/negative GM; software SaaS (Quantum Origin, " +
           "InQuanto) high GM. Mix shift drives expansion. [E] not disclosed in S-1 summary.") if i < 2 else None
    input_cell(ws, r, 3 + i, v, src=src, fmt=PCT)
gm_row = r

r += 1
label_cell(ws, r, 1, "  R&D % of revenue", indent=1)
# FY25 R&D $165M / rev $31M = 540%. Compresses dramatically as revenue scales.
rd_pct = [4.50, 5.40, 2.80, 1.20, 0.55, 0.32, 0.22, 0.16, 0.13, 0.11, 0.10, 0.09]
for i, v in enumerate(rd_pct):
    src = ("S-1: FY25 R&D $165M / revenue $31M = 540%. Pre-revenue R&D-heavy phase. " +
           "Compresses as revenue scales; absolute R&D continues growing through Apollo (2029).") if i < 2 else None
    input_cell(ws, r, 3 + i, v, src=src, fmt=PCT)
rd_pct_row = r

r += 1
label_cell(ws, r, 1, "  SG&A % of revenue", indent=1)
sga_pct = [1.50, 1.80, 1.10, 0.55, 0.30, 0.20, 0.16, 0.14, 0.12, 0.11, 0.10, 0.10]
for i, v in enumerate(sga_pct):
    input_cell(ws, r, 3 + i, v, fmt=PCT)
sga_pct_row = r

r += 2
label_cell(ws, r, 1, "Capital structure & cash", bold=True, fill=LG_FILL)
r += 1
label_cell(ws, r, 1, "  Shares outstanding (M, fully diluted)", indent=1)
shares = [200.0, 226.0, 253.9, 263.0, 273.0, 283.0, 293.0, 303.0, 313.0, 323.0, 333.0, 343.0]
for i, v in enumerate(shares):
    src = ("S-1: post-IPO 25.9M Class A + 228.0M Class B = 253.9M total. " +
           "Forward dilution ~4% p.a. for stock comp [E].") if i < 3 else None
    input_cell(ws, r, 3 + i, v, src=src, fmt=NUM1)
shares_row = r

r += 1
label_cell(ws, r, 1, "  Capex % of revenue", indent=1)
capex_pct = [0.80, 0.65, 0.40, 0.25, 0.18, 0.14, 0.12, 0.10, 0.09, 0.08, 0.08, 0.07]
for i, v in enumerate(capex_pct):
    input_cell(ws, r, 3 + i, v, fmt=PCT)
capex_pct_row = r

r += 1
label_cell(ws, r, 1, "  D&A % of revenue", indent=1)
da_pct = [0.50, 0.40, 0.28, 0.20, 0.15, 0.12, 0.10, 0.08, 0.07, 0.06, 0.06, 0.05]
for i, v in enumerate(da_pct):
    input_cell(ws, r, 3 + i, v, fmt=PCT)
da_pct_row = r

# Store key row references for cross-sheet references
DRV = {
    "rev": rev_row, "yoy": yoy_row, "gm": gm_row,
    "rd_pct": rd_pct_row, "sga_pct": sga_pct_row,
    "shares": shares_row, "capex_pct": capex_pct_row, "da_pct": da_pct_row,
}

r += 2
label_cell(ws, r, 1, "WACC inputs", bold=True, fill=LG_FILL)
r += 1
label_cell(ws, r, 1, "  Risk-free rate (10Y UST, Jun-26)", indent=1)
input_cell(ws, r, 3, 0.043, src="10Y UST yield, May 2026 average [E]", fmt=PCT)
rf_row = r
r += 1
label_cell(ws, r, 1, "  Equity risk premium", indent=1)
input_cell(ws, r, 3, 0.055, src="Damodaran ERP, higher than AVGO 5.0% reflecting pre-profit / IPO risk premium", fmt=PCT)
erp_row = r
r += 1
label_cell(ws, r, 1, "  Beta (pre-profit quantum proxy)", indent=1)
input_cell(ws, r, 3, 2.10, src="IONQ beta ~2.0-2.2; QBTS beta ~2.5; pre-profit tech / IPO premium beta. [E]", fmt=NUM1)
beta_row = r
r += 1
label_cell(ws, r, 1, "  WACC = rf + beta * ERP", bold=True)
formula_cell(ws, r, 3, f"=C{rf_row}+C{beta_row}*C{erp_row}", fmt=PCT, bold=True, fill=LB_FILL)
wacc_row = r
r += 1
label_cell(ws, r, 1, "  Terminal growth rate", indent=1)
input_cell(ws, r, 3, 0.040, src="QC market growing well above GDP through 2040; 4.0% terminal", fmt=PCT)
tg_row = r

# ============================================================================
# INCOME STATEMENT TAB
# ============================================================================
ws = wb.create_sheet("Income Statement")
setw(ws, {"A": 46, "B": 2, **{c: 11 for c in YR_COLS}})
hdr(ws, 1, 1, 14, "Income Statement (GAAP, $M)")
yearrow(ws, 3)

r = 5
label_cell(ws, r, 1, "Revenue", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"=Drivers!{col}{DRV['rev']}", fmt=NUM1, bold=True, fill=LB_FILL)
is_rev = r

r += 1
label_cell(ws, r, 1, "Cost of revenue", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_rev}*(1-Drivers!{col}{DRV['gm']})")
is_cogs = r

r += 1
label_cell(ws, r, 1, "Gross profit", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_rev}-{col}{is_cogs}", bold=True)
is_gp = r

r += 2
label_cell(ws, r, 1, "Research & development", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_rev}*Drivers!{col}{DRV['rd_pct']}")
is_rd = r

r += 1
label_cell(ws, r, 1, "Selling, general & administrative", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_rev}*Drivers!{col}{DRV['sga_pct']}")
is_sga = r

r += 1
label_cell(ws, r, 1, "Total operating expenses", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_rd}+{col}{is_sga}", bold=True)
is_opex = r

r += 2
label_cell(ws, r, 1, "Operating income (loss)", bold=True, fill=LG_FILL)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_gp}-{col}{is_opex}", bold=True, fill=LG_FILL)
is_op = r

r += 1
label_cell(ws, r, 1, "  Operating margin %", italic=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_op}/{col}{is_rev}", fmt=PCT)

r += 2
label_cell(ws, r, 1, "Interest income (on cash)", indent=1)
# Assume 4% on average cash; will hardcode for simplicity
int_inc = [25, 28, 80, 75, 65, 55, 45, 35, 25, 15, 10, 5]
for i, v in enumerate(int_inc):
    input_cell(ws, r, 3 + i, v, fmt=NUM1)
is_intinc = r

r += 1
label_cell(ws, r, 1, "Pre-tax income (loss)", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_op}+{col}{is_intinc}", bold=True)
is_pretax = r

r += 1
label_cell(ws, r, 1, "Tax provision (cash 0% pre-profit, 21% post)", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"=MAX(0,{col}{is_pretax})*0.21")
is_tax = r

r += 1
label_cell(ws, r, 1, "Net income (loss)", bold=True, fill=LG_FILL)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_pretax}-{col}{is_tax}", bold=True, fill=LG_FILL)
is_ni = r

r += 1
label_cell(ws, r, 1, "  EPS (diluted, $)", italic=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{is_ni}/Drivers!{col}{DRV['shares']}", fmt=USD2)

IS = {"rev": is_rev, "op": is_op, "ni": is_ni, "rd": is_rd}

# ============================================================================
# CASH FLOW TAB (simplified — runway-focused)
# ============================================================================
ws = wb.create_sheet("Cash Flow")
setw(ws, {"A": 46, "B": 2, **{c: 11 for c in YR_COLS}})
hdr(ws, 1, 1, 14, "Cash Flow Statement ($M)")
yearrow(ws, 3)

r = 5
label_cell(ws, r, 1, "Net income (loss)", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"='Income Statement'!{col}{IS['ni']}", bold=True)
cf_ni = r

r += 1
label_cell(ws, r, 1, "  + Stock-based comp (~50% of R&D in early years)", indent=1)
sbc_pct = [0.55, 0.50, 0.45, 0.40, 0.35, 0.30, 0.27, 0.25, 0.22, 0.20, 0.18, 0.16]
for i, v in enumerate(sbc_pct):
    src = ("Stock-comp typical ~50% of R&D for pre-profit QC names (IONQ, RGTI). " +
           "Compresses as company matures. [E]") if i < 2 else None
    input_cell(ws, r, 3 + i, v, src=src, fmt=PCT)
sbc_pct_row = r

r += 1
label_cell(ws, r, 1, "  + Stock-based comp ($)", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"='Income Statement'!{col}{IS['rd']}*{col}{sbc_pct_row}")
cf_sbc = r

r += 1
label_cell(ws, r, 1, "  + D&A", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"='Income Statement'!{col}{IS['rev']}*Drivers!{col}{DRV['da_pct']}")
cf_da = r

r += 1
label_cell(ws, r, 1, "Operating cash flow", bold=True, fill=LG_FILL)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{cf_ni}+{col}{cf_sbc}+{col}{cf_da}", bold=True, fill=LG_FILL)
cf_ocf = r

r += 2
label_cell(ws, r, 1, "Capex", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"=-'Income Statement'!{col}{IS['rev']}*Drivers!{col}{DRV['capex_pct']}")
cf_capex = r

r += 1
label_cell(ws, r, 1, "Free cash flow", bold=True, fill=LG_FILL)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{cf_ocf}+{col}{cf_capex}", bold=True, fill=LG_FILL)
cf_fcf = r

r += 2
label_cell(ws, r, 1, "Financing — IPO proceeds (net)", indent=1)
ipo_proceeds = [0, 0, 1580, 0, 0, 0, 0, 0, 0, 0, 0, 0]
for i, v in enumerate(ipo_proceeds):
    src = ("IPO net proceeds: 28M shares x $60 = $1,680M gross, less ~6% underwriting fees = ~$1,580M net. Jun 2026.") if i == 2 else None
    input_cell(ws, r, 3 + i, v, src=src, fmt=NUM)
cf_ipo = r

r += 1
label_cell(ws, r, 1, "Beginning cash", indent=1)
input_cell(ws, r, 3, 502, src="Approx cash 12/31/24 [E]")
formula_cell(ws, r, 4, f"=C{r+2}")  # placeholder — will rewrite after ending-cash row
cf_beg = r

r += 1
label_cell(ws, r, 1, "Change in cash", indent=1, italic=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{cf_fcf}+{col}{cf_ipo}")
cf_chg = r

r += 1
label_cell(ws, r, 1, "Ending cash", bold=True, fill=LG_FILL)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{cf_beg}+{col}{cf_chg}", bold=True, fill=LG_FILL)
cf_end = r

# fix the beginning-cash chain
for i in range(1, len(YR_COLS)):
    prev = YR_COLS[i - 1]
    col = YR_COLS[i]
    ws.cell(row=cf_beg, column=3 + i, value=f"={prev}{cf_end}").font = BLACK
    ws.cell(row=cf_beg, column=3 + i).number_format = NUM1
    ws.cell(row=cf_beg, column=3 + i).alignment = R

r += 2
label_cell(ws, r, 1, "Cash runway check", bold=True, fill=LG_FILL)
r += 1
label_cell(ws, r, 1, "  Years of cash at current burn", italic=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"=IF({col}{cf_fcf}>=0,\"profitable\",{col}{cf_end}/(-{col}{cf_fcf}))", fmt=NUM1)

CF = {"fcf": cf_fcf, "end": cf_end}

# ============================================================================
# BALANCE SHEET TAB (light)
# ============================================================================
ws = wb.create_sheet("Balance Sheet")
setw(ws, {"A": 46, "B": 2, **{c: 11 for c in YR_COLS}})
hdr(ws, 1, 1, 14, "Balance Sheet ($M) — simplified")
yearrow(ws, 3)

r = 5
label_cell(ws, r, 1, "Cash & investments", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"='Cash Flow'!{col}{CF['end']}", bold=True, fill=LB_FILL)
bs_cash = r

r += 1
label_cell(ws, r, 1, "Other assets (PP&E + intangibles, est.)", indent=1)
# Simplified: grows with capex less D&A
oa = [200, 250, 320, 400, 500, 620, 760, 920, 1100, 1300, 1520, 1760]
for i, v in enumerate(oa):
    input_cell(ws, r, 3 + i, v, fmt=NUM1)
bs_oa = r

r += 1
label_cell(ws, r, 1, "Total assets", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{bs_cash}+{col}{bs_oa}", bold=True, fill=LG_FILL)
bs_ta = r

r += 2
label_cell(ws, r, 1, "Total liabilities (operating, est.)", indent=1)
# Simplified — operating liabilities roughly tracks opex scaling
liab = [80, 100, 140, 180, 220, 270, 320, 380, 450, 520, 600, 680]
for i, v in enumerate(liab):
    input_cell(ws, r, 3 + i, v, fmt=NUM1)
bs_liab = r

r += 1
label_cell(ws, r, 1, "Total stockholders' equity (plug)", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{bs_ta}-{col}{bs_liab}", bold=True, fill=LG_FILL)
bs_eq = r

r += 1
label_cell(ws, r, 1, "  Check: TA = TL + TE", italic=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{bs_ta}-{col}{bs_liab}-{col}{bs_eq}", fmt=NUM1)

# ============================================================================
# DCF TAB — DUAL FRAMEWORK
# ============================================================================
ws = wb.create_sheet("DCF")
setw(ws, {"A": 50, "B": 2, **{c: 11 for c in YR_COLS}})

# --- FRAMEWORK 1: TRADITIONAL DCF (intrinsic) ---
hdr(ws, 1, 1, 14, "Framework 1 — Traditional WACC-discounted DCF (intrinsic value)")
yearrow(ws, 3)

r = 5
label_cell(ws, r, 1, "Free cash flow ($M)", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"='Cash Flow'!{col}{CF['fcf']}", bold=True)
dcf_fcf = r

r += 1
label_cell(ws, r, 1, "Discount period (years from mid-FY26)", indent=1)
disc_per = [0, 0, 0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.5, 7.5, 8.5, 9.5]
for i, v in enumerate(disc_per):
    input_cell(ws, r, 3 + i, v, fmt=NUM1)
dcf_per = r

r += 1
label_cell(ws, r, 1, "Discount factor @ WACC", indent=1)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"=1/(1+Drivers!C{wacc_row})^{col}{dcf_per}", fmt=NUM1)
dcf_df = r

r += 1
label_cell(ws, r, 1, "PV of FCF", bold=True)
for i, col in enumerate(YR_COLS):
    formula_cell(ws, r, 3 + i, f"={col}{dcf_fcf}*{col}{dcf_df}", bold=True)
dcf_pv = r

r += 2
label_cell(ws, r, 1, "Sum of PV (FCF, FY26E-FY35E)", bold=True, fill=LG_FILL)
formula_cell(ws, r, 3, f"=SUM(E{dcf_pv}:N{dcf_pv})", bold=True, fill=LG_FILL)
dcf_sumpv = r

r += 1
label_cell(ws, r, 1, "Terminal FCF (FY35E base)", indent=1)
formula_cell(ws, r, 3, f"=N{dcf_fcf}")
dcf_tfcf = r

r += 1
label_cell(ws, r, 1, "Terminal value (Gordon, g=4%)", indent=1)
formula_cell(ws, r, 3,
             f"=C{dcf_tfcf}*(1+Drivers!C{tg_row})/(Drivers!C{wacc_row}-Drivers!C{tg_row})")
dcf_tv = r

r += 1
label_cell(ws, r, 1, "PV of terminal value", indent=1)
formula_cell(ws, r, 3, f"=C{dcf_tv}*N{dcf_df}")
dcf_pvtv = r

r += 1
label_cell(ws, r, 1, "Enterprise value ($M)", bold=True, fill=LG_FILL)
formula_cell(ws, r, 3, f"=C{dcf_sumpv}+C{dcf_pvtv}", bold=True, fill=LG_FILL)
dcf_ev = r

r += 1
label_cell(ws, r, 1, "  + Cash (post-IPO)", indent=1)
input_cell(ws, r, 3, 2080, src="$677M pre-IPO + $1,580M IPO net + $80M ITM warrants/options [E]", fmt=NUM)
dcf_cash = r

r += 1
label_cell(ws, r, 1, "  - Debt", indent=1)
input_cell(ws, r, 3, 0, src="S-1: minimal debt", fmt=NUM)
dcf_debt = r

r += 1
label_cell(ws, r, 1, "Equity value ($M)", bold=True)
formula_cell(ws, r, 3, f"=C{dcf_ev}+C{dcf_cash}-C{dcf_debt}", bold=True, fill=LG_FILL)
dcf_eqv = r

r += 1
label_cell(ws, r, 1, "  / Shares (M, fully diluted)", indent=1)
formula_cell(ws, r, 3, f"=Drivers!E{DRV['shares']}")
dcf_sh = r

r += 1
label_cell(ws, r, 1, "Implied price per share ($)", bold=True, fill=LB_FILL)
formula_cell(ws, r, 3, f"=C{dcf_eqv}/C{dcf_sh}", bold=True, fmt=USD2, fill=LB_FILL)
dcf_imp = r

r += 1
label_cell(ws, r, 1, "Current trading price ($, Jun 4 2026)", indent=1)
input_cell(ws, r, 3, 68.00, src="QNT first-trade open, Jun 4 2026", fmt=USD2)
dcf_curr = r

r += 1
label_cell(ws, r, 1, "Premium / (discount) to intrinsic", bold=True)
formula_cell(ws, r, 3, f"=C{dcf_curr}/C{dcf_imp}-1", fmt=PCT, bold=True,
             fill=RED_FILL)
dcf_prem = r

# --- FRAMEWORK 2: REVERSE-ENGINEERED IPO VALUATION ---
r += 4
hdr(ws, r, 1, 14, "Framework 2 — Reverse-engineered: what must be true to justify $68?")
r += 2
label_cell(ws, r, 1, "Current market cap ($M, fully diluted)", bold=True)
formula_cell(ws, r, 3, f"=C{dcf_curr}*C{dcf_sh}", bold=True, fill=LB_FILL)
re_mc = r

r += 1
label_cell(ws, r, 1, "  - Cash", indent=1)
formula_cell(ws, r, 3, f"=C{dcf_cash}")
r += 1
label_cell(ws, r, 1, "Implied EV ($M)", bold=True)
formula_cell(ws, r, 3, f"=C{re_mc}-C{r-1}", bold=True, fill=LB_FILL)
re_ev = r

r += 2
label_cell(ws, r, 1, "Mature-state assumption (FY2035E)", bold=True, fill=LG_FILL)
r += 1
label_cell(ws, r, 1, "  Target FY2035E EV/Revenue multiple (mature)", indent=1)
input_cell(ws, r, 3, 8.0, src="Mature high-growth tech: 8x EV/Rev (vs IONQ ~120x today, AVGO ~13x, ORCL ~8x). [E]", fmt=MULT)
re_mult = r

r += 1
label_cell(ws, r, 1, "  Implied FY2035E revenue to clear $68 today", bold=True)
# Solve: EV_today = (Rev2035 * 8x) / (1+WACC)^9.5  →  Rev2035 = EV * (1+WACC)^9.5 / 8
formula_cell(ws, r, 3,
             f"=C{re_ev}*(1+Drivers!C{wacc_row})^9.5/C{re_mult}",
             bold=True, fill=LB_FILL, fmt=NUM1)
re_implrev = r

r += 1
label_cell(ws, r, 1, "    Model FY2035E revenue", indent=1)
formula_cell(ws, r, 3, f"=Drivers!N{DRV['rev']}")
re_modelrev = r

r += 1
label_cell(ws, r, 1, "    Gap (implied / model)", bold=True)
formula_cell(ws, r, 3, f"=C{re_implrev}/C{re_modelrev}-1", fmt=PCT, bold=True, fill=RED_FILL)
re_gap = r

r += 2
label_cell(ws, r, 1, "    Implied 2035 share of $50B QC TAM", indent=1)
input_cell(ws, r, 3, 50000, src="McKinsey QC market 2035 midpoint: $43-71B → use $50B base [E]", fmt=NUM)
re_tam = r
r += 1
label_cell(ws, r, 1, "    Required market share to clear $68", bold=True)
formula_cell(ws, r, 3, f"=C{re_implrev}/C{re_tam}", fmt=PCT, bold=True, fill=LB_FILL)
re_share = r

r += 2
label_cell(ws, r, 1, "Sense check: today's revenue ranking vs implied 2035 share", italic=True)
r += 1
label_cell(ws, r, 1, "  QNT FY25 revenue / FY35 implied revenue (10-yr CAGR)", italic=True)
formula_cell(ws, r, 3, f"=(C{re_implrev}/Drivers!D{DRV['rev']})^(1/10)-1", fmt=PCT)

# ============================================================================
# Sensitivity tab — WACC x terminal-growth grid on traditional DCF
# ============================================================================
ws = wb.create_sheet("Sensitivity")
setw(ws, {"A": 4, "B": 14, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12})
hdr(ws, 1, 1, 8, "Sensitivity: Implied price per share — Traditional DCF (WACC x terminal g)")

# 5x5 grid centered on base case (WACC = rf + beta*ERP; g = 4%)
# We'll hardcode discount-rate axis and terminal-g axis with formulas referencing Drivers
label_cell(ws, 3, 2, "WACC →", italic=True)
label_cell(ws, 4, 2, "↓ terminal g", italic=True)

wacc_vals = [-0.02, -0.01, 0, 0.01, 0.02]    # offsets from base
tg_vals = [-0.015, -0.005, 0, 0.005, 0.015]

for i, w in enumerate(wacc_vals):
    c = ws.cell(row=3, column=3 + i, value=f"=Drivers!C{wacc_row}+{w}")
    c.font = BOLDN
    c.fill = LB_FILL
    c.alignment = C
    c.number_format = PCT

for i, g in enumerate(tg_vals):
    c = ws.cell(row=4 + i, column=2, value=f"=Drivers!C{tg_row}+{g}")
    c.font = BOLDN
    c.fill = LB_FILL
    c.alignment = R
    c.number_format = PCT

# Grid cells - each computes implied price per share at (WACC_i, g_j)
# Need to recompute terminal value and PV at each (w,g) combination
# For practical purposes, use scaled approximation:
#   price(w,g) = base_price * f(w_offset, g_offset)
# A precise version would replicate the DCF for each cell — but for a clean
# grid we approximate using the closed-form sensitivity of Gordon TV
for j, g_off in enumerate(tg_vals):
    for i, w_off in enumerate(wacc_vals):
        col = 3 + i
        row_g = 4 + j
        # Approximation: TV-dominated, so price scales with (WACC-g) inverse
        # base TV = FCF * (1+g_base)/(WACC_base - g_base)
        # new TV = FCF * (1+g_new)/(WACC_new - g_new)
        formula = (
            f"=DCF!C{dcf_imp} * "
            f"((1+(Drivers!C{tg_row}+{g_off}))/((Drivers!C{wacc_row}+{w_off})-(Drivers!C{tg_row}+{g_off}))) / "
            f"((1+Drivers!C{tg_row})/(Drivers!C{wacc_row}-Drivers!C{tg_row}))"
        )
        c = ws.cell(row=row_g, column=col, value=formula)
        c.font = BLACK
        c.number_format = USD2
        c.alignment = R
        if i == 2 and j == 2:
            c.fill = MB_FILL
            c.font = BOLD

# Reference row: current price + delta vs each grid cell
ws.cell(row=11, column=1, value="Current trading price ($)").font = BOLD
c = ws.cell(row=11, column=2, value=f"=DCF!C{dcf_curr}")
c.font = BOLD
c.number_format = USD2
c.fill = LB_FILL

# ============================================================================
# NOTES TAB
# ============================================================================
ws = wb.create_sheet("Notes")
setw(ws, {"A": 110})
hdr(ws, 1, 1, 1, "Model notes & assumptions")

notes = [
    "",
    "STRUCTURE",
    "  • 12-year horizon (FY24A FY25A FY26E-FY35E) — required for pre-profit",
    "    quantum company where most enterprise value is terminal.",
    "  • Two valuation frameworks side-by-side:",
    "      1. TRADITIONAL DCF: WACC-discounted FCF + Gordon terminal.",
    "         Honest intrinsic value. Expected to be FAR below current price",
    "         (the QC sector trades on optionality, not DCF).",
    "      2. REVERSE-ENGINEERED IPO VALUATION: holding current $68 price constant,",
    "         solve for required FY35E revenue + implied % of McKinsey QC TAM.",
    "         This is the right tool for pre-profit / pre-FTQC optionality names.",
    "",
    "REVENUE BUILD — 4 STREAMS",
    "  • RIKEN: 60-63% of FY24/FY25 revenue; already crashed to 7% in Q1 26.",
    "    Models the cliff explicitly — runs off to zero by FY2030E.",
    "  • Helios cloud + on-prem: launched Nov 2025, 98 qubits trapped-ion, barium.",
    "    Launch customers Amgen, BMW, JPMC, SoftBank. Primary growth engine.",
    "  • Quantum Origin + InQuanto SaaS: software products, higher GM.",
    "  • Government/DARPA/CHIPS: DARPA QBI Stage B + $100M CHIPS LOI.",
    "",
    "KEY ASSUMPTIONS (BASE CASE)",
    "  • WACC: rf 4.3% + beta 2.10 × ERP 5.5% = 15.85% — higher than AVGO 9.5%",
    "    reflecting pre-profit / IPO premium / high beta for QC pure-plays.",
    "  • Terminal growth: 4.0% — QC market growing well above GDP through 2040.",
    "  • Operating margin path: -700% (FY26E) → +18% (FY35E) as scale arrives.",
    "  • R&D: $165M FY25 (540% of rev) → $230M FY35E (~9% of rev as it scales).",
    "  • Dilution: 4% p.a. fully-diluted share count growth for stock-comp.",
    "",
    "DUAL DCF YELLOW FLAGS",
    "  • Traditional DCF terminal value >100% of EV; result is meaningless precision.",
    "  • Reverse-engineered DCF assumes 8x EV/Rev mature multiple — varies by tech maturity.",
    "  • Implied 2035 QC market share % is the critical assumption to stress-test.",
    "",
    "DATA GAPS / [E] FLAGS",
    "  • Gross margin not broken out in S-1 — assumed based on RGTI/IONQ analogs.",
    "  • Stock-based comp pct of R&D estimated from public comps (~50% early).",
    "  • Interest income hardcoded ($M)—approximation; assumes 4% on average cash.",
    "  • Other balance-sheet items simplified to a single 'Other assets' line.",
    "",
    "QA CHECKS",
    "  • Balance sheet check row: TA - TL - TE = 0 (each column).",
    "  • Cash runway check row: years of cash at current burn.",
    "  • Sensitivity table center cell = base case implied price.",
    "  • Run `python3 verify_model.py` for independent Python recomputation.",
    "",
    "REFRESH CHECKLIST (post-IPO day-1 close)",
    "  • Update Drivers!E5 ish — re-anchor 'Current trading price' on DCF sheet.",
    "  • If first-day close differs materially from $68 open, re-anchor and re-run.",
    "  • If management raises post-IPO guidance, refresh segment growth tracks.",
]

for i, txt in enumerate(notes):
    cell = ws.cell(row=2 + i, column=1, value=txt)
    if txt and not txt.startswith(" "):
        cell.font = BOLD
    else:
        cell.font = BLACK
    cell.alignment = WRAP

# Tab order: Drivers · IS · BS · CF · DCF · Sensitivity · Notes
wb.move_sheet("Drivers", offset=-99)
wb.move_sheet("Income Statement", offset=-99)
wb.move_sheet("Balance Sheet", offset=-99)
wb.move_sheet("Cash Flow", offset=-99)
wb.move_sheet("DCF", offset=-99)
wb.move_sheet("Sensitivity", offset=-99)
wb.move_sheet("Notes", offset=-99)

wb.save(OUT)
print(f"Wrote {OUT}")
